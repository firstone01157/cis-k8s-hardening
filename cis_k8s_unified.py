#!/usr/bin/env python3
# pyright: strict
"""
CIS Kubernetes Benchmark - Unified Interactive Runner (Enhanced & Optimized)
คู่มือระบบตรวจสอบ CIS Kubernetes - ตัวรันโปรแกรมแบบรวมและโต้ตอบ (เพิ่มประสิทธิภาพ)

Features / ฟีเจอร์:
- Smart Kubeconfig Detection / การตรวจจับ Kubeconfig อัจฉริยะ
- Auto Backup System / ระบบสำรองข้อมูลอัตโนมัติ
- Oracle-Style Detailed XML/Text Reporting / รายงานแบบ Oracle
- Robust CSV Generation / สร้าง CSV ที่เชื่อถือได้
- Component-Based Summary / สรุปตามส่วนประกอบ
"""

import os
import sys
import shutil
import subprocess
import json
import csv
import time
import argparse
import tempfile
import requests
import difflib
import re
import shlex
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from types import ModuleType
from typing import List, Optional, Tuple, Dict, Any, Union, cast, Set, TypedDict
import glob
import socket
import urllib3
import modules.golden_configs as golden_configs
from modules.cis_level2_remediation import Level2Remediator
from modules.verification_utils import (
    get_component_pids as fetch_component_pids,
    verify_with_retry,
    wait_for_api_ready,
)

yaml: Optional[ModuleType] = None
try:
    import yaml as yaml_module
    yaml = yaml_module
except ImportError:
    # Fallback will be handled in classes
    yaml = None

openpyxl: Optional[ModuleType] = None
Font = PatternFill = Alignment = None
get_column_letter = None
openpyxl_available = False
try:
    import openpyxl as openpyxl_module
    from openpyxl.utils import get_column_letter
    openpyxl = openpyxl_module
    openpyxl_available = True
except ImportError:
    openpyxl_available = False

# --- Constants / ค่าคงที่ ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "cis_config.json")
LOG_FILE = os.path.join(BASE_DIR, "cis_runner.log")
REPORT_DIR = os.path.join(BASE_DIR, "reports")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
HISTORY_DIR = os.path.join(BASE_DIR, "history")

# Max workers for parallel execution / จำนวนผู้ปฏิบัติงานสูงสุดสำหรับการทำงานแบบขนาน
MAX_WORKERS = 8
SCRIPT_TIMEOUT = 60  # seconds / วินาที
REQUIRED_TOOLS = ["kubectl", "jq", "grep", "sed", "awk"]  # Required dependencies / การพึ่งพาที่จำเป็น



class Colors:
    """Terminal color codes / รหัสสีของเทอร์มินัล"""
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    WHITE = '\033[97m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class CISResult(TypedDict):
    id: str
    role: str
    level: str
    status: str
    duration: float
    reason: str
    fix_hint: str
    cmds: List[str]
    output: str
    path: str
    component: str

ResultTuple = Tuple[str, str, Optional[str], List[str]]

def _excel_column_letter(col_idx: int) -> str:
    """Return column letter even if openpyxl helper is unavailable."""
    if callable(get_column_letter):
        return get_column_letter(col_idx)

    if col_idx <= 0:
        return "A"

    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def save_yaml_robust(path: str, data: str, expected_flags: Optional[List[str]] = None) -> None:
    """Write YAML atomically, flush to disk, and validate critical flags."""
    tmp_path = f"{path}.tmp"
    content = str(data)
    expected = [flag for flag in (expected_flags or []) if flag]

    try:
        with open(tmp_path, "w", encoding="utf-8") as tmp_file:
            tmp_file.write(content)
            tmp_file.flush()
            os.fsync(tmp_file.fileno())
        os.chmod(tmp_path, 0o600)
        os.replace(tmp_path, path)
        os.chmod(path, 0o600)

        with open(path, "r", encoding="utf-8") as written_file:
            written_content = written_file.read()

        if written_content != content:
            raise IOError(f"Atomic write mismatch for {path}")

        missing_flags = [flag for flag in expected if flag not in written_content]
        if missing_flags:
            raise IOError(f"Missing critical flags after write: {missing_flags}")

    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


class YAMLSafeModifier:
    """
    Safely modify Kubernetes manifest YAML files.
    จัดการแก้ไขไฟล์ YAML ของ Kubernetes Manifest อย่างปลอดภัย
    """
    
    def __init__(self, file_path: str, dry_run: bool = False):
        """
        Initialize YAML modifier / เริ่มต้นตัวแก้ไข YAML
        :param file_path: Path to the YAML file / เส้นทางไปยังไฟล์ YAML
        :param dry_run: If True, don't save changes / หากเป็น True จะไม่บันทึกการเปลี่ยนแปลง
        """
        self.file_path = Path(file_path)
        self.dry_run = dry_run
        self.data: Dict[str, Any] = {}
        self.original_content: Optional[str] = None
        self._load_yaml()
    
    def _load_yaml(self) -> None:
        """Load YAML content from file / โหลดเนื้อหา YAML จากไฟล์"""
        if not self.file_path.exists():
            return
        try:
            with open(self.file_path, 'r') as f:
                self.original_content = f.read()

            if yaml is None:
                self.data = {}
                return

            loaded_raw = yaml.safe_load(self.original_content)
            if isinstance(loaded_raw, dict):
                self.data = cast(Dict[str, Any], loaded_raw)
            else:
                self.data = {}
        except Exception:
            # Fallback to empty dict if parsing fails / ใช้ dict ว่างหากการแยกวิเคราะห์ล้มเหลว
            self.data = {}
    
    def apply_modifications(self, modifications: Dict[str, Any], mod_type: str = "string") -> str:
        """
        Apply modifications to YAML data and return as string.
        ใช้การแก้ไขกับข้อมูล YAML และส่งกลับเป็นสตริง
        """
        if not self.data:
            return self.original_content or ""
            
        if 'flags' in modifications:
            self._update_command_flags(modifications['flags'], mod_type)
        if yaml is None:
            return self.original_content or ""

        return yaml.dump(
            self.data,
            default_flow_style=False,
            sort_keys=False,
            allow_unicode=True,
            width=float("inf")
        )

    def _update_command_flags(self, flags: List[str], mod_type: str = "string") -> None:
        """
        Update command-line flags in the container spec.
        อัปเดต flag คำสั่งในส่วนของ container spec
        """
        try:
            data = self.data
            spec_raw = data.get('spec', {})
            if not isinstance(spec_raw, dict):
                return
            spec = cast(Dict[str, Any], spec_raw)
            containers = spec.get('containers', [])
            if containers:
                command = containers[0].get('command', [])
                for flag_entry in flags:
                    if '=' in flag_entry:
                        flag_key, flag_value = flag_entry.split('=', 1)
                        self._update_flag_in_list(command, flag_key, flag_value, mod_type)
                    else:
                        self._update_flag_in_list(command, flag_entry, None, mod_type)
                containers[0]['command'] = command
        except Exception:
            pass

    def _update_flag_in_list(self, command_list: List[str], flag: str, value: Optional[str], mod_type: str = "string") -> None:
        """
        Update or append a flag in a list of command arguments.
        อัปเดตหรือเพิ่ม flag ในรายการอาร์กิวเมนต์คำสั่ง
        """
        # List of flags that should be treated as CSV lists / รายการ flag ที่ควรจัดการเป็นรายการ CSV
        CSV_FLAGS = [
            "--authorization-mode", 
            "--enable-admission-plugins", 
            "--disable-admission-plugins", 
            "--tls-cipher-suites"
        ]
        
        flag_index = -1
        for i, item in enumerate(command_list):
            if item.startswith(flag + '=') or item == flag:
                flag_index = i
                break
        
        if flag_index >= 0:
            if value is not None:
                if mod_type == "csv" or flag in CSV_FLAGS:
                    # Smart Append: Don't overwrite, append to CSV list if not present
                    # การเพิ่มแบบอัจฉริยะ: ไม่เขียนทับ แต่เพิ่มลงในรายการ CSV หากยังไม่มี
                    current_item = command_list[flag_index]
                    if '=' in current_item:
                        current_value = current_item.split('=', 1)[1]
                        # Split by comma and strip whitespace / แยกด้วยเครื่องหมายจุลภาคและตัดช่องว่าง
                        existing_values = [v.strip() for v in current_value.split(',') if v.strip()]
                        
                        if value not in existing_values:
                            existing_values.append(value)
                            command_list[flag_index] = f"{flag}={','.join(existing_values)}"
                    else:
                        # Flag exists but has no value, just set it / มี flag อยู่แต่ไม่มีค่า ให้ตั้งค่าเลย
                        command_list[flag_index] = f"{flag}={value}"
                else:
                    # Normal overwrite for non-CSV flags / การเขียนทับปกติสำหรับ flag ที่ไม่ใช่ CSV
                    command_list[flag_index] = f"{flag}={value}"
        else:
            if value is not None:
                command_list.append(f"{flag}={value}")
            else:
                command_list.append(flag)


class AtomicRemediationManager:
    """
    Manages atomic file writes and automatic rollback.
    จัดการการเขียนไฟล์แบบ atomic และการย้อนกลับอัตโนมัติ
    """
    
    def __init__(self, backup_dir: str = "/var/backups/cis-remediation"):
        """
        Initialize remediation manager / เริ่มต้นตัวจัดการการแก้ไข
        :param backup_dir: Directory for backups / ไดเรกทอรีสำหรับสำรองข้อมูล
        """
        self.backup_dir = Path(backup_dir)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.health_check_url = "https://127.0.0.1:6443/healthz"
        self.health_check_timeout = 300  # Increased to 5 minutes for slow hardware / เพิ่มเป็น 5 นาทีสำหรับฮาร์ดแวร์ที่ช้า
        self.health_check_interval = 5  # Check every 5 seconds / ตรวจสอบทุก 5 วินาที
        self.api_settle_time = 15
        
        # Smart SSL Verification / การตรวจสอบ SSL อัจฉริยะ
        self.ca_cert = "/etc/kubernetes/pki/ca.crt"
        if os.path.exists(self.ca_cert):
            self.ssl_verify = self.ca_cert
        else:
            self.ssl_verify = False
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    def create_backup(self, filepath: Union[str, Path]) -> Tuple[bool, str]:
        """
        Create a backup of the specified file / สร้างการสำรองข้อมูลของไฟล์ที่ระบุ
        """
        try:
            filepath = Path(filepath)
            if not filepath.exists():
                return False, ""
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.backup_dir / f"{filepath.name}.bak_{timestamp}"
            shutil.copy2(filepath, backup_path)
            return True, str(backup_path)
        except Exception:
            return False, ""
    
    def update_manifest_safely(self, filepath: Union[str, Path], modifications: Dict[str, Any], mod_type: str = "string") -> Tuple[bool, str]:
        """
        Update a manifest file atomically with backup / อัปเดตไฟล์ manifest แบบ atomic พร้อมการสำรองข้อมูล
        """
        filepath = Path(filepath)
        try:
            with open(filepath, 'r') as f:
                original_content = f.read()
            
            backup_success, backup_path = self.create_backup(str(filepath))
            if not backup_success:
                return False, "Failed to create backup"
            
            modifier = YAMLSafeModifier(str(filepath))
            modified_content = modifier.apply_modifications(modifications, mod_type)
            
            if modified_content == original_content:
                return True, "No changes needed"
            
            # Use temporary file for atomic write / ใช้ไฟล์ชั่วคราวสำหรับการเขียนแบบ atomic
            temp_fd, temp_path = tempfile.mkstemp(dir=filepath.parent, prefix=filepath.stem + '_', suffix='.tmp')
            try:
                with os.fdopen(temp_fd, 'w') as f:
                    f.write(modified_content)
                    f.flush()
                    os.fsync(f.fileno())
                # Atomic replacement / การแทนที่แบบ atomic
                os.replace(temp_path, filepath)
                return True, f"Updated successfully. Backup: {backup_path}"
            except Exception as e:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                raise e
        except Exception as e:
            return False, str(e)
    
    def wait_for_cluster_healthy(self) -> Tuple[bool, str]:
        """
        Wait for the Kubernetes API to become healthy / รอให้ Kubernetes API กลับมาใช้งานได้ปกติ
        Robust check for local remediation to prevent false-positive rollbacks on slow hardware.
        """
        print(f"{Colors.CYAN}[*] Waiting for API Server recovery (Timeout: {self.health_check_timeout}s)...{Colors.ENDC}")
        
        # Suppress insecure request warnings for this local check
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        start_time = time.time()
        while (time.time() - start_time) < self.health_check_timeout:
            elapsed = int(time.time() - start_time)
            try:
                # FORCE DISABLE VERIFY for localhost to avoid SSLError (hostname mismatch)
                # Accept 200 (OK) or 401 (Unauthorized) as success
                response = requests.get(
                    self.health_check_url, 
                    verify=False, 
                    timeout=5
                )
                
                if response.status_code in (200, 401):
                    print(f"{Colors.GREEN}    [✓] API Server is ALIVE (Status: {response.status_code}) after {elapsed}s. Settling...{Colors.ENDC}")
                    if self.api_settle_time > 0:
                        time.sleep(self.api_settle_time)
                    return True, f"Cluster is healthy (Status: {response.status_code})"
                elif response.status_code in (500, 503):
                    print(f"{Colors.YELLOW}    [*] API initializing... (Status: {response.status_code}, {elapsed}s/{self.health_check_timeout}s){Colors.ENDC}")
                else:
                    print(f"{Colors.YELLOW}    [*] API responding with status {response.status_code}... ({elapsed}s/{self.health_check_timeout}s){Colors.ENDC}")
                
            except requests.exceptions.ConnectionError:
                print(f"{Colors.YELLOW}    [*] API restarting... (Connection Refused, {elapsed}s/{self.health_check_timeout}s){Colors.ENDC}")
            except Exception as e:
                print(f"{Colors.YELLOW}    [*] Waiting for API... ({str(e)[:50]}, {elapsed}s/{self.health_check_timeout}s){Colors.ENDC}")
            
            time.sleep(self.health_check_interval)
            
        return False, f"Cluster health check timed out after {self.health_check_timeout}s"
    
    def rollback(self, filepath: str, backup_path: str) -> bool:
        """
        Restore file from backup / กู้คืนไฟล์จากการสำรองข้อมูล
        """
        try:
            shutil.copy2(backup_path, filepath)
            return True
        except Exception:
            return False

    def verify_remediation(self, check_id: str, audit_script_path: str) -> Tuple[bool, str]:
        """
        Verify remediation by running the audit script / ตรวจสอบการแก้ไขโดยการรันสคริปต์ตรวจสอบ
        """
        try:
            result = subprocess.run(["bash", audit_script_path], capture_output=True, text=True, timeout=30)
            if result.returncode == 0 and "PASS" in result.stdout:
                return True, "Audit passed"
            return False, f"Audit failed: {result.stdout}"
        except Exception as e:
            return False, str(e)


class CISUnifiedRunner:
    """
    Main CIS Kubernetes Benchmark Runner
    ตัวรันหลักของ CIS Kubernetes Benchmark
    """
    
    def __init__(self, verbose: int = 0, config_path: Optional[str] = None):
        """Initialize runner with configuration / เริ่มต้นตัวรันพร้อมการตั้งค่า"""
        self.base_dir = BASE_DIR
        self.config_file = config_path if config_path else CONFIG_FILE
        self.config_data: Dict[str, Any] = {}
        self.log_file = LOG_FILE
        self.report_dir = REPORT_DIR
        self.backup_dir = BACKUP_DIR
        self.history_dir = HISTORY_DIR
        
        # Execution settings / การตั้งค่าการดำเนิน
        self.verbose = verbose
        self.skip_manual = False
        self.script_timeout = SCRIPT_TIMEOUT
        
        # Results tracking / การติดตามผลลัพธ์
        self.results: List[CISResult] = []
        self.audit_results: Dict[str, Dict[str, Any]] = {}  # Track audit results by check ID for targeted remediation
        self.stats: Dict[str, Dict[str, int]] = {}
        self.stop_requested = False
        self.health_status = "UNKNOWN"
        self.manual_pending_items: List[Dict[str, Any]] = []  # Separate list for manual checks (NOT failures)
        self.current_level = "all"
        self.level_check_totals = {"L1": 0, "L2": 0}
        self.total_level_one_two_checks = 0
        self._aggressive_remediation_confirmed = False
        
        # Timestamp and directories / แสตมป์เวลาและไดเรกทอรี
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.date_dir = os.path.join(self.report_dir, datetime.now().strftime("%Y-%m-%d"))
        self.current_report_dir = None
        
        # Create required directories / สร้างไดเรกทอรีที่จำเป็น
        for directory in [self.report_dir, self.backup_dir, self.history_dir, self.date_dir]:
            os.makedirs(directory, exist_ok=True)
        
        # Initialize Atomic Remediation Manager
        self.atomic_manager = AtomicRemediationManager(backup_dir=self.backup_dir)
        
        # Smart SSL Verification for Runner / การตรวจสอบ SSL อัจฉริยะสำหรับ Runner
        self.ca_cert = "/etc/kubernetes/pki/ca.crt"
        self.health_check_url = "https://127.0.0.1:6443/healthz"
        
        if os.path.exists(self.ca_cert):
            self.ssl_verify = self.ca_cert
        else:
            self.ssl_verify = False
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        self.load_config()
        # Run pre-flight checks before proceeding
        self.run_preflight_checks()
        # Ensure audit log directory exists / ตรวจสอบให้แน่ใจว่ามีไดเรกทอรีบันทึกการตรวจสอบ
        self.ensure_audit_log_dir()

    def show_banner(self):
        """Display application banner / แสดงแบนเนอร์แอปพลิเคชัน"""
        banner = f"""
{Colors.CYAN}{'='*70}
  CIS Kubernetes Benchmark - Unified Interactive Runner
  ตัวรันเครื่องมือตรวจสอบ CIS Kubernetes แบบรวมและโต้ตอบ
  Version: 1.0 (Optimized)
  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*70}{Colors.ENDC}
"""
        print(banner)

    def load_config(self):
        """
        Load configuration from JSON file.
        โหลดการตั้งค่าจากไฟล์ JSON
        """
        self.excluded_rules: Dict[str, Any] = {}
        self.component_mapping: Dict[str, List[str]] = {}
        self.remediation_config: Dict[str, Any] = {}
        self.remediation_global_config: Dict[str, Any] = {}
        self.remediation_checks_config: Dict[str, Dict[str, Any]] = {}
        self.remediation_env_vars: Dict[str, Any] = {}
        self.variables: Dict[str, Any] = {}  # Store variables section for reference resolution / เก็บส่วนตัวแปรสำหรับการแก้ไขการอ้างอิง
        
        # Initialize API timeout settings with defaults / เริ่มต้นการตั้งค่า API timeout ด้วยค่าเริ่มต้น
        self.api_check_interval = 5  # seconds / วินาที
        self.api_max_retries = 60    # 60 * 5 = 300 seconds total (5 minutes) / รวม 5 นาที
        self.api_settle_time = 15    # settle time after API becomes ready (seconds) / เวลารอให้ API นิ่งหลังจากพร้อมใช้งาน
        self.wait_for_api_enabled = True
        
        if not os.path.exists(self.config_file):
            print(f"{Colors.YELLOW}[!] Config not found. Using defaults.{Colors.ENDC}")
            return
        
        try:
            with open(self.config_file, 'r') as f:
                config_raw = json.load(f)

            if not isinstance(config_raw, dict):
                raise ValueError("Configuration root must be an object")

            config = cast(Dict[str, Any], config_raw)
            self.config_data = config
            self.excluded_rules = cast(Dict[str, Any], config.get("excluded_rules", {}))
            self.component_mapping = cast(Dict[str, List[str]], config.get("component_mapping", {}))
            self.variables = cast(Dict[str, Any], config.get("variables", {}))  # Load variables for reference resolution / โหลดตัวแปรสำหรับการแก้ไขการอ้างอิง
            
            # Load remediation configuration / โหลดการตั้งค่าการแก้ไข
            self.remediation_config = config.get("remediation_config", {})
            self.remediation_global_config = self.remediation_config.get("global", {})
            checks_config_raw = self.remediation_config.get("checks")
            normalized_checks: Dict[str, Dict[str, Any]] = {}
            if isinstance(checks_config_raw, dict):
                checks_config = cast(Dict[str, Any], checks_config_raw)
                for check_id, check_cfg in checks_config.items():
                    if not isinstance(check_cfg, dict):
                        continue
                    normalized_checks[str(check_id)] = check_cfg
                self.remediation_checks_config = normalized_checks
            else:
                self.remediation_checks_config = {}
            self.remediation_env_vars = self.remediation_config.get("environment_overrides", {})
            
            # Load API timeout settings from global config / โหลดการตั้งค่า API timeout จากการตั้งค่าส่วนกลาง
            self.wait_for_api_enabled = self.remediation_global_config.get("wait_for_api", True)
            self.api_check_interval = self.remediation_global_config.get("api_check_interval", 5)
            self.api_max_retries = self.remediation_global_config.get("api_max_retries", 60)
            self.api_settle_time = self.remediation_global_config.get("api_settle_time", 15)
            
            # Resolve all variable references in checks / แก้ไขการอ้างอิงตัวแปรทั้งหมดในการตรวจสอบ
            self._resolve_references()
            self._calculate_level_totals()
            
            if self.verbose >= 1:
                print(f"{Colors.BLUE}[DEBUG] Loaded remediation config for {len(self.remediation_checks_config)} checks{Colors.ENDC}")
                print(f"{Colors.BLUE}[DEBUG] API timeout settings: interval={self.api_check_interval}s, max_retries={self.api_max_retries}, settle_time={self.api_settle_time}s{Colors.ENDC}")
                print(f"{Colors.BLUE}[DEBUG] Resolved variable references in checks{Colors.ENDC}")
        except json.JSONDecodeError as e:
            print(f"{Colors.RED}[!] Config parse error: {e}{Colors.ENDC}")
        except Exception as e:
            print(f"{Colors.RED}[!] Config load error: {e}{Colors.ENDC}")

    def _resolve_references(self):
        """
        Resolve all variable references in remediation checks.
        แก้ไขการอ้างอิงตัวแปรทั้งหมดในการตั้งค่าการแก้ไข
        
        Algorithm / อัลกอริทึม:
        1. Iterate through each check in remediation_checks_config / วนลูปผ่านการตรวจสอบแต่ละรายการ
        2. Identify all keys ending with '_ref' / ระบุคีย์ทั้งหมดที่ลงท้ายด้วย '_ref'
        3. Parse the dotted path from the reference value / แยกวิเคราะห์ dotted path จากค่าอ้างอิง
        4. Fetch the actual value from self.variables / ดึงค่าจริงจาก self.variables
        5. Inject/Overwrite the target key with the fetched value / ใส่หรือเขียนทับคีย์เป้าหมายด้วยค่าที่ดึงมา
        """
        reference_count = 0
        invalid_refs: List[Dict[str, Any]] = []
        
        for check_id, check_config in self.remediation_checks_config.items():
            
            # Find all keys ending with '_ref' / ค้นหาคีย์ทั้งหมดที่ลงท้ายด้วย '_ref'
            ref_keys = [key for key in check_config.keys() if key.endswith('_ref')]
            
            for ref_key in ref_keys:
                ref_path = check_config[ref_key]
                # Resolve target key by removing '_ref' and stripping leading underscore
                # e.g., '_required_value_ref' -> 'required_value'
                target_key = ref_key.replace('_ref', '').lstrip('_')
                
                # Fetch the value from variables using dotted path / ดึงค่าจากตัวแปรโดยใช้ dotted path
                if ref_path.startswith("variables."):
                    var_path = ref_path[len("variables."):]  # Remove "variables." prefix
                    resolved_value = self._get_nested_value(self.variables, var_path)
                else:
                    resolved_value = None
                
                if resolved_value is None:
                    invalid_refs.append(
                        {
                            'check_id': check_id,
                            'ref_key': ref_key,
                            'ref_path': ref_path,
                        }
                    )
                    if self.verbose >= 1:
                        print(f"{Colors.YELLOW}[!] Invalid reference in check {check_id}: {ref_path} not found{Colors.ENDC}")
                else:
                    # Type conversion: Convert JSON boolean to string if appropriate / การแปลงประเภท: แปลง JSON boolean เป็น string
                    if isinstance(resolved_value, bool):
                        resolved_value = str(resolved_value).lower()  # true -> "true", false -> "false"
                    
                    # Inject the resolved value into the check config / ใส่ค่าที่แก้ไขแล้วลงในการตั้งค่าการตรวจสอบ
                    check_config[target_key] = resolved_value
                    reference_count += 1
                    
                    if self.verbose >= 2:
                        print(f"{Colors.BLUE}[DEBUG] Resolved {check_id}.{target_key} = {resolved_value}{Colors.ENDC}")
        
        if self.verbose >= 1 and reference_count > 0:
            print(f"{Colors.GREEN}[+] Resolved {reference_count} variable references{Colors.ENDC}")
        
        if invalid_refs and self.verbose >= 1:
            print(f"{Colors.YELLOW}[!] Found {len(invalid_refs)} invalid references{Colors.ENDC}")
            for invalid in invalid_refs:
                print(f"    - {invalid['check_id']}: {invalid['ref_path']}")

    def _calculate_level_totals(self):
        """Track how many checks exist at Level 1 and Level 2 for reporting."""
        totals = {"L1": 0, "L2": 0}

        for check_config in self.remediation_checks_config.values():
            level_tag = str(check_config.get("level", "")).upper()
            if level_tag in totals:
                totals[level_tag] += 1

        self.level_check_totals = totals
        self.total_level_one_two_checks = totals.get("L1", 0) + totals.get("L2", 0)

        if self.verbose >= 2:
            print(f"{Colors.BLUE}[DEBUG] Level 1+2 check totals: {self.level_check_totals} ({self.total_level_one_two_checks} total){Colors.ENDC}")

    def _get_nested_value(self, data: Dict[str, Any], dotted_path: str) -> Optional[Any]:
        """
        Retrieve a value from nested dictionary using dotted path notation.
        ดึงค่าจากพจนานุกรมที่ซ้อนกันโดยใช้รูปแบบ dotted path
        
        Example / ตัวอย่าง:
            dotted_path = "api_server_flags.secure_port"
            Returns data['api_server_flags']['secure_port']
        """
        try:
            keys = dotted_path.split('.')
            node: Any = data

            for key in keys:
                if not isinstance(node, dict):
                    return None

                node_dict = cast(Dict[str, Any], node)
                next_value: Any = node_dict.get(key)
                if next_value is None:
                    return None

                node = next_value

            return node
        except Exception:
            return None

    def is_rule_excluded(self, rule_id: str) -> bool:
        """Check if rule is excluded / ตรวจสอบว่ากฎถูกยกเว้นหรือไม่"""
        return rule_id in self.excluded_rules

    def get_component_for_rule(self, rule_id: str) -> str:
        """Get component category for a rule / ได้รับหมวดหมู่ส่วนประกอบสำหรับกฎ"""
        for component, rules in self.component_mapping.items():
            if rule_id in rules:
                return component
        return "Other"

    def get_remediation_config_for_check(self, check_id: str) -> Dict[str, Any]:
        """Get remediation configuration for a specific check / ได้รับการตั้งค่าการแก้ไขสำหรับการตรวจสอบเฉพาะ"""
        if check_id in self.remediation_checks_config:
            return self.remediation_checks_config[check_id]
        return {}

    def log_activity(self, action: str, details: Optional[str] = None):
        """Log activities to file / บันทึกกิจกรรมลงไฟล์"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {action}"
        if details:
            log_entry += f" - {details}"
        
        try:
            with open(self.log_file, 'a') as f:
                f.write(log_entry + "\n")
        except Exception as e:
            if self.verbose >= 2:
                print(f"{Colors.RED}[DEBUG] Logging error: {e}{Colors.ENDC}")

    def run_preflight_checks(self):
        """
        Run comprehensive pre-flight checks before application starts.
        รันการตรวจสอบความพร้อมก่อนเริ่มโปรแกรม
        """
        print(f"\n{Colors.CYAN}[*] Running pre-flight checks...{Colors.ENDC}")
        
        checks_passed = True
        
        # Check 1: Required Tools (kubectl, jq) / ตรวจสอบเครื่องมือที่จำเป็น
        if not self._check_required_tools():
            checks_passed = False
        
        # Check 2: Root Permissions / ตรวจสอบสิทธิ์ Root
        if not self._check_root_permissions():
            checks_passed = False
        
        # Check 3: Config File Validity / ตรวจสอบความถูกต้องของไฟล์ตั้งค่า
        if not self._check_config_validity():
            checks_passed = False
        
        # Check 4: Optional Dependencies (openpyxl) / ตรวจสอบโมดูลเสริม
        self._check_optional_dependencies()
        
        if not checks_passed:
            print(f"\n{Colors.RED}{'='*70}")
            print(f"[ERROR] Pre-flight checks FAILED. Cannot proceed.")
            print(f"{'='*70}{Colors.ENDC}\n")
            sys.exit(1)
        
        print(f"{Colors.GREEN}[✓] All pre-flight checks passed!{Colors.ENDC}\n")

    def _check_required_tools(self):
        """Check if kubectl and jq are installed and executable / ตรวจสอบว่า kubectl และ jq ถูกติดตั้งหรือไม่"""
        tools_to_check = ["kubectl", "jq"]
        missing_tools: List[str] = []
        
        for tool in tools_to_check:
            tool_path = shutil.which(tool)
            if tool_path is None:
                missing_tools.append(tool)
                print(f"{Colors.RED}[ERROR] Required tool not found: {tool}{Colors.ENDC}")
            else:
                print(f"{Colors.GREEN}[✓] Tool found:{Colors.ENDC} {tool} ({tool_path})")
        
        if missing_tools:
            print(f"{Colors.RED}[ERROR] Missing critical tools: {', '.join(missing_tools)}{Colors.ENDC}")
            return False
        
        return True

    def _check_root_permissions(self):
        """Check if running as root (UID 0) / ตรวจสอบว่ารันด้วยสิทธิ์ root หรือไม่"""
        current_uid = os.getuid()
        
        if current_uid != 0:
            print(f"{Colors.RED}[ERROR] This application must run as root (UID 0){Colors.ENDC}")
            print(f"{Colors.RED}[ERROR] Current UID: {current_uid}{Colors.ENDC}")
            print(f"{Colors.YELLOW}[*] Please run with: sudo python3 {sys.argv[0]}{Colors.ENDC}")
            return False
        
        print(f"{Colors.GREEN}[✓] Running as root{Colors.ENDC} (UID: 0)")
        return True

    def _check_config_validity(self):
        """Check if cis_config.json is valid JSON and readable"""
        if not os.path.exists(self.config_file):
            print(f"{Colors.YELLOW}[!] Config file not found (optional): {self.config_file}{Colors.ENDC}")
            return True  # Optional - don't fail if missing, defaults will be used
        
        if not os.path.isfile(self.config_file):
            print(f"{Colors.RED}[ERROR] Config path is not a file: {self.config_file}{Colors.ENDC}")
            return False
        
        try:
            with open(self.config_file, 'r') as f:
                json.load(f)
            print(f"{Colors.GREEN}[✓] Config file is valid JSON:{Colors.ENDC} {self.config_file}")
            return True
        except json.JSONDecodeError as e:
            print(f"{Colors.RED}[ERROR] Config file is not valid JSON:{Colors.ENDC}")
            print(f"{Colors.RED}[ERROR] {self.config_file}{Colors.ENDC}")
            print(f"{Colors.RED}[ERROR] Details: {str(e)}{Colors.ENDC}")
            return False
        except PermissionError:
            print(f"{Colors.RED}[ERROR] Permission denied reading config file:{Colors.ENDC} {self.config_file}")
            return False
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to read config file:{Colors.ENDC} {str(e)}")
            return False

    def _check_optional_dependencies(self):
        """Check for optional dependencies like openpyxl"""
        if openpyxl_available:
            print(f"{Colors.GREEN}[✓] openpyxl is installed (Excel reporting enabled){Colors.ENDC}")
        else:
            print(f"{Colors.YELLOW}[!] openpyxl is not installed (Excel reporting disabled){Colors.ENDC}")
            print(f"{Colors.YELLOW}[!] Install with: pip install openpyxl{Colors.ENDC}")

    def ensure_audit_log_dir(self):
        """
        Ensure the audit log directory and file exist with correct permissions.
        ตรวจสอบให้แน่ใจว่าไดเรกทอรีและไฟล์บันทึกการตรวจสอบมีอยู่พร้อมสิทธิ์ที่ถูกต้อง
        """
        audit_dir = "/var/log/kubernetes/audit"
        audit_file = os.path.join(audit_dir, "audit.log")
        
        # Only attempt if running as root / ดำเนินการเฉพาะเมื่อรันด้วยสิทธิ์ root
        if os.getuid() != 0:
            if self.verbose >= 1:
                print(f"{Colors.YELLOW}[WARN] Not running as root. Skipping audit log directory creation.{Colors.ENDC}")
            return False

        try:
            # 1. Create directory if not exists (recursive) / สร้างไดเรกทอรีหากไม่มีอยู่ (แบบ recursive)
            if not os.path.exists(audit_dir):
                print(f"{Colors.CYAN}[*] Creating audit log directory: {audit_dir}{Colors.ENDC}")
                os.makedirs(audit_dir, mode=0o700, exist_ok=True)
            
            # Ensure directory permissions are 700 / ตรวจสอบให้แน่ใจว่าสิทธิ์ไดเรกทอรีคือ 700
            os.chmod(audit_dir, 0o700)
            
            # 2. Create audit.log if not exists / สร้าง audit.log หากไม่มีอยู่
            if not os.path.exists(audit_file):
                print(f"{Colors.CYAN}[*] Creating audit log file: {audit_file}{Colors.ENDC}")
                with open(audit_file, 'a'):
                    os.utime(audit_file, None)
            
            # 3. Set file permissions to 600 / ตั้งค่าสิทธิ์ไฟล์เป็น 600
            os.chmod(audit_file, 0o600)
            
            # 4. Set ownership to root:root (uid 0, gid 0) / ตั้งค่าความเป็นเจ้าของเป็น root:root
            try:
                # Use uid/gid 0 directly for maximum reliability / ใช้ uid/gid 0 โดยตรงเพื่อความน่าเชื่อถือสูงสุด
                os.chown(audit_dir, 0, 0)
                os.chown(audit_file, 0, 0)
                if self.verbose >= 2:
                    print(f"{Colors.BLUE}[DEBUG] Set ownership of {audit_dir} to root:root{Colors.ENDC}")
            except Exception as e:
                if self.verbose >= 1:
                    print(f"{Colors.YELLOW}[WARN] Failed to set ownership: {str(e)}{Colors.ENDC}")
            
            return True
        except Exception as e:
            print(f"{Colors.RED}[!] Error ensuring audit log directory: {str(e)}{Colors.ENDC}")
            return False

    def check_dependencies(self):
        """Verify required tools are installed / ตรวจสอบว่าเครื่องมือที่จำเป็นได้ถูกติดตั้ง"""
        missing = [tool for tool in REQUIRED_TOOLS if shutil.which(tool) is None]
        
        if missing:
            print(f"{Colors.RED}[-] Missing: {', '.join(missing)}{Colors.ENDC}")
            sys.exit(1)

    def detect_node_role(self) -> Optional[str]:
        """
        Detect current node role using multi-method approach.
        ตรวจจับบทบาทโหนดปัจจุบันโดยใช้หลายวิธี
        
        PRIORITY 1: Check running processes (most reliable) / ตรวจสอบโปรเซสที่กำลังรัน (น่าเชื่อถือที่สุด)
        PRIORITY 2: Check config/manifest files / ตรวจสอบไฟล์ตั้งค่าหรือ manifest
        PRIORITY 3: Fallback to kubectl with node labels / ใช้ kubectl ตรวจสอบ label ของโหนด
        """
        # PRIORITY 1: Check running processes (most reliable) / ตรวจสอบโปรเซสที่กำลังรัน
        try:
            # Check if kube-apiserver is running → Master node / ตรวจสอบว่า kube-apiserver รันอยู่หรือไม่
            cmd: List[str] = ["pgrep", "-l", "kube-apiserver"]
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode == 0:
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kube-apiserver process found → Master")
                return "master"
        except Exception:
            pass

        try:
            # Check if kubelet is running (without apiserver) → Worker node / ตรวจสอบว่า kubelet รันอยู่หรือไม่
            cmd: List[str] = ["pgrep", "-l", "kubelet"]
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode == 0:
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kubelet process found → Worker")
                return "worker"
        except Exception:
            pass

        # PRIORITY 2: Check config/manifest files / ตรวจสอบไฟล์ตั้งค่าหรือ manifest
        try:
            # Check for kube-apiserver manifest → Master node / ตรวจสอบไฟล์ manifest ของ kube-apiserver
            if os.path.exists("/etc/kubernetes/manifests/kube-apiserver.yaml"):
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kube-apiserver.yaml manifest found → Master")
                return "master"
        except Exception:
            pass

        try:
            # Check for kubelet config → Worker node / ตรวจสอบไฟล์ตั้งค่าของ kubelet
            if os.path.exists("/var/lib/kubelet/config.yaml"):
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kubelet config.yaml found → Worker")
                return "worker"
        except Exception:
            pass

        try:
            # Additional check: /etc/kubernetes/kubelet.conf exists → likely Worker / ตรวจสอบไฟล์ kubelet.conf
            if os.path.exists("/etc/kubernetes/kubelet.conf") and not os.path.exists("/etc/kubernetes/manifests/kube-apiserver.yaml"):
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kubelet.conf found → Worker")
                return "worker"
        except Exception:
            pass

        # PRIORITY 3: Fallback to kubectl node labels (original method) / ใช้ kubectl ตรวจสอบ label
        try:
            hostname = socket.gethostname()
            kubectl_cmd = self.get_kubectl_cmd()
            
            if not kubectl_cmd:
                if self.verbose >= 2:
                    print("[DEBUG] Node role detection: kubectl not available, cannot determine role")
                return None

            labels_cmd: List[str] = kubectl_cmd + [
                "get",
                "node",
                hostname,
                "--no-headers",
                "-o",
                "jsonpath={.metadata.labels}"
            ]
            result = subprocess.run(
                labels_cmd,
                capture_output=True,
                text=True,
                timeout=10
            )

            if result.returncode == 0:
                labels = result.stdout.lower()
                if labels:
                    if "control-plane" in labels or "master" in labels:
                        if self.verbose >= 2:
                            print("[DEBUG] Node role detection: kubectl labels indicate → Master")
                        return "master"
                    if self.verbose >= 2:
                        print("[DEBUG] Node role detection: kubectl labels indicate → Worker")
                    return "worker"
        except Exception:
            pass

        # All detection methods failed / วิธีการตรวจจับทั้งหมดล้มเหลว
        if self.verbose >= 2:
            print("[DEBUG] Node role detection: all methods failed, unable to determine role")
        return None

    def get_kubectl_cmd(self) -> List[str]:
        """
        Detect kubeconfig and return kubectl command.
        ตรวจจับ kubeconfig และส่งกลับคำสั่ง kubectl
        """
        kubeconfig_paths = [
            os.environ.get('KUBECONFIG'),
            "/etc/kubernetes/admin.conf",
            os.path.expanduser("~/.kube/config"),
            f"/home/{os.environ.get('SUDO_USER', '')}/.kube/config"
        ]
        
        for config_path in kubeconfig_paths:
            if config_path and os.path.exists(config_path):
                if self.verbose >= 2:
                    print(f"{Colors.BLUE}[DEBUG] Using Kubeconfig: {config_path}{Colors.ENDC}")
                return ["kubectl", "--kubeconfig", config_path]
        
        return ["kubectl"]

    def check_health(self):
        """
        Check Kubernetes cluster health using secure API endpoint.
        ตรวจสอบสุขภาพของคลัสเตอร์ Kubernetes โดยใช้ endpoint API ที่ปลอดภัย
        """
        print(f"{Colors.CYAN}[*] Checking Cluster Health...{Colors.ENDC}")
        
        # 1. Try Secure API Health Check / ตรวจสอบสุขภาพ API แบบปลอดภัย
        health_url = "https://127.0.0.1:6443/healthz"
        try:
            # Suppress insecure request warnings for this local check
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            
            # FORCE DISABLE VERIFY for localhost to avoid SSLError (hostname mismatch)
            # Accept 200 (OK) or 401 (Unauthorized) as success
            response = requests.get(health_url, verify=False, timeout=5)
            
            if response.status_code in (200, 401):
                self.health_status = f"OK (Healthy - Status {response.status_code})"
                print(f"{Colors.GREEN}    -> {self.health_status}{Colors.ENDC}")
                return self.health_status
        except Exception as e:
            if self.verbose >= 1:
                print(f"{Colors.YELLOW}[WARN] API Healthz check failed: {str(e)}{Colors.ENDC}")

        # 2. Fallback to kubectl if API check fails or for more detail
        admin_conf = "/etc/kubernetes/admin.conf"

        if os.path.exists(admin_conf):
            kubectl = self.get_kubectl_cmd()

            try:
                # Check nodes / ตรวจสอบโหนด
                nodes_result = subprocess.run(
                    kubectl + ["get", "nodes"],
                    capture_output=True, text=True, timeout=10
                )
            
                if nodes_result.returncode != 0:
                    self.health_status = "CRITICAL (Nodes Unreachable)"
                    print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")
                    return self.health_status
                
                # Check pod status / ตรวจสอบสถานะพอด
                pods_result = subprocess.run(
                    kubectl + ["get", "pods", "-n", "kube-system", 
                              "--field-selector", "status.phase!=Running"],
                    capture_output=True, text=True, timeout=10
                )
                
                unhealthy_pods = [
                    line for line in pods_result.stdout.split('\n')
                    if line.strip() and "NAME" not in line
                ]
                
                if unhealthy_pods:
                    self.health_status = f"WARNING ({len(unhealthy_pods)} Unhealthy Pods)"
                    print(f"{Colors.YELLOW}    -> {self.health_status}{Colors.ENDC}")
                else:
                    self.health_status = "OK (Healthy)"
                    print(f"{Colors.GREEN}    -> {self.health_status}{Colors.ENDC}")
                
            except subprocess.TimeoutExpired:
                self.health_status = "ERROR (Timeout)"
                print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")
            except Exception as e:
                self.health_status = f"ERROR ({str(e)})"
                print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")
        else:
            # Worker node health check / ตรวจสอบสุขภาพของ Worker node
            try:
                kubelet = subprocess.run(
                    ["systemctl", "is-active", "kubelet"],
                    capture_output=True, text=True, timeout=10
                )
                if kubelet.returncode == 0 and kubelet.stdout.strip() == "active":
                    self.health_status = "OK (Worker Kubelet Running)"
                    print(f"{Colors.GREEN}    -> {self.health_status}{Colors.ENDC}")
                else:
                    self.health_status = "CRITICAL (Kubelet Down)"
                    print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")
            except subprocess.TimeoutExpired:
                self.health_status = "ERROR (Timeout)"
                print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")
            except Exception as e:
                self.health_status = f"ERROR ({str(e)})"
                print(f"{Colors.RED}    -> {self.health_status}{Colors.ENDC}")

        return self.health_status

    def wait_for_healthy_cluster(self, skip_health_check: bool = False) -> bool:
        """
        Wait for cluster to be healthy after API server restart.
        รอให้คลัสเตอร์กลับมาใช้งานได้ปกติหลังจากรีสตาร์ท API server
        
        MASTER NODE (3-Step Verification):
        - Step 1 (TCP): Verify API server port is open / ตรวจสอบพอร์ต TCP 6443
        - Step 2 (Application Ready): Verify API responds to requests / ตรวจสอบว่า API ตอบสนอง
        - Step 3 (Settle Time): Force sleep to allow components to sync / รอให้ส่วนประกอบต่างๆ ซิงค์ข้อมูล
        
        WORKER NODE: Checks systemctl is-active kubelet / ตรวจสอบสถานะ kubelet
        """
        # CRITICAL: If skip_health_check=True, bypass ALL verification logic / ข้ามการตรวจสอบทั้งหมดหากระบุ
        if skip_health_check:
            if self.verbose >= 1:
                print(f"{Colors.GREEN}[*] Health check skipped (safe operation - no service impact).{Colors.ENDC}")
            return True
        
        if not self.wait_for_api_enabled:
            if self.verbose >= 1:
                print(f"{Colors.CYAN}[*] API health check disabled in config.{Colors.ENDC}")
            return True
        
        admin_conf = "/etc/kubernetes/admin.conf"
        total_timeout = self.api_check_interval * self.api_max_retries
        
        # --- WORKER NODE LOGIC ---
        if not os.path.exists(admin_conf):
            print(f"{Colors.YELLOW}[*] Worker Node detected. Checking kubelet health...{Colors.ENDC}")
            
            try:
                kubelet_result = subprocess.run(
                    ["systemctl", "is-active", "kubelet"],
                    capture_output=True,
                    text=True,
                    timeout=10
                )
                
                kubelet_status = kubelet_result.stdout.strip()
                
                if kubelet_result.returncode == 0 and kubelet_status == "active":
                    self.health_status = "OK (Worker Kubelet Running)"
                    print(f"{Colors.GREEN}    [OK] Worker kubelet is active.{Colors.ENDC}")
                    return True
                else:
                    self.health_status = "CRITICAL (Kubelet Not Active)"
                    print(f"{Colors.RED}    [FAIL] Worker kubelet is {kubelet_status or 'not responding'}.{Colors.ENDC}")
                    return False
            
            except subprocess.TimeoutExpired:
                self.health_status = "CRITICAL (Kubelet Check Timeout)"
                print(f"{Colors.RED}    [FAIL] Kubelet health check timed out.{Colors.ENDC}")
                return False
            except Exception as e:
                self.health_status = f"CRITICAL (Kubelet Check Error: {str(e)})"
                print(f"{Colors.RED}    [FAIL] Error checking kubelet: {str(e)}{Colors.ENDC}")
                return False
        
        # --- MASTER NODE LOGIC (3-STEP VERIFICATION) ---
        print(f"{Colors.YELLOW}[*] Master Node detected. 3-Step health verification (Timeout: {total_timeout}s)...{Colors.ENDC}")
        
        kubectl = self.get_kubectl_cmd()
        count = 0
        step1_passed = False
        
        while count < self.api_max_retries:
            elapsed = count * self.api_check_interval
            
            # --- STEP 1: TCP Check (Port Open) / ตรวจสอบพอร์ต TCP ---
            if not step1_passed:
                try:
                    result = subprocess.run(
                        ["nc", "-zv", "localhost", "6443"],
                        capture_output=True,
                        text=True,
                        timeout=5
                    )
                    
                    if result.returncode == 0:
                        print(f"{Colors.GREEN}    [Step 1/3 OK] TCP port 6443 open (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                        step1_passed = True
                    else:
                        print(f"{Colors.YELLOW}    [Step 1/3 WAIT] TCP port 6443 not responding... (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                        time.sleep(self.api_check_interval)
                        count += 1
                        continue
                
                except Exception:
                    print(f"{Colors.YELLOW}    [Step 1/3 WAIT] TCP check failed... (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                    time.sleep(self.api_check_interval)
                    count += 1
                    continue
            
            # --- STEP 2: Application Ready (API healthz/readyz) / ตรวจสอบความพร้อมของ API ---
            if step1_passed:
                try:
                    # Suppress insecure request warnings for this local check
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                    
                    # FORCE DISABLE VERIFY for localhost to avoid SSLError (hostname mismatch)
                    # Accept 200 (OK) or 401 (Unauthorized) as success
                    response = requests.get(self.health_check_url, verify=False, timeout=5)
                    
                    if response.status_code in (200, 401):
                        print(f"{Colors.GREEN}    [Step 2/3 OK] API server is ready (Status {response.status_code}) (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                        
                        # --- STEP 3: Settle Time (Allow components to sync) / เวลารอให้ระบบนิ่ง ---
                        print(f"{Colors.CYAN}    [Step 3/3] Settling ({self.api_settle_time}s for etcd/scheduler/controller-manager sync)...{Colors.ENDC}")
                        time.sleep(self.api_settle_time)
                        
                        self.health_status = f"OK (Healthy - Status {response.status_code})"
                        print(f"{Colors.GREEN}    [OK] Cluster is online and stable.{Colors.ENDC}")
                        return True
                    elif response.status_code in (500, 503):
                        print(f"{Colors.YELLOW}    [Step 2/3 WAIT] API initializing... (Status {response.status_code}) (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                    else:
                        print(f"{Colors.YELLOW}    [Step 2/3 WAIT] API returned {response.status_code}... (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                
                except requests.exceptions.ConnectionError:
                    print(f"{Colors.YELLOW}    [Step 2/3 WAIT] API restarting... (Connection Refused) (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                except Exception as e:
                    # Fallback to kubectl if requests fails for other reasons
                    if self.verbose >= 2:
                        print(f"{Colors.BLUE}[DEBUG] Step 2 requests failed: {str(e)}{Colors.ENDC}")
                        
                    try:
                        readyz_result = subprocess.run(
                            kubectl + ["get", "--raw=/readyz"],
                            capture_output=True,
                            text=True,
                            timeout=10
                        )
                        
                        if readyz_result.returncode == 0:
                            print(f"{Colors.GREEN}    [Step 2/3 OK] API server is ready via kubectl (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                            
                            # --- STEP 3: Settle Time ---
                            print(f"{Colors.CYAN}    [Step 3/3] Settling ({self.api_settle_time}s)...{Colors.ENDC}")
                            time.sleep(self.api_settle_time)
                            
                            self.health_status = "OK (Healthy - 3-Step Verified via kubectl)"
                            return True
                        else:
                            print(f"{Colors.YELLOW}    [Step 2/3 WAIT] API not responding... (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                    except Exception:
                        print(f"{Colors.YELLOW}    [Step 2/3 WAIT] API check error... (Attempt {count+1}/{self.api_max_retries}, {elapsed}s){Colors.ENDC}")
                
                time.sleep(self.api_check_interval)
                count += 1

        print(f"{Colors.RED}    [FAIL] Cluster did not achieve healthy state within {total_timeout} seconds (3-step verification failed).{Colors.ENDC}")
        self.health_status = "CRITICAL (Recovery Timeout - 3-Step Failed)"
        return False

    def extract_metadata_from_script(self, script_path: Optional[str]):
        """
        Extract title from script file header.
        แยกชื่อเรื่องจากส่วนหัวไฟล์สคริปต์
        """
        if not script_path or not os.path.exists(script_path):
            return "Title not found"
        
        try:
            with open(script_path, 'r', encoding='utf-8', errors='ignore') as f:
                # Read first 30 lines / อ่านบรรทัดแรก 30 บรรทัด
                lines = f.readlines()[:30]
            
            for line in lines:
                line = line.strip()
                if not line.startswith("#"):
                    continue
                
                # Look for "Title:" tag / มองหาแท็ก "Title:"
                if "Title:" in line:
                    return line.split("Title:", 1)[1].strip()
                
                # Look for common patterns / มองหารูปแบบทั่วไป
                clean = line.lstrip("#").strip()
                prefixes = ["Ensure", "Enable", "Disable", "Configure", "Restrict"]
                if any(clean.startswith(p) for p in prefixes):
                    if 10 < len(clean) < 150:
                        return clean
        except Exception:
            pass
        
        return "Title not found"

    def get_scripts(self, mode: str, target_level: str, target_role: str) -> List[Dict[str, Any]]:
        """
        Get list of audit/remediation scripts.
        ได้รับรายชื่อสคริปต์ตรวจสอบ/การแก้ไข
        """
        suffix = "_remediate.sh" if mode == "remediate" else "_audit.sh"
        scripts: List[Dict[str, Any]] = []
        
        # Determine levels to scan / กำหนดระดับที่จะสแกน
        levels = ['1', '2'] if target_level == "all" else [target_level]
        roles = ['Master', 'Worker'] if target_role == "all" else [target_role.capitalize()]
        
        for role in roles:
            for level in levels:
                dir_path = os.path.join(self.base_dir, f"Level_{level}_{role}_Node")
                
                if not os.path.isdir(dir_path):
                    continue
                
                for filename in sorted(os.listdir(dir_path)):
                    if not filename.endswith(suffix):
                        continue
                    
                    scripts.append({
                        "path": os.path.join(dir_path, filename),
                        "id": filename.replace(suffix, ""),
                        "role": role.lower(),
                        "name": filename,
                        "level": level
                    })
        
        return scripts

    def _prepare_remediation_env(self, script_id: str, remediation_cfg: Optional[Dict[str, Any]]) -> Dict[str, str]:
        """Prepare environment variables for remediation scripts."""
        env = os.environ.copy()
        
        # Explicitly add KUBECONFIG
        kubeconfig_paths = [
            os.environ.get('KUBECONFIG'),
            "/etc/kubernetes/admin.conf",
            os.path.expanduser("~/.kube/config"),
            f"/home/{os.environ.get('SUDO_USER', '')}/.kube/config"
        ]
        for config_path in kubeconfig_paths:
            if config_path and os.path.exists(config_path):
                env["KUBECONFIG"] = config_path
                break
        
        # Add global remediation config
        env.update({
            "BACKUP_ENABLED": str(self.remediation_global_config.get("backup_enabled", True)).lower(),
            "BACKUP_DIR": self.remediation_global_config.get("backup_dir", "/var/backups/cis-remediation"),
            "DRY_RUN": str(self.remediation_global_config.get("dry_run", False)).lower(),
            "WAIT_FOR_API": str(self.wait_for_api_enabled).lower(),
            "API_CHECK_INTERVAL": str(self.api_check_interval),
            "API_MAX_RETRIES": str(self.api_max_retries)
        })
        
        # Add check-specific remediation config
        if remediation_cfg:
            for key, value in remediation_cfg.items():
                if key.startswith('_') or key in ['skip', 'enabled', 'id', 'path', 'role', 'level', 'requires_health_check']:
                    continue
                
                env_key = key.upper()
                if isinstance(value, bool):
                    env[env_key] = "true" if value else "false"
                elif isinstance(value, (list, dict)):
                    env[env_key] = json.dumps(value)
                elif isinstance(value, (int, float)):
                    env[env_key] = str(value)
                elif value is None:
                    env[env_key] = ""
                else:
                    str_value = str(value)
                    if str_value.startswith('"') and str_value.endswith('"'):
                        str_value = str_value[1:-1]
                    env[env_key] = str_value
        
        # Add global environment overrides
        env.update(self.remediation_env_vars)
        return env

    def _protect_kube_flannel_pss(self, env: Optional[Dict[str, str]] = None):
        """Ensure kube-flannel keeps a privileged PSS label to avoid breaking the CNI."""
        namespace = "kube-flannel"
        label = "pod-security.kubernetes.io/enforce=privileged"
        cmd = ["kubectl", "label", "--overwrite", "ns", namespace, label]
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=30,
                env=env
            )
        except FileNotFoundError:
            print(f"{Colors.RED}[!] kubectl not found while protecting {namespace}{Colors.ENDC}")
            return
        except Exception as exc:
            print(f"{Colors.YELLOW}[WARN] Failed to label {namespace} as privileged: {exc}{Colors.ENDC}")
            return

        if result.returncode == 0:
            print(f"{Colors.GREEN}[✓] Ensured {namespace} is marked pod-security.kubernetes.io/enforce=privileged{Colors.ENDC}")
        else:
            stderr = result.stderr.strip()
            print(f"{Colors.YELLOW}[WARN] kube-flannel PSS override failed (code {result.returncode}): {stderr}{Colors.ENDC}")

    def _is_kube_flannel_privileged(self) -> bool:
        """Return True when kube-flannel already has a privileged PSS label."""
        try:
            result = subprocess.run(
                ["kubectl", "get", "namespace", "kube-flannel", "-o", "json"],
                capture_output=True,
                text=True,
                timeout=15
            )
        except FileNotFoundError:
            return False
        except Exception:
            return False

        if result.returncode != 0 or not result.stdout:
            return False

        try:
            data = json.loads(result.stdout)
        except json.JSONDecodeError:
            return False

        labels = cast(Dict[str, str], data.get("metadata", {}).get("labels") or {})
        return bool(labels.get("pod-security.kubernetes.io/enforce") == "privileged")

    def is_namespace_exempt(self, namespace: str) -> bool:
        if not namespace:
            return False
        cmd = self.get_kubectl_cmd() + [
            "get",
            "namespace",
            namespace,
            "-o",
            "jsonpath={.metadata.labels.cis-compliance/exempt}",
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
        except FileNotFoundError:
            return False
        if result.returncode != 0:
            return False
        return result.stdout.strip().lower() == "true"

    def label_namespace_as_exempt(self, namespace: str) -> bool:
        if not namespace:
            return False
        if not wait_for_api_ready(
            log_callback=lambda msg: print(f"{Colors.YELLOW}[WARN] {msg}{Colors.ENDC}")
        ):
            return False
        cmd = self.get_kubectl_cmd() + [
            "label",
            "namespace",
            namespace,
            "cis-compliance/exempt=true",
            "--overwrite",
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
        except FileNotFoundError:
            print(f"{Colors.YELLOW}[WARN] kubectl missing — cannot label {namespace}.{Colors.ENDC}")
            return False
        if result.returncode != 0:
            stderr = result.stderr.strip()
            print(f"{Colors.YELLOW}[WARN] Failed to label namespace {namespace}: {stderr}{Colors.ENDC}")
            return False
        print(f"{Colors.GREEN}[✓] Marked {namespace} as cis-compliance/exempt=true{Colors.ENDC}")
        return True

    def fix_item_internal(self, script: Dict[str, Any], remediation_cfg: Optional[Dict[str, Any]]) -> ResultTuple:
        """
        Main remediation router for internal Python-based fixes.
        ตัวเลือกหลักสำหรับการแก้ไขโดยใช้ Python ภายใน
        
        Handles YAML config checks internally and falls back to bash for system checks.
        จัดการการตรวจสอบการตั้งค่า YAML ภายใน และใช้ bash สำหรับการตรวจสอบระบบ
        """
        script_id = script["id"]
        script_path = script["path"]
        remediation_cfg = remediation_cfg or {}
        
        # 1. Detect check type / ตรวจสอบประเภทของการตรวจสอบ
        is_yaml, manifest_path = self._classify_remediation_type(script_id)
        
        if is_yaml and manifest_path and os.path.exists(manifest_path):
            print(f"{Colors.CYAN}[*] Internal YAML Fix: {script_id} -> {manifest_path}{Colors.ENDC}")
            
            # Identify CSV Checks / ระบุการตรวจสอบที่เป็น CSV
            CSV_CHECKS = ["1.2.8", "1.2.11", "1.2.14", "1.2.29"]
            mod_type = "csv" if script_id in CSV_CHECKS else "string"
            
            # Prepare modifications from config / เตรียมการแก้ไขจากการตั้งค่า
            modifications: Dict[str, List[str]] = {'flags': []}
            
            # Handle single flag / จัดการแฟล็กเดียว
            flag_name = remediation_cfg.get('flag_name')
            expected_value = remediation_cfg.get('expected_value')
            
            if flag_name:
                if not flag_name.startswith('--'):
                    flag_name = '--' + flag_name
                modifications['flags'].append(f"{flag_name}={expected_value}")
            
            # Handle multiple flags (multi_flag_check) / จัดการหลายแฟล็ก
            # Normalize and validate to an explicit dict so static analyzers know the type
            flags_obj: Dict[str, Any] = {}
            cfg_flags = remediation_cfg.get('flags', {})
            if isinstance(cfg_flags, dict):
                flags_obj = cast(Dict[str, Any], cfg_flags)
            # flags_obj is now a Dict[str, Any] for iteration
            for f_name, f_val in flags_obj.items():
                # Normalize key to string for robustness (avoid redundant isinstance checks)
                try:
                    f_name_str = str(f_name)
                except Exception:
                    continue
                if not f_name_str.startswith('--'):
                    f_name_str = '--' + f_name_str
                modifications['flags'].append(f"{f_name_str}={f_val}")
            
            # 2. Execute Atomic Remediation Flow / ดำเนินการขั้นตอนการแก้ไขแบบ Atomic
            # Phase 1: Backup & Apply / ขั้นตอนที่ 1: สำรองข้อมูลและนำไปใช้
            success, msg = self.atomic_manager.update_manifest_safely(manifest_path, modifications, mod_type)
            if not success:
                return "FAIL", f"Atomic write failed: {msg}", None, []
            
            if msg == "No changes needed":
                # DEEP RUNTIME VERIFICATION (Case B: File OK but Process might be STALE)
                # การตรวจสอบรันไทม์เชิงลึก (กรณี B: ไฟล์ถูกต้องแต่โปรเซสอาจจะยังไม่อัปเดต)
                process_name = self._get_process_name_for_check(script_id)
                if process_name and modifications['flags']:
                    # Check the first flag as a representative for runtime verification
                    first_flag = modifications['flags'][0]
                    if '=' in first_flag:
                        f_key, f_val = first_flag.split('=', 1)
                        
                        if self.verify_runtime_flag(process_name, f_key, f_val):
                            return "PASS", "No changes needed (File & Runtime compliant)", None, []
                        else:
                            print(f"{Colors.YELLOW}[STALE CONFIG] File is compliant but Runtime is STALE for {process_name}. Forcing reload...{Colors.ENDC}")
                            self.log_activity("STALE_CONFIG_DETECTED", f"{script_id}: {process_name}")
                            self._aggressive_restart_component(process_name)
                            
                            # Re-verify after reload / ตรวจสอบอีกครั้งหลังโหลดใหม่
                            print(f"{Colors.CYAN}[*] Re-verifying runtime after reload...{Colors.ENDC}")
                            time.sleep(10) # Wait for process to stabilize
                            
                            if self.verify_runtime_flag(process_name, f_key, f_val):
                                return "PASS", "Fixed (Stale config reloaded successfully)", None, []
                            else:
                                return "FAIL", f"Stale config reload failed to apply changes for {process_name}", None, []
                
                return "PASS", "No changes needed (already compliant)", None, []

            # Phase 2: Health Check (Wait for API) / ขั้นตอนที่ 2: ตรวจสอบสุขภาพ (รอ API)
            print(f"{Colors.YELLOW}[*] Waiting for cluster health...{Colors.ENDC}")
            is_healthy, health_msg = self.atomic_manager.wait_for_cluster_healthy()
            
            # Extract backup path for potential rollback
            backup_path = None
            match = re.search(r"Backup: (.*)", msg)
            if match:
                backup_path = match.group(1)

            if not is_healthy:
                print(f"{Colors.RED}[!] Cluster unhealthy after fix. Rolling back...{Colors.ENDC}")
                if backup_path:
                    print(f"{Colors.YELLOW}[ROLLBACK] Restoring original config due to health failure...{Colors.ENDC}")
                    self.atomic_manager.rollback(manifest_path, backup_path)
                    print(f"{Colors.CYAN}[*] Waiting 15s for API server to recover from rollback...{Colors.ENDC}")
                    time.sleep(15)
                
                return "FAIL", f"Health check failed: {health_msg}. Rolled back.", None, []
            
            # Phase 3: Deep Runtime Verification / ขั้นตอนที่ 3: การตรวจสอบรันไทม์เชิงลึก
            process_name = self._get_process_name_for_check(script_id)
            if process_name and modifications['flags']:
                # Check the first flag as a representative for runtime verification
                first_flag = modifications['flags'][0]
                if '=' in first_flag:
                    f_key, f_val = first_flag.split('=', 1)
                    
                    # Initial verification / การตรวจสอบเบื้องต้น
                    if not self.verify_runtime_flag(process_name, f_key, f_val):
                        print(f"{Colors.YELLOW}[STALE CONFIG] Manifest updated but runtime state is stale for {process_name}. Attempting recovery...{Colors.ENDC}")
                        self.log_activity("STALE_CONFIG_AFTER_FIX", f"{script_id}: {process_name}")
                        
                        # Try Pod reload first for static pods / ลองโหลด Pod ใหม่ก่อนสำหรับ static pods
                        if process_name in {"kube-apiserver", "kube-controller-manager", "kube-scheduler"}:
                            self._aggressive_restart_component(process_name)
                            time.sleep(10)
                        
                        # If still stale, we rely on the later runtime check to decide / หากยังคงค้างอยู่ ให้ใช้การตรวจสอบรันไทม์ต่อ
                        if not self.verify_runtime_flag(process_name, f_key, f_val):
                            print(f"{Colors.YELLOW}[WARN] Runtime still stale after forced pod reload.{Colors.ENDC}")

            # Phase 4: Audit Verification / ขั้นตอนที่ 4: ตรวจสอบการตรวจสอบ
            print(f"{Colors.YELLOW}[*] Verifying fix with audit script...{Colors.ENDC}")
            audit_script_path = script_path.replace("_remediate.sh", "_audit.sh")
            if os.path.exists(audit_script_path):
                audit_passed, audit_msg = self.atomic_manager.verify_remediation(script_id, audit_script_path)
                
                if not audit_passed:
                    print(f"{Colors.RED}[ERROR] Verification failed after remediation. Config is invalid.{Colors.ENDC}")
                    if backup_path:
                        print(f"{Colors.YELLOW}[ROLLBACK] Restored original config to ensure cluster stability.{Colors.ENDC}")
                        self.atomic_manager.rollback(manifest_path, backup_path)
                        self.log_activity("REMEDIATION_ROLLBACK", f"{script_id}: Audit failed, rolled back to {backup_path}")
                    return "FAIL", f"Audit failed: {audit_msg}", None, []
            
            return "PASS", "Remediation applied and verified successfully", None, []
            
        else:
            # System Check (chmod/chown/etc) or manifest not found - Fallback to Bash / การตรวจสอบระบบหรือหา manifest ไม่พบ - ใช้ Bash แทน
            if is_yaml and not os.path.exists(manifest_path):
                print(f"{Colors.YELLOW}[!] Manifest {manifest_path} not found. Falling back to bash script.{Colors.ENDC}")
            
            env = self._prepare_remediation_env(script_id, remediation_cfg)
            result = subprocess.run(["bash", script_path], capture_output=True, text=True, timeout=self.script_timeout, env=env)
            # Help static analyzers by casting the parsed output to a known tuple type
            status, reason, fix_hint, cmds = self._parse_script_output(result, script_id, "remediate", False)

            if script_id.startswith("5.2."):
                self._protect_kube_flannel_pss(env=env)
            return status, reason, fix_hint, cmds

    def _get_process_name_for_check(self, check_id: str) -> Optional[str]:
        """Map check ID to process name / แมป ID การตรวจสอบกับชื่อโปรเซส"""
        if check_id.startswith('1.2.'): return "kube-apiserver"
        if check_id.startswith('1.3.'): return "kube-controller-manager"
        if check_id.startswith('1.4.'): return "kube-scheduler"
        if check_id.startswith('2.'): return "etcd"
        if check_id.startswith('4.'): return "kubelet"
        return None

    def verify_process_runtime(self, process_name: str, flag: str, expected_value: Any) -> Tuple[str, str]:
        """
        Verify if the running process has the expected flag and value.
        ตรวจสอบว่าโปรเซสที่กำลังรันมี flag และค่าที่คาดหวังหรือไม่

        Returns: (status, message)
        Status values: "VERIFIED", "NOT_APPLIED", "CRASHED", "ERROR"
        """
        try:
            # Use pgrep -a to get full command line / ใช้ pgrep -a เพื่อรับบรรทัดคำสั่งแบบเต็ม
            cmd = ["pgrep", "-a", process_name]
            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode != 0 or not result.stdout:
                # Fallback to ps -ef for some systems / ใช้ ps -ef เป็นทางเลือกสำหรับบางระบบ
                ps_cmd = f"ps -ef | grep {process_name} | grep -v grep"
                ps_result = subprocess.run(ps_cmd, shell=True, capture_output=True, text=True)
                if not ps_result.stdout:
                    return "CRASHED", "Process not running"
                process_cmdline = ps_result.stdout
            else:
                process_cmdline = result.stdout

            # Normalize flag (ensure it starts with --)
            if not flag.startswith('--'):
                flag = '--' + flag

            if flag == '--authorization-mode' and expected_value:
                auth_match = re.search(r"--authorization-mode(?:=|\s+)([^\s]+)", process_cmdline)
                if auth_match:
                    actual_value = auth_match.group(1)
                    actual_tokens = {token.strip() for token in actual_value.split(',') if token.strip()}
                    expected_tokens = {token.strip() for token in str(expected_value).split(',') if token.strip()}
                    if expected_tokens and expected_tokens.issubset(actual_tokens):
                        return "VERIFIED", f"Running with {flag}={actual_value}"
                    return "NOT_APPLIED", f"Flag found but value mismatch (Expected: {expected_value}, actual: {actual_value})"
                return "NOT_APPLIED", f"Flag {flag} not found in runtime"

            # Pattern matches --flag=value or --flag value
            # Also handles boolean flags where --flag=true might be just --flag
            pattern = rf"{re.escape(flag)}([=\s]+{re.escape(str(expected_value))})?(\s|$)"

            # Special case for boolean 'true' / กรณีพิเศษสำหรับค่าบูลีน 'true'
            if str(expected_value).lower() == 'true':
                # Match --flag=true OR just --flag
                pattern = rf"{re.escape(flag)}([=\s]+true)?(\s|$)"
            elif str(expected_value).lower() == 'false':
                # Match --flag=false OR absence of flag (though absence is harder to verify here)
                pattern = rf"{re.escape(flag)}[=\s]+false(\s|$)"

            if re.search(pattern, process_cmdline):
                return "VERIFIED", f"Running with {flag}={expected_value}"

            # Check if flag exists but value is wrong
            if flag in process_cmdline:
                return "NOT_APPLIED", f"Flag found but value mismatch (Expected: {expected_value})"
            else:
                return "NOT_APPLIED", f"Flag {flag} not found in runtime"

        except Exception as e:
            return "ERROR", f"Verification error: {str(e)}"

    def verify_runtime_flag(self, process_name: str, flag: str, value: Any) -> bool:
        """
        Deep Runtime Verification: Check if the running process has the specific flag=value.
        การตรวจสอบรันไทม์เชิงลึก: ตรวจสอบว่าโปรเซสที่กำลังรันมี flag=value ที่ระบุหรือไม่
        """
        return self.verify_runtime_config(process_name, flag, value)

    def verify_runtime_config(self, process_name: str, flag_key: str, expected_value: Any) -> bool:
        """
        Verify if the running process has the expected flag and value.
        Returns: True (Applied) / False (Not Applied)
        """
        status, _ = self.verify_process_runtime(process_name, flag_key, expected_value)
        return status == "VERIFIED"

    def _aggressive_restart_component(self, component: str) -> bool:
        """Hard kill every container for the component and verify kubelet restarts it."""
        component_aliases = {
            "kube-apiserver": "kube-apiserver",
            "kube-controller-manager": "kube-controller-manager",
            "kube-scheduler": "kube-scheduler",
            "etcd": "etcd"
        }
        comp_name = component_aliases.get(component)
        if not comp_name:
            print(f"{Colors.YELLOW}[WARN] No crictl helper available for {component}.{Colors.ENDC}")
            return False

        try:
            existing_ids = self._list_crictl_ids(comp_name)
        except Exception as exc:
            print(f"{Colors.RED}[!] Failed to list {component} containers: {exc}{Colors.ENDC}")
            existing_ids = []

        if existing_ids:
            print(f"{Colors.YELLOW}[WARN] Performing HARD KILL on {component} to force config reload...{Colors.ENDC}")
            for cid in existing_ids:
                subprocess.run(["crictl", "stop", cid], check=False)
                subprocess.run(["crictl", "rm", cid], check=False)
                print(f"[✓] Killed & Removed container: {cid}")
        else:
            print(f"{Colors.YELLOW}[WARN] No containers found for {component}; skipping kill phase.{Colors.ENDC}")

        start_time = time.time()
        seen_ids = set(existing_ids)

        while time.time() - start_time < 60:
            running_ids = set(self._list_crictl_ids(comp_name, state="Running"))
            new_ids = running_ids - seen_ids
            if new_ids:
                cid = sorted(new_ids)[0]
                print(f"{Colors.GREEN}[✓] {component} resurrected with new container {cid}.{Colors.ENDC}")
                return True
            seen_ids.update(running_ids)
            time.sleep(2)

        print(f"{Colors.RED}[!] Critical: {component} failed to resurrect. Check kubelet logs.{Colors.ENDC}")
        return False

    def _list_crictl_ids(self, name: str, state: Optional[str] = None) -> List[str]:
        """
        List container IDs via crictl filtered by name and optional state.
        Returns a list of container id strings; returns empty list on error.
        """
        try:
            cmd = ["crictl", "ps", "--name", name, "-q"]
            if state:
                cmd.extend(["--state", state])
            result = subprocess.run(cmd, capture_output=True, text=True, check=False)
            return [line.strip() for line in result.stdout.splitlines() if line.strip()]
        except Exception:
            return []

    def _get_component_pids(self, component: Optional[str]) -> Set[str]:
        """Wrapper around the shared PID tracker for compatibility."""
        # Safely handle None to satisfy callers that may pass Optional[str]
        if component is None:
            return set()
        return fetch_component_pids(component)

    def _get_crictl_component_name(self, component: Optional[str]) -> Optional[str]:
        aliases = {
            "kube-apiserver": "kube-apiserver",
            "kube-controller-manager": "kube-controller-manager",
            "kube-scheduler": "kube-scheduler",
            "etcd": "etcd"
        }
        # Avoid passing None to dict.get (static type checkers require str key)
        if component is None:
            return None
        return aliases.get(component)

    def _hard_kill_containers(self, component: Optional[str]) -> None:
        comp_name = self._get_crictl_component_name(component)
        if not comp_name:
            print(f"{Colors.YELLOW}[WARN] No crictl helper available for {component} to kill containers.{Colors.ENDC}")
            return
        before_pids = self._get_component_pids(component)
        container_ids = self._list_crictl_ids(comp_name)
        if container_ids:
            print(f"{Colors.YELLOW}[WARN] Performing HARD KILL on {component} to force config reload...{Colors.ENDC}")
            for cid in container_ids:
                subprocess.run(["crictl", "stop", cid], check=False)
                subprocess.run(["crictl", "rm", cid], check=False)
                print(f"[✓] Hard Killed container: {cid} to force config reload.")
        else:
            print(f"{Colors.YELLOW}[WARN] No containers found for {component}; skipping hard kill.{Colors.ENDC}")

        print(f"{Colors.YELLOW}[WARN] Waiting 30s for kubelet to launch {component}...{Colors.ENDC}")
        time.sleep(30)
        post_pids = self._get_component_pids(component)
        if not post_pids or post_pids.issubset(before_pids):
            print(f"{Colors.YELLOW}[WARN] {component} PID unchanged after hard kill; restarting kubelet as fallback...{Colors.ENDC}")
            subprocess.run(["systemctl", "restart", "kubelet"], check=False)
            print(f"{Colors.YELLOW}[WARN] Sleeping 30s after kubelet restart for {component}...{Colors.ENDC}")
            time.sleep(30)
            post_pids = self._get_component_pids(component)
            if post_pids:
                print(f"{Colors.GREEN}[✓] {component} restarted with new PID(s): {', '.join(sorted(post_pids))}.{Colors.ENDC}")
            else:
                print(f"{Colors.RED}[!] {component} still missing after kubelet restart!{Colors.ENDC}")
        else:
            print(f"{Colors.GREEN}[✓] {component} restarted with new PID(s): {', '.join(sorted(post_pids))}.{Colors.ENDC}")

    def _run_audit_verification_loop(self, script: Dict[str, Any], component: Optional[str]) -> Tuple[str, str]:
        audit_script_path = script['path'].replace("_remediate.sh", "_audit.sh")
        if not os.path.exists(audit_script_path):
            return "FAIL", "Audit script missing"

        cfg = self.get_remediation_config_for_check(script['id'])
        env = self._prepare_remediation_env(script['id'], cfg)
        comp_name = self._get_crictl_component_name(component)
        timeout = 60
        max_retries = 10

        def log_attempt(message: str) -> None:
            print(f"{Colors.YELLOW}[WARN] {message}{Colors.ENDC}")

        def audit_callable() -> Tuple[str, str]:
            audit_res = subprocess.run([
                "bash",
                audit_script_path,
            ], capture_output=True, text=True, timeout=self.script_timeout, env=env)
            status, reason, _, _ = self._parse_script_output(audit_res, script['id'], "audit", False)
            return status, reason

        status, reason = verify_with_retry(
            audit_callable,
            script['id'],
            component_name=comp_name,
            timeout=timeout,
            max_retries=max_retries,
            log_callback=log_attempt,
        )

        if status in ["PASS", "FIXED"]:
            print(f"{Colors.GREEN}[✓] Verified PASS for {script['id']}.{Colors.ENDC}")

        return status, reason

    def run_script(self, script: Dict[str, Any], mode: str) -> Optional[CISResult]:
        """
        Execute audit/remediation script with error handling.
        ดำเนินการสคริปต์ตรวจสอบ/การแก้ไขพร้อมการจัดการข้อผิดพลาด
        """
        if self.stop_requested:
            return None
        
        start_time = time.time()
        script_id = script["id"]
        
        # Check if rule is excluded / ตรวจสอบว่ากฎถูกยกเว้นหรือไม่
        if self.is_rule_excluded(script_id):
            return self._create_result(
                script, "IGNORED",
                f"Excluded: {self.excluded_rules.get(script_id, 'No reason')}",
                time.time() - start_time
            )
        
        try:
            # 1. REMEDIATION MODE / โหมดการแก้ไข
            if mode == "remediate":
                remediation_cfg = self.get_remediation_config_for_check(script_id)
                
                # Check if remediation is enabled / ตรวจสอบว่าเปิดใช้งานการแก้ไขหรือไม่
                if remediation_cfg.get("skip", False) or not remediation_cfg.get("enabled", True):
                    return self._create_result(
                        script, "SKIPPED",
                        f"Skipped by remediation config",
                        time.time() - start_time
                    )
                
                # Call internal remediation router / เรียกใช้ตัวเลือกการแก้ไขภายใน
                status, reason, fix_hint, cmds = self.fix_item_internal(script, remediation_cfg)
                duration = round(time.time() - start_time, 2)

                reason_text = reason or ""
                fix_hint_text = fix_hint or ""
                cmds_list = cmds or []

                result_record = self._create_result(script, status, reason_text, duration)
                result_record["fix_hint"] = fix_hint_text
                result_record["cmds"] = cmds_list
                result_record["output"] = reason_text
                return result_record

            # 2. AUDIT MODE / โหมดการตรวจสอบ
            is_manual = self._is_manual_check(script["path"])
            
            if is_manual and self.skip_manual:
                return self._create_result(
                    script, "SKIPPED",
                    "Manual check skipped by user",
                    time.time() - start_time
                )
            
            # Prepare environment variables for audit scripts / เตรียมตัวแปรสภาพแวดล้อมสำหรับสคริปต์ตรวจสอบ
            env = self._prepare_remediation_env(script_id, self.get_remediation_config_for_check(script_id))
            
            # Run script / รันสคริปต์
            result = subprocess.run(
                ["bash", script["path"]],
                capture_output=True,
                text=True,
                timeout=self.script_timeout,
                env=env
            )
            
            duration = round(time.time() - start_time, 2)
            
            # Parse output / แยกวิเคราะห์ผลลัพธ์
            status, reason, fix_hint, cmds = self._parse_script_output(
                result, script_id, mode, is_manual
            )
            
            # NEW: Deep Verification for Audit Mode (Config checks)
            if status == "PASS":
                process_name = self._get_process_name_for_check(script_id)
                remediation_cfg = self.get_remediation_config_for_check(script_id)
                
                # Check if it's a flag check / ตรวจสอบว่าเป็น flag check หรือไม่
                if process_name and remediation_cfg.get("check_type") == "flag_check":
                    flag = remediation_cfg.get("flag")
                    expected = remediation_cfg.get("required_value")
                    
                    if flag and expected:
                        # Use the new verify_runtime_config method
                        is_applied = self.verify_runtime_config(process_name, flag, expected)
                        runtime_status, runtime_msg = self.verify_process_runtime(process_name, flag, expected)
                        
                        if self.verbose >= 1:
                            print(f"    [INFO] File Config:   ✅ Correct ({flag}={expected})")
                            if is_applied:
                                print(f"    [INFO] Runtime State: ✅ VERIFIED ({runtime_msg})")
                            else:
                                print(f"    [INFO] Runtime State: ❌ {runtime_status} ({runtime_msg})")
                        
                        if not is_applied:
                            # Downgrade to FAIL if runtime doesn't match / ลดระดับเป็น FAIL หาก runtime ไม่ตรงกัน
                            status = "FAIL"
                            reason = f"[STALE_PROCESS] File is correct but runtime state is {runtime_status}: {runtime_msg}"
            
            # Handle silent script output / จัดการผลลัพธ์สคริปต์ที่ไม่มีข้อความ
            combined_output = result.stdout.strip() + result.stderr.strip()
            if not combined_output:
                if status == "PASS":
                    reason = "[INFO] Check completed successfully with no output"
                elif status == "FAIL":
                    reason = "[ERROR] Script failed silently without output"
                elif status == "MANUAL":
                    reason = "[WARN] Manual check completed with no output"
            
            reason_text = reason or ""
            fix_hint_text = fix_hint or ""
            cmds_list = cmds or []

            result_record = self._create_result(script, status, reason_text, duration)
            result_record["fix_hint"] = fix_hint_text
            result_record["cmds"] = cmds_list
            result_record["output"] = result.stdout + result.stderr
            return result_record
        
        except subprocess.TimeoutExpired:
            return self._create_result(
                script, "ERROR",
                f"Script timeout after {self.script_timeout}s",
                time.time() - start_time
            )
        except FileNotFoundError:
            return self._create_result(
                script, "ERROR",
                f"Script not found: {script['path']}",
                time.time() - start_time
            )
        except PermissionError:
            return self._create_result(
                script, "ERROR",
                "Permission denied executing script",
                time.time() - start_time
            )
        except Exception as e:
            return self._create_result(
                script, "ERROR",
                f"Unexpected error: {str(e)}",
                time.time() - start_time
            )

    def _run_etcd_remediation(self, script: Dict[str, Any], remediation_cfg: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        start_time = time.time()
        manifest_path = "/etc/kubernetes/manifests/etcd.yaml"
        temp_config_path = None
        combined_output = ""
        csv_checks = ["1.2.8", "1.2.11", "1.2.14", "1.2.29"]

        def build_result(status: str, reason: str, fix_hint: Optional[str] = None, output: str = "") -> Dict[str, Any]:
            return {
                "id": script["id"],
                "role": script.get("role", "master"),
                "level": script.get("level", "1"),
                "status": status,
                "duration": round(time.time() - start_time, 2),
                "reason": reason,
                "fix_hint": fix_hint,
                "cmds": [],
                "output": output,
                "path": script.get("path"),
                "component": self.get_component_for_rule(script["id"])
            }

        modifications: Dict[str, List[str]] = {"flags": []}
        flag_name = remediation_cfg.get("flag_name") if remediation_cfg else None
        expected_value = remediation_cfg.get("expected_value") if remediation_cfg else None

        if flag_name and expected_value is not None:
            if not flag_name.startswith("--"):
                flag_name = f"--{flag_name}"
            modifications["flags"].append(f"{flag_name}={expected_value}")

        # Safely obtain flags mapping and make its type explicit for static analyzers
        flags_dict_raw = remediation_cfg.get("flags") if remediation_cfg else None
        if isinstance(flags_dict_raw, dict):
            flags_dict = cast(Dict[str, Any], flags_dict_raw)
        else:
            flags_dict = {}

        for name, value in flags_dict.items():
            # Defensive: ensure key can be represented as a string
            try:
                name = str(name)
            except Exception:
                continue
            if not name.startswith("--"):
                name = f"--{name}"
            modifications["flags"].append(f"{name}={value}")

        if not modifications["flags"]:
            return build_result(
                "PASS",
                "[PASS] No etcd-specific flags configured for this check.",
                fix_hint="No changes needed."
            )

        mod_type = "csv" if script["id"] in csv_checks else "string"
        modifier = YAMLSafeModifier(manifest_path)
        modified_content = modifier.apply_modifications(modifications, mod_type)
        original_content = modifier.original_content or ""

        if not modified_content:
            return build_result(
                "FAIL",
                "[ERROR] Failed to render etcd manifest with requested flags.",
                fix_hint="Inspect /etc/kubernetes/manifests/etcd.yaml manually."
            )

        if modified_content == original_content:
            return build_result(
                "PASS",
                "[PASS] Etcd manifest already contains the desired flags.",
                fix_hint="No modification required."  # nothing to apply
            )

        try:
            with tempfile.NamedTemporaryFile("w", delete=False, suffix=".yaml") as temp_file:
                temp_file.write(modified_content)
                temp_config_path = temp_file.name

            timeout = int(self.remediation_global_config.get("etcd_timeout", 60))
            hardener_script = os.path.join(BASE_DIR, "modules", "etcd_hardener.py")
            cmd = [
                sys.executable,
                hardener_script,
                "--manifest", manifest_path,
                "--config-path", temp_config_path,
                "--backup-dir", str(self.backup_dir),
                "--timeout", str(timeout)
            ]
            proc = subprocess.run(cmd, capture_output=True, text=True)
            combined_output = (proc.stdout or "") + (proc.stderr or "")

            if proc.returncode == 0:
                reason = "[FIXED] Etcd manifest updated via etcd_hardener."
                status = "FIXED"
                fix_hint = "Etcd hardening helper applied the change."
                self.log_activity("ETCD_HARDENING", f"{script['id']}: success")
            else:
                reason = f"[ERROR] Etcd hardening helper failed (exit {proc.returncode})."
                status = "FAIL"
                fix_hint = "Inspect etcd_hardener output for details."
                self.log_activity("ETCD_HARDENING", f"{script['id']}: failure (exit {proc.returncode})")

            return build_result(status, reason, fix_hint, combined_output)

        except Exception as exc:
            combined_output = combined_output or str(exc)
            self.log_activity("ETCD_HARDENING", f"{script['id']}: exception {exc}")
            return build_result(
                "FAIL",
                f"[ERROR] Etcd hardening failed: {exc}",
                fix_hint="Run modules/etcd_hardener.py manually.",
                output=combined_output
            )

        finally:
            if temp_config_path and os.path.exists(temp_config_path):
                try:
                    os.remove(temp_config_path)
                except Exception:
                    pass

    def _get_backup_file_path(self, check_id: str, env: Optional[Dict[str, str]]) -> Optional[str]:
        """
        Identify backup file location for the manifest being remediated.
        ระบุตำแหน่งไฟล์สำรองข้อมูลสำหรับ manifest ที่กำลังถูกแก้ไข
        
        Priority 1: Check environment variable BACKUP_FILE / ลำดับความสำคัญ 1: ตรวจสอบตัวแปรสภาพแวดล้อม BACKUP_FILE
        Priority 2: Check standard backup path / ลำดับความสำคัญ 2: ตรวจสอบเส้นทางสำรองข้อมูลมาตรฐาน
        """
        if env is None:
            if self.verbose >= 1:
                print(f"{Colors.YELLOW}[!] Environment mapping is None when searching backups for {check_id}{Colors.ENDC}")
            return None

        # PRIORITY 1: Check environment variable / ตรวจสอบตัวแปรสภาพแวดล้อม
        backup_file = env.get("BACKUP_FILE")
        if backup_file and os.path.exists(backup_file):
            if self.verbose >= 1:
                print(f"{Colors.BLUE}[DEBUG] Found backup file from env: {backup_file}{Colors.ENDC}")
            return backup_file
        
        # PRIORITY 2: Search standard backup path / ค้นหาในเส้นทางสำรองข้อมูลมาตรฐาน
        backup_dir = env.get("BACKUP_DIR", "/var/backups/cis-remediation")
        if backup_dir and os.path.exists(backup_dir):
            # Find most recent backup for this check / ค้นหาการสำรองข้อมูลล่าสุดสำหรับการตรวจสอบนี้
            backup_pattern = os.path.join(backup_dir, f"{check_id}_*.bak")
            backups = sorted(glob.glob(backup_pattern), reverse=True)
            
            if backups:
                if self.verbose >= 1:
                    print(f"{Colors.BLUE}[DEBUG] Found backup file: {backups[0]}{Colors.ENDC}")
                return backups[0]
        
        if self.verbose >= 1:
            print(f"{Colors.YELLOW}[!] No backup file found for {check_id}{Colors.ENDC}")
        return None

    def _wait_for_api_healthy(self, check_id: str, timeout: int = 300) -> bool:
        """
        Wait for API Server to become healthy after remediation.
        รอให้ API Server กลับมาใช้งานได้ปกติหลังจากการแก้ไข
        """
        print(f"{Colors.CYAN}[*] Waiting for API Server to become healthy (timeout: {timeout}s)...{Colors.ENDC}")
        
        start_time = time.time()
        check_interval = 5  # Check every 5 seconds / ตรวจสอบทุก 5 วินาที
        
        while time.time() - start_time < timeout:
            elapsed = int(time.time() - start_time)
            try:
                # Use curl with -k (insecure) for localhost to avoid SSLError
                # ใช้ curl พร้อม -k (ไม่ปลอดภัย) สำหรับ localhost เพื่อหลีกเลี่ยง SSLError
                curl_cmd = ["curl", "-s", "-k", "-m", "5", "-w", "%{http_code}"]
                curl_cmd += ["https://127.0.0.1:6443/healthz"]
                
                result = subprocess.run(
                    curl_cmd,
                    capture_output=True,
                    text=True,
                    timeout=10
                )
                
                # Check if response is "ok" or status code is 200/401
                # ตรวจสอบว่าการตอบสนองคือ "ok" หรือรหัสสถานะคือ 200/401
                output = result.stdout.strip()
                http_code = output[-3:] if len(output) >= 3 else ""
                
                if result.returncode == 0 and ("ok" in output.lower() or http_code in ("200", "401")):
                    print(f"{Colors.GREEN}[✓] API Server healthy (Status: {http_code or 'OK'}) after {elapsed}s{Colors.ENDC}")
                    self.log_activity("API_HEALTH_OK", f"{check_id}: {elapsed}s")
                    return True
                elif http_code in ("500", "503"):
                    print(f"{Colors.YELLOW}    [*] API initializing... (Status: {http_code}, {elapsed}s/{timeout}s){Colors.ENDC}")
                elif result.returncode != 0:
                    print(f"{Colors.YELLOW}    [*] API restarting... (Connection Refused, {elapsed}s/{timeout}s){Colors.ENDC}")
                else:
                    print(f"{Colors.YELLOW}    [*] API responding with status {http_code}... ({elapsed}s/{timeout}s){Colors.ENDC}")
            
            except Exception as e:
                if self.verbose >= 2:
                    print(f"{Colors.BLUE}[DEBUG] API check attempt failed: {str(e)[:60]}...{Colors.ENDC}")
            
            # Wait before next attempt / รอก่อนการพยายามครั้งต่อไป
            time.sleep(check_interval)
        
        # Timeout reached / หมดเวลา
        print(f"{Colors.RED}[✗] API Server did not become healthy within {timeout}s{Colors.ENDC}")
        self.log_activity("API_HEALTH_TIMEOUT", f"{check_id}: No response after {timeout}s")
        return False

    def _rollback_manifest(self, check_id: str, backup_file: Optional[str]) -> bool:
        """
        Rollback manifest file to backup copy.
        ย้อนกลับไฟล์ manifest ไปยังชุดข้อมูลสำรอง
        """
        if not backup_file or not os.path.exists(backup_file):
            print(f"{Colors.RED}[✗] Backup file not found: {backup_file}{Colors.ENDC}")
            self.log_activity("ROLLBACK_NO_BACKUP", f"{check_id}: {backup_file}")
            return False
        
        # Determine original manifest path based on check ID / กำหนดเส้นทาง manifest เดิมตาม ID การตรวจสอบ
        if check_id.startswith("1.2"):
            original_path: str = "/etc/kubernetes/manifests/kube-apiserver.yaml"
        elif check_id.startswith("2."):
            original_path = "/etc/kubernetes/manifests/etcd.yaml"
        elif check_id.startswith("4."):
            original_path = "/var/lib/kubelet/config.yaml"
        else:
            # backup_file is Optional[str], guard above ensures it's not None here
            original_path = backup_file.replace(".bak", "")
        
        try:
            print(f"{Colors.YELLOW}[*] Rolling back: {original_path}{Colors.ENDC}")
            print(f"    From backup: {backup_file}")
            
            # Create backup of current broken state before rollback / สร้างการสำรองข้อมูลของสถานะที่เสียก่อนการย้อนกลับ
            if os.path.exists(original_path):
                broken_backup = f"{original_path}.broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                shutil.copy2(original_path, broken_backup)
                print(f"    Saved broken config: {broken_backup}")
            
            # Restore original file / กู้คืนไฟล์เดิม
            shutil.copy2(backup_file, original_path)
            
            print(f"{Colors.GREEN}[✓] Rollback completed successfully{Colors.ENDC}")
            self.log_activity("ROLLBACK_SUCCESS", f"{check_id}: {original_path}")
            
            # Wait briefly for manifest reload / รอสักครู่เพื่อให้ manifest โหลดใหม่
            time.sleep(2)
            
            return True
        
        except FileNotFoundError as e:
            print(f"{Colors.RED}[✗] Rollback failed - file not found: {str(e)}{Colors.ENDC}")
            self.log_activity("ROLLBACK_FILE_NOT_FOUND", f"{check_id}: {str(e)}")
            return False
        
        except PermissionError as e:
            print(f"{Colors.RED}[✗] Rollback failed - permission denied: {str(e)}{Colors.ENDC}")
            self.log_activity("ROLLBACK_PERMISSION_DENIED", f"{check_id}: {str(e)}")
            return False
        
        except Exception as e:
            print(f"{Colors.RED}[✗] Rollback failed - unexpected error: {str(e)}{Colors.ENDC}")
            self.log_activity("ROLLBACK_ERROR", f"{check_id}: {str(e)}")
            return False

    def update_manifest_safely(self, filepath: str, key: str, value: str) -> Tuple[bool, str]:
        """
        Atomically update a manifest file using safe copy-paste pattern.
        อัปเดตไฟล์ manifest แบบ atomic โดยใช้รูปแบบการคัดลอกและวางที่ปลอดภัย
        
        ATOMIC COPY-PASTE MODIFIER / ตัวแก้ไขแบบ ATOMIC:
        1. Read original file content / อ่านเนื้อหาไฟล์เดิม
        2. Modify specific line / แก้ไขบรรทัดที่ระบุ
        3. Write to temporary file / เขียนลงไฟล์ชั่วคราว
        4. Atomic overwrite / เขียนทับแบบ atomic
        5. Preserve indentation / รักษาระยะย่อหน้า
        """
        # ===== STEP 1: Validation & Read / ขั้นตอนที่ 1: การตรวจสอบและการอ่าน =====
        if not os.path.exists(filepath):
            msg = f"[ERROR] Manifest file not found: {filepath}"
            print(f"{Colors.RED}{msg}{Colors.ENDC}")
            return False, msg
        
        if not os.access(filepath, os.R_OK):
            msg = f"[ERROR] Cannot read manifest file: {filepath}"
            print(f"{Colors.RED}{msg}{Colors.ENDC}")
            return False, msg
        
        try:
            # Read original file content / อ่านเนื้อหาไฟล์เดิม
            with open(filepath, 'r', encoding='utf-8') as f:
                original_content = f.read()
                original_lines = original_content.split('\n')
            
            if self.verbose >= 2:
                print(f"{Colors.BLUE}[DEBUG] Read manifest: {filepath} ({len(original_lines)} lines){Colors.ENDC}")
            
            # ===== STEP 2-4: Parse and Modify / ขั้นตอนที่ 2-4: การแยกวิเคราะห์และการแก้ไข =====
            modified_lines: List[str] = []
            changes_made = 0
            found_key = False
            in_command_section = False
            command_indent = None
            
            for i, line in enumerate(original_lines):
                # Detect command section / ตรวจจับส่วนของคำสั่ง
                if line.strip().startswith(('command:', 'args:')):
                    in_command_section = True
                    command_indent = len(line) - len(line.lstrip())
                    modified_lines.append(line)
                    if self.verbose >= 2:
                        print(f"{Colors.BLUE}[DEBUG] Found command section at line {i+1}, indent={command_indent}{Colors.ENDC}")
                    continue
                
                # If in command section, look for the key to modify / หากอยู่ในส่วนคำสั่ง ให้ค้นหาคีย์ที่จะแก้ไข
                if in_command_section:
                    current_indent = len(line) - len(line.lstrip()) if line.strip() else 0
                    
                    # Exit command section if we dedent back to original level / ออกจากส่วนคำสั่งหากระยะย่อหน้ากลับมาเท่าเดิม
                    if line.strip() and command_indent is not None and current_indent <= command_indent:
                        in_command_section = False
                        command_indent = None
                    
                    # Look for key in this line / ค้นหาคีย์ในบรรทัดนี้
                    if key in line:
                        # Found the key - modify the value / พบคีย์ - แก้ไขค่า
                        indent = len(line) - len(line.lstrip())
                        stripped = line.strip()
                        
                        # List of flags that should be treated as CSV lists / รายการ flag ที่ควรจัดการเป็นรายการ CSV
                        CSV_FLAGS = [
                            "--authorization-mode", 
                            "--enable-admission-plugins", 
                            "--disable-admission-plugins", 
                            "--tls-cipher-suites"
                        ]
                        
                        if '=' in stripped:
                            # Split by key to get prefix / แยกด้วยคีย์เพื่อรับคำนำหน้า
                            parts = stripped.split(key, 1)
                            prefix = parts[0] + key
                            
                            if key in CSV_FLAGS:
                                # Smart Append: Don't overwrite, append to CSV list if not present
                                # การเพิ่มแบบอัจฉริยะ: ไม่เขียนทับ แต่เพิ่มลงในรายการ CSV หากยังไม่มี
                                current_val_str = parts[1].lstrip('=')
                                existing_values = [v.strip() for v in current_val_str.split(',') if v.strip()]
                                
                                if value not in existing_values:
                                    existing_values.append(value)
                                    new_value = ','.join(existing_values)
                                    modified_line = ' ' * indent + prefix + '=' + new_value
                                    changes_made += 1
                                else:
                                    # Value already exists, no change needed / มีค่าอยู่แล้ว ไม่จำเป็นต้องเปลี่ยนแปลง
                                    modified_line = line
                            else:
                                # Normal overwrite for non-CSV flags / การเขียนทับปกติสำหรับ flag ที่ไม่ใช่ CSV
                                modified_line = ' ' * indent + prefix + value
                                changes_made += 1
                        else:
                            # No '=' in line, just append it / ไม่มี '=' ในบรรทัด ให้เพิ่มเข้าไปเลย
                            modified_line = ' ' * indent + stripped + '=' + value
                            changes_made += 1
                        
                        modified_lines.append(modified_line)
                        found_key = True
                        
                        if self.verbose >= 1 and modified_line != line:
                            print(f"{Colors.YELLOW}[*] Modified line {i+1}: {key}={value} (Smart Append){Colors.ENDC}")
                        continue
                
                # No change needed for this line
                modified_lines.append(line)
            
            # ===== STEP 5: Append if key not found =====
            if not found_key and in_command_section:
                # Key not found - append to command section
                print(f"{Colors.YELLOW}[*] Key '{key}' not found in command section - appending...{Colors.ENDC}")
                
                # Find where to append (end of command list)
                for i in range(len(modified_lines) - 1, -1, -1):
                    line = modified_lines[i]
                    if line.strip() and (line.strip().startswith('-') or 
                                         (i > 0 and modified_lines[i-1].strip().startswith(('command:', 'args:')))):
                        # Found last item in list
                        list_item_indent = len(line) - len(line.lstrip())
                        new_item = ' ' * list_item_indent + f'- {key}{value}'
                        modified_lines.insert(i + 1, new_item)
                        changes_made += 1
                        if self.verbose >= 1:
                            print(f"{Colors.YELLOW}[*] Appended line {i+2}: {key}{value}{Colors.ENDC}")
                        break
            
            # ===== STEP 6: Create backup before writing =====
            backup_path = f"{filepath}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                shutil.copy2(filepath, backup_path)
                if self.verbose >= 1:
                    print(f"{Colors.CYAN}[✓] Backup created: {backup_path}{Colors.ENDC}")
            except Exception as e:
                msg = f"[ERROR] Failed to create backup: {str(e)}"
                print(f"{Colors.RED}{msg}{Colors.ENDC}")
                return False, msg
            
            # ===== STEP 7: Write to temporary file =====
            temp_file = f"{filepath}.tmp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            modified_content = '\n'.join(modified_lines)
            
            # Visual Diff Generation / การสร้างความแตกต่างของไฟล์ด้วยสายตา
            print(f"\n{Colors.BOLD}[DIFF] {filepath}:{Colors.ENDC}")
            diff = difflib.unified_diff(
                original_content.splitlines(),
                modified_content.splitlines(),
                fromfile='original',
                tofile='modified',
                lineterm=''
            )
            
            diff_summary: List[str] = []
            for line in diff:
                if line.startswith('---') or line.startswith('+++') or line.startswith('@@'):
                    continue
                if line.startswith('+'):
                    formatted_line = f"[+] {line[1:]}"
                    print(f"{Colors.GREEN}{formatted_line}{Colors.ENDC}")
                    diff_summary.append(formatted_line)
                elif line.startswith('-'):
                    formatted_line = f"[-] {line[1:]}"
                    print(f"{Colors.RED}{formatted_line}{Colors.ENDC}")
                    diff_summary.append(formatted_line)
            
            diff_text = "\n".join(diff_summary)
            if diff_text:
                self.log_activity("MANIFEST_DIFF", f"Changes for {filepath}:\n{diff_text}")

            try:
                with open(temp_file, 'w', encoding='utf-8') as f:
                    f.write(modified_content)
            except Exception as e:
                msg = f"[ERROR] Failed to write temporary file: {str(e)}"
                print(f"{Colors.RED}{msg}{Colors.ENDC}")
                if os.path.exists(temp_file): os.unlink(temp_file)
                return False, msg
            
            # ===== STEP 8: Atomic replace (CRITICAL) =====
            try:
                os.replace(temp_file, filepath)
                
                # Post-Write Read-Back Verification (Double Check)
                with open(filepath, 'r', encoding='utf-8') as f:
                    verified_content = f.read()
                
                if key in verified_content and value in verified_content:
                    print(f"{Colors.GREEN}[SUCCESS] Verified change on disk: {key}={value}{Colors.ENDC}")
                else:
                    msg = f"[ERROR] Panic! File write reported success but content verification failed."
                    print(f"{Colors.RED}{msg}{Colors.ENDC}")
                    # Trigger Immediate Rollback
                    shutil.copy2(backup_path, filepath)
                    return False, msg

                success_msg = f"[SUCCESS] Manifest updated atomically with {changes_made} change(s)"
                print(f"{Colors.GREEN}[✓] {success_msg}{Colors.ENDC}")
                self.log_activity("MANIFEST_UPDATE_SUCCESS", f"{filepath}: {changes_made} change(s), backup: {backup_path}")
                
                return True, success_msg
            
            except OSError as e:
                msg = f"[ERROR] Atomic replace failed: {str(e)}"
                print(f"{Colors.RED}{msg}{Colors.ENDC}")
                if os.path.exists(temp_file): os.unlink(temp_file)
                return False, msg
        
        except Exception as e:
            msg = f"[ERROR] Unexpected error: {str(e)}"
            print(f"{Colors.RED}{msg}{Colors.ENDC}")
            return False, msg

    def apply_remediation_with_health_gate(self, filepath: str, key: str, value: str, check_id: str, script_dict: Dict[str, Any], timeout: int = 60) -> Dict[str, Any]:
        """
        Apply remediation using atomic modifier with health-gated rollback.
        
        HEALTH-GATED ROLLBACK FLOW:
        1. Backup: Archive current filepath to filepath.bak
        2. Apply: Call update_manifest_safely() for atomic modification
        3. Wait: Loop check https://127.0.0.1:6443/healthz for up to 'timeout' seconds
        4. Decision:
           - IF Unhealthy (Timeout):
             * Log: [CRITICAL] API Server failed to restart. Rolling back...
             * Restore: Copy backup_path back to filepath
             * Return: False (Fail)
           - IF Healthy:
             * Run audit check immediately
             * If Audit Pass: Return True (Success)
             * If Audit Fail: Rollback and Return False (Config invalid)
        
        Args:
            filepath (str): Path to manifest file
            key (str): Key-value pair to modify
            value (str): Value for the key
            check_id (str): CIS check ID for logging
            script_dict (dict): Script metadata dict (contains 'path' for audit script location)
            timeout (int): Seconds to wait for API health (default: 60)
        
        Returns:
            dict: {
                'success': bool,
                'status': str ('FIXED', 'REMEDIATION_FAILED', 'REMEDIATION_PARTIAL'),
                'reason': str,
                'backup_path': str,
                'audit_verified': bool
            }
        """
        result = {
            'success': False,
            'status': 'REMEDIATION_FAILED',
            'reason': '',
            'backup_path': None,
            'audit_verified': False
        }
        
        print(f"\n{Colors.BOLD}{Colors.CYAN}{'='*70}")
        print(f"[HEALTH-GATED REMEDIATION] {check_id}")
        print(f"{'='*70}{Colors.ENDC}")
        
        # ========== STEP 1: Backup ==========
        print(f"{Colors.CYAN}[STEP 1/4] Creating backup...{Colors.ENDC}")
        backup_path = f"{filepath}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            if not os.path.exists(filepath):
                result['reason'] = f"[ERROR] Manifest not found: {filepath}"
                print(f"{Colors.RED}{result['reason']}{Colors.ENDC}")
                return result
            
            shutil.copy2(filepath, backup_path)
            result['backup_path'] = backup_path
            print(f"{Colors.GREEN}[✓] Backup created: {backup_path}{Colors.ENDC}")
            self.log_activity("REMEDIATION_BACKUP_CREATED", f"{check_id}: {backup_path}")
        
        except Exception as e:
            result['reason'] = f"[ERROR] Failed to create backup: {str(e)}"
            print(f"{Colors.RED}{result['reason']}{Colors.ENDC}")
            return result
        
        # ========== STEP 2: Apply Atomic Modification ==========
        print(f"{Colors.CYAN}[STEP 2/4] Applying atomic modification...{Colors.ENDC}")
        success, msg = self.update_manifest_safely(filepath, key, value)
        
        if not success:
            result['reason'] = f"[ERROR] Atomic modification failed: {msg}"
            print(f"{Colors.RED}{result['reason']}{Colors.ENDC}")
            return result
        
        print(f"{Colors.GREEN}[✓] Manifest modified atomically.{Colors.ENDC}")
        self.log_activity("REMEDIATION_APPLIED", f"{check_id}: {msg}")
        
        # ========== STEP 3: Wait for Health Check ==========
        print(f"{Colors.CYAN}[STEP 3/4] Waiting for API Server health check (timeout: {timeout}s)...{Colors.ENDC}")
        api_healthy = self._wait_for_api_healthy(check_id, timeout=timeout)
        
        if not api_healthy:
            # ========== FAILURE: ROLLBACK ==========
            print(f"{Colors.RED}[✗] CRITICAL: API Server failed to restart within {timeout}s.{Colors.ENDC}")
            print(f"{Colors.RED}[!] Triggering automatic rollback...{Colors.ENDC}")
            
            rollback_ok = self._rollback_manifest(check_id, backup_path)
            
            if rollback_ok:
                print(f"{Colors.YELLOW}[✓] Rollback succeeded. Waiting for cluster recovery...{Colors.ENDC}")
                time.sleep(5)
                self.wait_for_healthy_cluster(skip_health_check=False)
                result['status'] = 'REMEDIATION_FAILED_ROLLED_BACK'
                result['reason'] = (
                    f"[CRITICAL] API Server failed to restart. "
                    f"Automatic rollback succeeded. Manual investigation required."
                )
            else:
                result['status'] = 'REMEDIATION_FAILED_ROLLBACK_FAILED'
                result['reason'] = (
                    f"[CRITICAL] API Server failed to restart AND rollback failed! "
                    f"MANUAL INTERVENTION REQUIRED immediately!"
                )
                print(f"{Colors.RED}{Colors.BOLD}[!!!] EMERGENCY: Rollback failed!{Colors.ENDC}")
            
            self.log_activity("REMEDIATION_HEALTH_FAILED", f"{check_id}: Rollback {'success' if rollback_ok else 'FAILED'}")
            return result
        
        # ========== SUCCESS: API is Healthy - Verify with Audit ==========
        print(f"{Colors.GREEN}[✓] API Server became healthy. Running deep verification...{Colors.ENDC}")
        
        # ========== NEW: Deep Verification (Runtime State) ==========
        process_name = self._get_process_name_for_check(check_id)
        if process_name:
            print(f"{Colors.CYAN}[INFO] File Config:   ✅ Correct ({key}={value}){Colors.ENDC}")

            # Check if runtime matches / ตรวจสอบว่า runtime ตรงกันหรือไม่
            if not self.verify_runtime_config(process_name, key, value):
                # STALE PROCESS DETECTED / ตรวจพบโปรเซสที่ค้างอยู่
                print(f"{Colors.YELLOW}[WARN] Configuration updated on disk but process is stale.{Colors.ENDC}")
                self._aggressive_restart_component(process_name)

                # Re-verify after force reload / ตรวจสอบอีกครั้งหลังบังคับโหลดใหม่
                print(f"{Colors.CYAN}[*] Re-verifying runtime state after forced reload...{Colors.ENDC}")
                time.sleep(5)

                # If still stale, log warning / หากยังคงค้างอยู่ ให้เตือนผู้ใช้
                if not self.verify_runtime_config(process_name, key, value):
                    print(f"{Colors.YELLOW}[WARN] Runtime still stale after forced restart.{Colors.ENDC}")
            
            runtime_status, runtime_msg = self.verify_process_runtime(process_name, key, value)
            
            if runtime_status == "VERIFIED":
                print(f"{Colors.GREEN}[INFO] Runtime State: ✅ {runtime_status} ({runtime_msg}){Colors.ENDC}")
            else:
                print(f"{Colors.RED}[INFO] Runtime State: ❌ {runtime_status} ({runtime_msg}){Colors.ENDC}")
                print(f"{Colors.RED}[!] Configuration not reflected in runtime. Triggering rollback...{Colors.ENDC}")
                
                rollback_ok = self._rollback_manifest(check_id, backup_path)
                
                if rollback_ok:
                    print(f"{Colors.YELLOW}[✓] Rollback succeeded. Waiting for cluster recovery...{Colors.ENDC}")
                    time.sleep(5)
                    self.wait_for_healthy_cluster(skip_health_check=False)
                    result['status'] = 'REMEDIATION_FAILED_ROLLED_BACK'
                    result['reason'] = f"[STALE_PROCESS] {runtime_msg}. Rolled back for safety."
                else:
                    result['status'] = 'REMEDIATION_FAILED_ROLLBACK_FAILED'
                    result['reason'] = f"[CRITICAL] Stale process AND rollback failed!"
                
                self.log_activity("REMEDIATION_RUNTIME_FAILED", f"{check_id}: {runtime_msg}")
                return result

        # ========== STEP 4: Audit Verification ==========
        print(f"{Colors.CYAN}[STEP 4/4] Running audit verification...{Colors.ENDC}")
        
        audit_script_path = script_dict.get("path", "").replace("_remediate.sh", "_audit.sh")
        
        if not os.path.exists(audit_script_path):
            print(f"{Colors.YELLOW}[!] Audit script not found: {audit_script_path}{Colors.ENDC}")
            print(f"{Colors.YELLOW}[!] Skipping audit verification (proceeding with health check success){Colors.ENDC}")
            
            result['success'] = True
            result['status'] = 'FIXED'
            result['reason'] = f"[FIXED] Remediation applied and API healthy (audit script not found)"
            result['audit_verified'] = False
            self.log_activity("REMEDIATION_SUCCESS_NO_AUDIT", check_id)
            return result
        
        try:
            # Wait for config propagation
            time.sleep(2)
            
            # Run audit script
            env = os.environ.copy()
            audit_result = subprocess.run(
                ["bash", audit_script_path],
                capture_output=True,
                text=True,
                timeout=self.script_timeout,
                env=env
            )
            
            # Parse audit result
            audit_status, audit_reason, _, _ = self._parse_script_output(
                audit_result, check_id, "audit", False
            )
            
            if audit_status == "PASS":
                # ========== SUCCESS: All checks passed ==========
                print(f"{Colors.GREEN}[✓] VERIFIED: Audit check PASSED{Colors.ENDC}")
                print(f"{Colors.GREEN}{'='*70}")
                print(f"[SUCCESS] Remediation complete and verified")
                print(f"{'='*70}{Colors.ENDC}")
                
                result['success'] = True
                result['status'] = 'FIXED'
                result['reason'] = f"[FIXED] Remediation applied, API healthy, and audit verified"
                result['audit_verified'] = True
                
                self.log_activity("REMEDIATION_SUCCESS_VERIFIED", check_id)
                return result
            
            else:
                # ========== FAILURE: Audit failed despite healthy API ==========
                print(f"{Colors.RED}[✗] AUDIT FAILED: API healthy but audit verification failed.{Colors.ENDC}")
                print(f"{Colors.RED}[!] Audit reason: {audit_reason}{Colors.ENDC}")
                print(f"{Colors.RED}[!] Triggering automatic rollback...{Colors.ENDC}")
                
                rollback_ok = self._rollback_manifest(check_id, backup_path)
                
                if rollback_ok:
                    print(f"{Colors.YELLOW}[✓] Rollback succeeded. Waiting for cluster recovery...{Colors.ENDC}")
                    time.sleep(5)
                    self.wait_for_healthy_cluster(skip_health_check=False)
                    result['status'] = 'REMEDIATION_FAILED_ROLLED_BACK'
                    result['reason'] = (
                        f"[REMEDIATION_FAILED] Audit verification failed: {audit_reason}. "
                        f"Automatic rollback succeeded."
                    )
                else:
                    result['status'] = 'REMEDIATION_FAILED_ROLLBACK_FAILED'
                    result['reason'] = (
                        f"[CRITICAL] Audit verification failed AND rollback failed! "
                        f"MANUAL INTERVENTION REQUIRED!"
                    )
                    print(f"{Colors.RED}{Colors.BOLD}[!!!] EMERGENCY: Rollback failed!{Colors.ENDC}")
                
                self.log_activity("REMEDIATION_AUDIT_FAILED", 
                                  f"{check_id}: {audit_reason}, Rollback {'success' if rollback_ok else 'FAILED'}")
                return result
        
        except subprocess.TimeoutExpired:
            print(f"{Colors.YELLOW}[!] Audit verification TIMEOUT for {check_id}{Colors.ENDC}")
            result['status'] = 'REMEDIATION_FAILED'
            result['reason'] = "[REMEDIATION_FAILED] Audit verification timed out"
            self.log_activity("REMEDIATION_AUDIT_TIMEOUT", check_id)
            return result
        
        except Exception as e:
            print(f"{Colors.RED}[✗] Unexpected error during audit verification: {str(e)}{Colors.ENDC}")
            result['status'] = 'REMEDIATION_FAILED'
            result['reason'] = f"[REMEDIATION_FAILED] Unexpected audit error: {str(e)}"
            self.log_activity("REMEDIATION_AUDIT_ERROR", f"{check_id}: {str(e)}")
            return result

    def _is_manual_check(self, script_path: Optional[str]) -> bool:
        """Check if script is marked as manual / ตรวจสอบว่าสคริปต์ถูกทำเครื่องหมายเป็นด้วยตนเองหรือไม่"""
        if not script_path:
            return False
        try:
            with open(script_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f.readlines()[:10]:
                    if "# Title:" in line and "(Manual)" in line:
                        return True
        except Exception:
            # Any file error means we cannot determine; treat as non-manual to avoid blocking automation
            return False
        return False

    def _extract_pod_security_failure_names(self, combined_output: str) -> Set[str]:
        pattern = re.compile(r"^-+\s*([a-z0-9][a-z0-9\.\-]*)$", re.IGNORECASE)
        failures: Set[str] = set()
        for line in combined_output.splitlines():
            stripped = line.strip()
            match = pattern.match(stripped)
            if match:
                failures.add(match.group(1).lower())
        return failures

    def _parse_script_output(self, result: subprocess.CompletedProcess[str], script_id: str, mode: str, is_manual: bool) -> ResultTuple:
        """
        Parse structured output from script with Smart Override for MANUAL checks
        แยกวิเคราะห์ผลลัพธ์ที่มีโครงสร้างจากสคริปต์
        
        Algorithm / อัลกอริทึม:
        - STEP 1: Initialize fields / เริ่มต้นฟิลด์
        - STEP 2: Parse structured tags / แยกแท็ก
        - STEP 3: SMART OVERRIDE - Check if MANUAL checks have explicit PASS/FAIL in output
        - STEP 4: PRIORITY 1 - Check exit code 3 for MANUAL / ตรวจสอบรหัสออก 3 สำหรับด้วยตนเอง
        - STEP 5: PRIORITY 2 - Determine status based on return code / กำหนดสถานะตามรหัสคืน
        - STEP 6: PRIORITY 3 - Fallback to text-based detection / ใช้ตัวตรวจจับตามข้อความ
        - STEP 7: Return result / คืนผลลัพธ์
        
        SMART OVERRIDE: If a MANUAL check contains [PASS] in output, override to PASS.
        If it contains [FAIL], override to FAIL. Only keep MANUAL if explicitly marked or exit code = 3.
        This allows automation to confirm compliance for checks that can be verified programmatically.
        """
        status = "FAIL"
        reason = ""
        fix_hint = ""
        cmds: List[str] = []
        
        # Parse structured tags from stdout / แยกแท็กจาก stdout
        for line in result.stdout.split('\n'):
            if "[FAIL_REASON]" in line:
                reason = line.split("[FAIL_REASON]", 1)[1].strip()
            elif "[FIX_HINT]" in line:
                fix_hint = line.split("[FIX_HINT]", 1)[1].strip()
            elif "[CMD]" in line:
                cmds.append(line.split("[CMD]", 1)[1].strip())
        
        # Prepare combined output for keyword detection / เตรียม output รวมสำหรับการตรวจหาคำสำคัญ
        combined_output = result.stdout + result.stderr
        manual_keywords = {"manual remediation", "manual intervention", "requires manual", "check requires manual"}
        is_manual_check = any(kw in combined_output.lower() for kw in manual_keywords)

        # ========== STEP 3: SMART OVERRIDE FOR MANUAL CHECKS ==========
        # If check is marked as MANUAL, but script output shows explicit PASS/FAIL/MANUAL,
        # override the status based on the script output. This allows automation to confirm
        # compliance for checks that CAN be verified programmatically.
        # 
        # Logic: Some checks are marked MANUAL but the underlying script can still verify compliance.
        # Example: 5.6.4 (Default Namespace) - Script can check if namespace is empty.
        # If script confirms PASS, override MANUAL to PASS (automation did the verification).
        # If script confirms FAIL, override MANUAL to FAIL (automation found the issue).
        # Only keep MANUAL if script explicitly says [MANUAL] or exit code = 3.
        if is_manual:
            # Check for explicit PASS in output (Smart Override)
            if "[PASS]" in combined_output:
                status = "PASS"
                if not reason:
                    reason = "[INFO] Manual check confirmed PASS by automation"
                if self.verbose >= 2:
                    print(f"{Colors.GREEN}[DEBUG] SMART OVERRIDE: {script_id} - [PASS] found in output, overriding MANUAL to PASS{Colors.ENDC}")
                return status, reason, fix_hint, cmds
            
            # Check for explicit FAIL in output (Smart Override)
            elif "[FAIL]" in combined_output:
                status = "FAIL"
                if not reason:
                    reason = "[INFO] Manual check confirmed FAIL by automation"
                if self.verbose >= 2:
                    print(f"{Colors.RED}[DEBUG] SMART OVERRIDE: {script_id} - [FAIL] found in output, overriding MANUAL to FAIL{Colors.ENDC}")
                return status, reason, fix_hint, cmds
            
            # Check for explicit MANUAL in output (keep as MANUAL)
            elif "[MANUAL]" in combined_output or result.returncode == 3:
                status = "MANUAL"
                if not reason:
                    if "[MANUAL]" in combined_output:
                        reason = "[INFO] Script output indicates manual verification required"
                    else:
                        reason = "[INFO] Script returned exit code 3 (Manual Intervention Required)"
                if self.verbose >= 2:
                    print(f"{Colors.YELLOW}[DEBUG] MANUAL CHECK: {script_id} - Explicitly marked MANUAL (output or exit code){Colors.ENDC}")
                return status, reason, fix_hint, cmds
            
            # No explicit status in output - use default MANUAL enforcement
            status = "MANUAL"
            if not reason:
                # Provide context about script execution
                if result.returncode == 0:
                    reason = "[INFO] Manual check executed successfully. Human verification required before marking as resolved."
                elif result.returncode == 3:
                    reason = "[INFO] Script returned exit code 3 (Manual Intervention Required)"
                else:
                    reason = "[INFO] Manual check requires human verification."
            
            if self.verbose >= 2:
                print(f"{Colors.BLUE}[DEBUG] MANUAL ENFORCEMENT: {script_id} - No explicit status in output, keeping as MANUAL{Colors.ENDC}")
            
            return status, reason, fix_hint, cmds
        
        # ========== STEP 4: PRIORITY 1 - Check specific exit codes ==========
        # Exit code 3 is standardized as "Manual Intervention Required"
        # Only applied to NON-MANUAL checks (since manual checks already handled above)
        # ตรวจสอบรหัสออกเฉพาะ - รหัส 3 ถูกกำหนดเป็น "ต้องการการแทรกแซงด้วยตนเอง"
        if result.returncode == 3:
            status = "MANUAL"
            if not reason:
                reason = "[INFO] Script returned exit code 3 (Manual Intervention Required)"
            if self.verbose >= 2:
                print(f"{Colors.BLUE}[DEBUG] Exit code 3 detected for {script_id} - Setting status to MANUAL{Colors.ENDC}")
        
        # ========== STEP 5: PRIORITY 2 - Check for success (0) or other failures ==========
        elif result.returncode == 0:
            # Success case / กรณีสำเร็จ
            # This only applies to automated (non-manual) checks
            if mode == "remediate":
                status = "FIXED" if "fixed" in combined_output.lower() else "PASS"
            else:  # mode == "audit"
                status = "PASS"
        
        # ========== STEP 6: PRIORITY 3 - Fallback to text-based detection for non-zero exit codes ==========
        # For audit mode, check if output contains manual intervention keywords
        # สำหรับโหมดการตรวจสอบ ตรวจสอบว่า output มีคำสำคัญเกี่ยวกับการแทรกแซง
        else:
            # Return code is non-zero (failure)
            if mode == "remediate":
                status = "FAIL"
            else:  # mode == "audit"
                # Check if failure reason is manual intervention (fallback detection)
                if is_manual_check:
                    status = "MANUAL"
                else:
                    status = "FAIL"
                    if not reason:
                        lines = result.stdout.split('\n')
                        reason = next((l for l in lines if l.strip()), "Check failed")
        
        # Generate default fix hint / สร้างคำแนะนำการแก้ไขเริ่มต้น
        if status == "FAIL" and not fix_hint:
            remediate_path = result.stdout.replace("audit", "remediate")
            if os.path.exists(remediate_path):
                fix_hint = f"Run: sudo bash {remediate_path}"

        if script_id.startswith("5.2."):
            combined = f"{result.stdout}\n{result.stderr}"
            failures = self._extract_pod_security_failure_names(combined)
            if "kube-flannel" in failures and self._is_kube_flannel_privileged():
                failures.discard("kube-flannel")
                suffix = "[INFO] kube-flannel privileges are intentionally excluded from enforcement."
                if failures and status == "FAIL":
                    reason = f"{reason} {suffix}".strip() if reason else suffix
                elif not failures and status == "FAIL":
                    status = "PASS"
                    reason = reason or suffix
                else:
                    reason = reason or suffix

        return status, reason, fix_hint, cmds

    def _create_result(self, script: Dict[str, Any], status: str, reason: str, duration: float) -> CISResult:
        """Create result dictionary / สร้างพจนานุกรมผลลัพธ์"""
        # Use safe .get lookups to avoid KeyError and provide stable types for static analysis
        return {
            "id": str(script.get("id", "")),
            "role": str(script.get("role", "")),
            "level": str(script.get("level", "")),
            "status": status,
            "duration": round(float(duration), 2),
            "reason": reason,
            "fix_hint": "",
            "cmds": [],
            "output": "",
            "path": str(script.get("path", "")),
            "component": self.get_component_for_rule(str(script.get("id", "")))
        }

    def _coerce_cis_result(self, raw: Any) -> Optional[CISResult]:
        """
        Normalize any dict-like result into a CISResult so callers can safely append to self.results.
        This prevents type checkers from seeing `Dict[str, Any] | CISResult` at append sites.
        """
        if raw is None:
            return None

        if not isinstance(raw, dict):
            return None

        raw_dict = cast(Dict[str, Any], raw)

        check_id_obj = raw_dict.get("id")
        check_id = str(check_id_obj) if check_id_obj is not None else ""
        try:
            duration_val = float(raw_dict.get("duration", 0.0))
        except Exception:
            duration_val = 0.0

        cmds_raw = raw_dict.get("cmds", [])
        cmds_list: List[str] = []
        if isinstance(cmds_raw, list):
            cmds_raw_list = cast(List[Any], cmds_raw)
            for item in cmds_raw_list:
                try:
                    cmds_list.append(str(item))
                except Exception:
                    continue

        component_raw = raw_dict.get("component")
        component = (
            str(component_raw)
            if component_raw not in (None, "")
            else self.get_component_for_rule(check_id)
        )

        return cast(CISResult, {
            "id": check_id,
            "role": str(raw_dict.get("role", "")),
            "level": str(raw_dict.get("level", "")),
            "status": str(raw_dict.get("status", "")),
            "duration": round(duration_val, 2),
            "reason": str(raw_dict.get("reason", "")),
            "fix_hint": str(raw_dict.get("fix_hint", "")),
            "cmds": cmds_list,
            "output": str(raw_dict.get("output", "")),
            "path": str(raw_dict.get("path", "")),
            "component": component,
        })

    def update_stats(self, result: Optional[CISResult]) -> None:
        """
        Update statistics based on result and keep audit_results fresh.
        อัปเดตสถิติตามผลลัพธ์และรักษาข้อมูล audit_results ให้เป็นปัจจุบัน
        """
        if not result:
            return
        
        role = "master" if "master" in result["role"] else "worker"
        status = result["status"].upper()
        check_id = result.get('id')
        
        # Map various statuses to counter keys
        if status in ("PASS", "FIXED"):
            counter_key = "pass"
        elif status in ("FAIL", "ERROR", "REMEDIATION_FAILED"):
            counter_key = "fail"
        elif status == "MANUAL":
            counter_key = "manual"
        elif status in ("SKIPPED", "IGNORED"):
            counter_key = "skipped"
        else:
            return
        
        # Update the live audit_results map to ensure reporting is always current
        if check_id:
            self.audit_results[check_id] = {
                'status': status,
                'role': result.get('role', role),
                'level': result.get('level', '1')
            }
        
        self.stats[role][counter_key] += 1
        self.stats[role]["total"] += 1
        
        if self.verbose >= 2:
            print(f"{Colors.BLUE}[DEBUG] Updated stats: {result['id']} -> {status} ({counter_key}){Colors.ENDC}")

    def _rotate_backups(self, max_backups: int = 5) -> None:
        """
        Maintain only the N most recent backup folders.
        Deletes older backups automatically to save disk space.
        
        Args:
            max_backups (int): Maximum number of backups to keep (default: 5)
        """
        if not os.path.exists(self.backup_dir):
            return
        
        try:
            # List all backup directories (format: backup_YYYYMMDD_HHMMSS)
            backup_folders: List[Tuple[str, str, float]] = []
            for item in os.listdir(self.backup_dir):
                item_path = os.path.join(self.backup_dir, item)
                if os.path.isdir(item_path) and item.startswith("backup_"):
                    # Get modification time for sorting
                    mtime = os.path.getmtime(item_path)
                    backup_folders.append((item_path, item, mtime))
            
            # Sort by modification time (newest first)
            backup_folders.sort(key=lambda x: x[2], reverse=True)
            
            # Delete old backups beyond the limit
            if len(backup_folders) > max_backups:
                removed_count = 0
                for backup_path, backup_name, _ in backup_folders[max_backups:]:
                    try:
                        shutil.rmtree(backup_path)
                        removed_count += 1
                        self.log_activity("BACKUP_ROTATION", 
                                        f"Removed old backup: {backup_name}")
                        print(f"{Colors.YELLOW}[INFO] Cleaned up old backups: removed {backup_name}{Colors.ENDC}")
                    except Exception as e:
                        print(f"{Colors.RED}[!] Failed to delete backup {backup_name}: {e}{Colors.ENDC}")
                
                if removed_count > 0:
                    print(f"{Colors.GREEN}[+] Backup rotation complete: {removed_count} old backup(s) removed{Colors.ENDC}")
        
        except Exception as e:
            print(f"{Colors.RED}[!] Backup rotation error: {e}{Colors.ENDC}")

    def perform_backup(self):
        """
        Create backup of critical Kubernetes configs with automatic rotation.
        สร้างสำรองข้อมูลของการตั้งค่า Kubernetes ที่สำคัญพร้อมการหมุนเวียนอัตโนมัติ
        
        Features:
        - Backs up critical Kubernetes configuration files
        - Automatically rotates backups (keeps only 5 most recent)
        - Logs all backup operations
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(self.backup_dir, f"backup_{timestamp}")
        os.makedirs(backup_path, exist_ok=True)
        
        print(f"\n{Colors.CYAN}[*] Creating Backup...{Colors.ENDC}")
        
        targets = ["/etc/kubernetes/manifests", "/var/lib/kubelet/config.yaml"]
        
        for target in targets:
            if not os.path.exists(target):
                continue
            
            try:
                name = os.path.basename(target)
                if os.path.isdir(target):
                    shutil.copytree(target, os.path.join(backup_path, name))
                else:
                    shutil.copy2(target, backup_path)
            except Exception as e:
                print(f"{Colors.RED}[!] Backup error {target}: {e}{Colors.ENDC}")
        
        print(f"   -> Saved to: {backup_path}")
        self.log_activity("BACKUP_CREATED", f"New backup created: backup_{timestamp}")
        
        # Perform automatic backup rotation
        print(f"{Colors.CYAN}[*] Checking for old backups...{Colors.ENDC}")
        self._rotate_backups(max_backups=5)

    def scan(self, target_level: str, target_role: str, skip_menu: bool = False) -> None:
        """
        Execute audit scan with parallel execution
        ดำเนินการสแกนการตรวจสอบพร้อมการดำเนินการแบบขนาน
        
        Args:
            target_level: CIS level to audit ("1", "2", or "all")
            target_role: Target node role ("master", "worker", or "all")
            skip_menu: If True, skip results menu (used when in "Both" mode)
        """
        print(f"\n{Colors.CYAN}[*] Starting Audit Scan...{Colors.ENDC}")
        self.log_activity("AUDIT_START", 
                         f"Level:{target_level}, Role:{target_role}, Timeout:{self.script_timeout}s")
        self.current_level = target_level
        
        self._prepare_report_dir("audit")
        scripts = self.get_scripts("audit", target_level, target_role)
        self.results = []
        self._init_stats()
        
        # Execute scripts in parallel / ดำเนินการสคริปต์แบบขนาน
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            self._run_scripts_parallel(executor, scripts, "audit")
        
        print(f"\n{Colors.GREEN}[+] Audit Complete.{Colors.ENDC}")
        self.save_reports("audit")
        self.print_stats_summary()
        
        # Store audit results for potential targeted remediation
        self._store_audit_results()
        
        # Trend analysis / การวิเคราะห์แนวโน้ม
        current_score = self.calculate_score(self.stats)
        previous = self.get_previous_snapshot("audit", target_role, target_level)
        if previous:
            self.show_trend_analysis(current_score, previous)
        
        self.save_snapshot("audit", target_role, target_level)
        
        # Show results menu only if not skipped (e.g., in "Both" mode)
        if not skip_menu:
            self.show_results_menu("audit")
        else:
            print(f"\n{Colors.CYAN}[*] Audit Complete. Proceeding to Remediation phase...{Colors.ENDC}")

    def _run_scripts_parallel(self, executor: ThreadPoolExecutor, scripts: List[Dict[str, Any]], mode: str) -> None:
        """
        Run scripts in parallel with progress tracking
        รันสคริปต์แบบขนานพร้อมการติดตามความคืบหน้า
        
        For remediation mode: Implements "Emergency Brake" - stops execution if cluster becomes unhealthy
        สำหรับโหมดการแก้ไข: ใช้ "เบรกฉุกเฉิน" - หยุดการทำงานหากคลัสเตอร์ไม่สมบูรณ์
        """
        futures = {executor.submit(self.run_script, s, mode): s for s in scripts}
        total_checks = len(scripts)
        
        try:
            completed = 0
            for future in as_completed(futures):
                if self.stop_requested:
                    break

                result_raw = future.result()
                typed_result = self._coerce_cis_result(result_raw)
                if typed_result is not None:
                    self.results.append(typed_result)
                    self.update_stats(typed_result)
                    completed += 1

                    # Show progress / แสดงความคืบหน้า
                    progress_pct = (completed / total_checks) * 100
                    self._print_progress(typed_result, completed, total_checks, progress_pct)

                    # EMERGENCY BRAKE: Check cluster health after remediation
                    # If cluster fails after remediation script, stop immediately to prevent cascading damage
                    # เบรกฉุกเฉิน: ตรวจสอบความสมบูรณ์ของคลัสเตอร์หลังการแก้ไข
                    # หากคลัสเตอร์ล้มเหลวหลังสคริปต์แก้ไข ให้หยุดทันทีเพื่อป้องกันความเสียหายต่อเนื่อง
                    if mode == "remediate" and typed_result['status'] in ['PASS', 'FIXED']:
                        if not self.wait_for_healthy_cluster():
                            # CRITICAL: Cluster became unhealthy after this remediation
                            # วิกฤต: คลัสเตอร์ไม่สมบูรณ์หลังจากการแก้ไขนี้
                            error_banner = (
                                f"\n{Colors.RED}{'='*70}\n"
                                f"⛔ CRITICAL: CLUSTER UNHEALTHY - EMERGENCY BRAKE ACTIVATED\n"
                                f"{'='*70}\n"
                                f"Last Remediation: CIS {typed_result['id']}\n"
                                f"Status: {self.health_status}\n"
                                f"Cluster Health: FAILED\n\n"
                                f"Remediation loop aborted to prevent cascading failures.\n\n"
                                f"Recovery Steps:\n"
                                f"  1. Check cluster status: kubectl get nodes\n"
                                f"  2. Check API server: kubectl get pods -n kube-system -l component=kube-apiserver\n"
                                f"  3. Review kubelet logs: journalctl -u kubelet -n 50\n"
                                f"  4. Review API server logs: kubectl logs -n kube-system -l component=kube-apiserver --tail=50\n"
                                f"  5. Restore from backup if needed: /var/backups/cis-remediation/\n\n"
                                f"For manual recovery:\n"
                                f"  - Check /var/log/cis_runner.log for remediation history\n"
                                f"  - Review recent manifest changes in /etc/kubernetes/manifests/\n"
                                f"  - Revert the last remediation if necessary\n"
                                f"{'='*70}{Colors.ENDC}\n"
                            )
                            print(error_banner)
                            self.log_activity("EMERGENCY_BRAKE", f"Cluster failed after CIS {typed_result['id']} remediation")
                            sys.exit(1)
        
        except KeyboardInterrupt:
            self.stop_requested = True
            print("\n[!] Aborted.")

    def _show_verbose_result(self, res: CISResult) -> None:
        """
        Verbose result printer with explicit typing so static analyzers know the signature.
        Prints a compact, detailed line for each result in verbose mode.
        """
        try:
            status = str(res.get("status", ""))
            comp = str(res.get("component", ""))
            rid = str(res.get("id", ""))
            reason = str(res.get("reason", "") or res.get("output", ""))
            duration: float = res["duration"]
            duration_str = f"{duration}s"
            status_color = {
                "PASS": Colors.GREEN,
                "FAIL": Colors.RED,
                "MANUAL": Colors.YELLOW,
                "SKIPPED": Colors.CYAN,
                "FIXED": Colors.GREEN,
                "ERROR": Colors.RED,
                "REMEDIATION_FAILED": Colors.RED
            }
            color = status_color.get(status, Colors.WHITE)
            print(f"   [VERBOSE] {rid} | {comp} | {color}{status}{Colors.ENDC} {duration_str} - {reason}")
        except Exception:
            # Fallback to safe print to avoid raising in progress path
            try:
                print(f"   [VERBOSE] {res}")
            except Exception:
                pass

    def _print_progress(self, result: CISResult, completed: int, total: int, percentage: float) -> None:
        """
        Print progress line
        พิมพ์บรรทัดความคืบหน้า
        """
        if self.verbose > 0:
            self.show_verbose_result(result)
            return
        
        status_color = {
            "PASS": Colors.GREEN,
            "FAIL": Colors.RED,
            "MANUAL": Colors.YELLOW,
            "SKIPPED": Colors.CYAN,
            "FIXED": Colors.GREEN,
            "ERROR": Colors.RED,
            "REMEDIATION_FAILED": Colors.RED
        }
        
        color = status_color.get(result.get('status', ''), Colors.WHITE)
        print(f"   [{percentage:5.1f}%] [{completed}/{total}] {result.get('id', '')} -> {color}{result.get('status', '')}{Colors.ENDC}")

    def fix(self, target_level: str, target_role: str, fix_failed_only: bool = False) -> None:
        """
        Execute remediation with split execution strategy
        Group A (Critical/Config - IDs 1.x, 2.x, 3.x, 4.x): Run SEQUENTIALLY with health checks
        Group B (Resources - IDs 5.x): Run in PARALLEL
        
        Args:
            target_level: CIS level ("1", "2", or "all")
            target_role: Target node role ("master", "worker", or "all")
            fix_failed_only: If True, only remediate checks that FAILED in audit. Otherwise, remediate ALL.
        
        ดำเนินการแก้ไขพร้อมกลยุทธ์การแยกการทำงาน
        กลุ่ม A (วิกฤต/การตั้งค่า - ID 1.x, 2.x, 3.x, 4.x): รันแบบลำดับ (SEQUENTIAL) พร้อมการตรวจสอบความสมบูรณ์
        กลุ่ม B (ทรัพยากร - ID 5.x): รันแบบขนาน (PARALLEL)
        """
        self.current_level = target_level
        # Prevent remediation if cluster is critical
        # ป้องกันการแก้ไขหากคลัสเตอร์อยู่ในสถานะวิกฤต
        if "CRITICAL" in self.health_status:
            print(f"{Colors.RED}[-] Cannot remediate: Cluster health is CRITICAL.{Colors.ENDC}")
            self.log_activity("FIX_SKIPPED", "Cluster health critical")
            return
        
        if not self._aggressive_remediation_confirmed:
            print(f"{Colors.YELLOW}[WARN] Auto-confirming Aggressive Hard Kill mode. Proceeding immediately...{Colors.ENDC}")
            self._aggressive_remediation_confirmed = True
        

        self._prepare_report_dir("remediation")
        self.log_activity("FIX_START", f"Level:{target_level}, Role:{target_role}, FailedOnly:{fix_failed_only}")
        self.perform_backup()
        
        # Ensure audit log directory exists before starting remediation
        # ตรวจสอบให้แน่ใจว่ามีไดเรกทอรีบันทึกการตรวจสอบก่อนเริ่มการแก้ไข
        self.ensure_audit_log_dir()
        
        print(f"\n{Colors.YELLOW}[*] Starting Remediation with Split Strategy...{Colors.ENDC}")
        scripts: List[Dict[str, Any]] = self.get_scripts("remediate", target_level, target_role) or []
        
        # Filter scripts if "fix failed only" mode is enabled
        # กรองสคริปต์หากเปิดใช้งานโหมด "แก้ไขเฉพาะที่ล้มเหลว"
        if fix_failed_only:
            scripts = self._filter_failed_checks(scripts)
            if not scripts:
                print(f"{Colors.GREEN}[+] No failed items to remediate. All checks passed!{Colors.ENDC}")
                return
        
        self.results = []
        self._init_stats()
        
        # Execute remediation with split strategy
        # ดำเนินการแก้ไขด้วยกลยุทธ์การแยกการทำงาน
        self._run_remediation_with_split_strategy(scripts)
        
        print(f"\n{Colors.GREEN}[+] Remediation Complete.{Colors.ENDC}")
        self.save_reports("remediate")
        self.print_compliance_report()
        
        # Trend analysis / การวิเคราะห์แนวโน้ม
        current_score = self.calculate_score(self.stats)
        previous = self.get_previous_snapshot("remediate", target_role, target_level)
        if previous:
            self.show_trend_analysis(current_score, previous)
        
        self.save_snapshot("remediate", target_role, target_level)
        self._run_level2_preferences(target_level)
        self.show_results_menu("remediate")

    def _run_level2_preferences(self, target_level: str) -> None:
        """Run the Level 2 safe defaults once Level 2 remediation finishes."""
        if target_level not in {"2", "all"}:
            return
        if not self.config_data.get("level2_preferences"):
            return

        try:
            helper = Level2Remediator(
                self.config_data,
                kubectl_cmd=shlex.join(self.get_kubectl_cmd()),
                namespace_exempt_checker=self.is_namespace_exempt,
                namespace_label_applier=self.label_namespace_as_exempt,
            )
            summary = helper.run_all()
            enforced_tasks = [res["task_id"] for res in summary["enforced"]]
            skipped_messages = [f"{res['task_id']}: {res['detail']}" for res in summary["skipped"]]
            if enforced_tasks:
                print(f"{Colors.GREEN}[✓] Level 2 tasks enforced: {', '.join(enforced_tasks)}{Colors.ENDC}")
            if skipped_messages:
                print(f"{Colors.YELLOW}[WARN] Level 2 tasks skipped: {', '.join(skipped_messages)}{Colors.ENDC}")
            self.log_activity(
                "LEVEL2_PREFERENCES",
                f"Enforced: {enforced_tasks}; Skipped: {skipped_messages}",
            )
        except Exception as exc:
            warning = f"Level 2 preferences application failed: {exc}"
            print(f"{Colors.YELLOW}[WARN] {warning}{Colors.ENDC}")
            self.log_activity("LEVEL2_PREFERENCES_ERROR", warning)

    def _filter_failed_checks(self, scripts: Optional[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """
        Filter scripts to include only those that FAILED in the audit phase.
        
        Logic:
        - Only includes checks that are present in self.audit_results
        - Only includes checks with status FAIL or ERROR
        - EXCLUDES checks marked as REMEDIATION_FAILED (already failed remediation attempt)
        - Optionally includes MANUAL checks if configured
        
        Args:
            scripts: List of script objects to filter
        
        Returns:
            Filtered list containing only failed checks from audit
            
        กรองสคริปต์เพื่อรวมเฉพาะรายการที่ล้มเหลว (FAILED) ในขั้นตอนการตรวจสอบ (Audit)
        """
        # Defensive guard to make typing explicit for static analyzers
        if not scripts:
            if not self.audit_results:
                print(f"{Colors.YELLOW}[!] No audit results available. Running full remediation.{Colors.ENDC}")
            # Ensure a list is always returned
            return []

        if not self.audit_results:
            print(f"{Colors.YELLOW}[!] No audit results available. Running full remediation.{Colors.ENDC}")
            return list(scripts)
        
        failed_scripts: List[Dict[str, Any]] = []
        skipped_pass: List[Dict[str, Any]] = []
        skipped_remediation_failed: List[Dict[str, Any]] = []
        
        for script in scripts:
            check_id = script['id']
            
            # Check if this ID was audited
            # ตรวจสอบว่า ID นี้ได้รับการตรวจสอบแล้วหรือไม่
            if check_id not in self.audit_results:
                # Not audited, include it in remediation
                failed_scripts.append(script)
                continue
            
            audit_status = self.audit_results[check_id].get('status', 'UNKNOWN')
            
            # CRITICAL: Skip items marked as REMEDIATION_FAILED
            # These already failed verification after remediation and should NOT be re-attempted
            # They require manual intervention instead
            # วิกฤต: ข้ามรายการที่ทำเครื่องหมายว่าเป็น REMEDIATION_FAILED
            # รายการเหล่านี้ล้มเหลวในการตรวจสอบหลังการแก้ไขแล้ว และไม่ควรพยายามซ้ำ
            # ต้องมีการดำเนินการด้วยตนเองแทน
            if audit_status == 'REMEDIATION_FAILED':
                skipped_remediation_failed.append(script)
                print(f"{Colors.RED}    [SKIP] {check_id}: Previously failed remediation verification - requires manual intervention{Colors.ENDC}")
                continue
            
            # CRITICAL: Skip items that already PASSED in audit
            # They don't need remediation
            # วิกฤต: ข้ามรายการที่ผ่าน (PASS) ในการตรวจสอบแล้ว
            if audit_status == 'PASS':
                skipped_pass.append(script)
                continue
            
            # Include only FAIL, ERROR, and MANUAL status checks
            # รวมเฉพาะการตรวจสอบสถานะ FAIL, ERROR และ MANUAL
            if audit_status in ['FAIL', 'ERROR']:
                failed_scripts.append(script)
            elif audit_status == 'MANUAL':
                # Include MANUAL checks as they might need re-verification
                failed_scripts.append(script)
        
        # Summary report / รายงานสรุป
        print(f"{Colors.CYAN}[*] Remediation Filter Summary:{Colors.ENDC}")
        print(f"    Total checks available: {len(scripts)}")
        print(f"    {Colors.GREEN}Already PASSED:{Colors.ENDC} {len(skipped_pass)} (SKIPPED - no remediation needed)")
        print(f"    {Colors.RED}REMEDIATION_FAILED:{Colors.ENDC} {len(skipped_remediation_failed)} (SKIPPED - manual intervention required)")
        print(f"    {Colors.RED}FAILED/ERROR:{Colors.ENDC} {len([s for s in failed_scripts if self.audit_results.get(s['id'], {}).get('status') in ['FAIL', 'ERROR']])}")
        print(f"    {Colors.YELLOW}MANUAL:{Colors.ENDC} {len([s for s in failed_scripts if self.audit_results.get(s['id'], {}).get('status') == 'MANUAL'])}")
        print(f"    {Colors.CYAN}NOT AUDITED:{Colors.ENDC} {len([s for s in failed_scripts if s['id'] not in self.audit_results])}")
        print(f"    → Will remediate: {len(failed_scripts)} checks\n")
        
        if self.verbose >= 2:
            if skipped_pass:
                print(f"{Colors.CYAN}[DEBUG] Skipped PASSED checks:{Colors.ENDC}")
                for script in skipped_pass:
                    print(f"        {script['id']}")
            if skipped_remediation_failed:
                print(f"{Colors.RED}[DEBUG] Skipped REMEDIATION_FAILED checks (manual intervention required):{Colors.ENDC}")
                for script in skipped_remediation_failed:
                    print(f"        {script['id']}")
            print()
        
        return failed_scripts
    
    def _store_audit_results(self):
        """
        Store current audit results in a dictionary keyed by check ID.
        Used for targeted remediation (fixing only failed items).
        
        จัดเก็บผลการตรวจสอบปัจจุบันในพจนานุกรมโดยใช้ ID การตรวจสอบเป็นคีย์
        ใช้สำหรับการแก้ไขที่ตรงเป้าหมาย (แก้ไขเฉพาะรายการที่ล้มเหลว)
        """
        self.audit_results.clear()
        
        for result in self.results:
            check_id = result.get('id')
            if check_id:
                self.audit_results[check_id] = {
                    'status': result.get('status'),
                    'role': result.get('role'),
                    'level': result.get('level')
                }
        
        if self.verbose >= 1:
            print(f"{Colors.BLUE}[DEBUG] Stored {len(self.audit_results)} audit results for targeted remediation{Colors.ENDC}")

    def _classify_remediation_type(self, check_id: Optional[str]) -> Tuple[bool, str]:
        """
        SMART WAIT OPTIMIZATION: Classify remediation type to determine if health check is needed.
        Also identifies manifest path for internal YAML fixes.
        
        Classification Rules:
        - SYSTEM_CHECK: Permission/ownership changes (1.1.x) or other non-YAML fixes
        - YAML_CONFIG: Config file changes in static pod manifests (1.2.x, 1.3.x, 1.4.x, 2.x)
        
        Args:
            check_id (Optional[str]): The CIS check ID (e.g., '1.1.1', '1.2.1', '5.6.4')
        
        Returns:
            tuple: (requires_health_check: bool, manifest_path: str)
            
        การเพิ่มประสิทธิภาพการรอแบบอัจฉริยะ: จำแนกประเภทการแก้ไขเพื่อพิจารณาว่าจำเป็นต้องตรวจสอบความสมบูรณ์หรือไม่
        """
        # Defensive guard for callers that might pass None or empty values
        if not check_id:
            return False, "System Check"

        # YAML Config Checks - API Server
        if check_id.startswith('1.2.'):
            return True, "/etc/kubernetes/manifests/kube-apiserver.yaml"
            
        # YAML Config Checks - Controller Manager
        if check_id.startswith('1.3.'):
            return True, "/etc/kubernetes/manifests/kube-controller-manager.yaml"
            
        # YAML Config Checks - Scheduler
        if check_id.startswith('1.4.'):
            return True, "/etc/kubernetes/manifests/kube-scheduler.yaml"
            
        # YAML Config Checks - Etcd
        if check_id.startswith('2.'):
            return True, "/etc/kubernetes/manifests/etcd.yaml"
            
        # Safe checks - file permissions (chmod) or ownership (chown), no service impact
        # การตรวจสอบที่ปลอดภัย - การอนุญาตไฟล์ (chmod) หรือความเป็นเจ้าของ (chown) ไม่มีผลกระทบต่อบริการ
        if check_id.startswith('1.1.'):
            return False, "File Permissions/Ownership"
        
        # All other remediation checks default to system check
        return False, "System Check"

    def apply_batch_remediation(self, manifest_path: str, scripts: List[Dict[str, Any]]) -> None:
        """
        Requirement 1: Enforce Batch Write (Write-Once-Per-File)
        Apply multiple remediations to a single manifest file in one go.
        """
        if not scripts:
            return
        
        component = self._get_process_name_for_check(scripts[0]['id'])
        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"BATCH REMEDIATION: {component}")
        print(f"Manifest: {manifest_path}")
        print(f"Checks: {', '.join([s['id'] for s in scripts])}")
        print(f"{'='*70}{Colors.ENDC}")

        if not self._aggressive_remediation_confirmed:
            print(f"{Colors.YELLOW}[WARN] Auto-confirming Aggressive Hard Kill mode. Proceeding immediately...{Colors.ENDC}")
            self._aggressive_remediation_confirmed = True
        
        # 1. Load the target YAML file to extract dynamic values
        try:
            with open(manifest_path, 'r') as f:
                if yaml:
                    raw = yaml.safe_load(f)
                else:
                    raw = {}
            # Normalize to a concrete Dict[str, Any] so .get has a stable type for static analyzers
            if isinstance(raw, dict):
                data = cast(Dict[str, Any], raw)
            else:
                data = {}
        except Exception as e:
            print(f"{Colors.YELLOW}[WARN] Could not read {manifest_path} for dynamic values: {e}{Colors.ENDC}")
            data = {}

        # 2. Golden Config Strategy (Force Overwrite for Static Pods)
        is_static_pod = any(c in manifest_path for c in ["kube-apiserver", "kube-controller-manager", "kube-scheduler"])
        
        if is_static_pod:
            print(f"{Colors.YELLOW}[BATCH] Force Overwriting with Golden Config Strategy...{Colors.ENDC}")
            template = None
            if "kube-apiserver" in manifest_path:
                template = golden_configs.APISERVER_HARDENED_TEMPLATE
                
                # Step 1: Ensure Admission Control Config exists
                os.makedirs("/etc/kubernetes/admission-control", exist_ok=True)
                admission_config = """apiVersion: apiserver.config.k8s.io/v1
kind: AdmissionConfiguration
plugins:
- name: EventRateLimit
  configuration:
    apiVersion: eventratelimit.admission.k8s.io/v1alpha1
    kind: Configuration
    limits:
    - type: Server
      burst: 20000
      qps: 5000
"""
                try:
                    save_yaml_robust("/etc/kubernetes/admission-control/admission-control.yaml", admission_config)
                    print(f"{Colors.GREEN}[BATCH] Created admission-control.yaml{Colors.ENDC}")
                except Exception as e:
                    print(f"{Colors.RED}[BATCH ERROR] Failed to create admission-control.yaml: {e}{Colors.ENDC}")

                # Step 2: Ensure Audit Policy exists (minimal valid policy)
                audit_policy = """apiVersion: audit.k8s.io/v1
kind: Policy
rules:
- level: Metadata
"""
                try:
                    if not os.path.exists("/etc/kubernetes/audit-policy.yaml"):
                        save_yaml_robust("/etc/kubernetes/audit-policy.yaml", audit_policy)
                        print(f"{Colors.GREEN}[BATCH] Created audit-policy.yaml{Colors.ENDC}")
                except Exception as e:
                    print(f"{Colors.RED}[BATCH ERROR] Failed to create audit-policy.yaml: {e}{Colors.ENDC}")

            elif "kube-controller-manager" in manifest_path:
                template = golden_configs.CONTROLLER_HARDENED_TEMPLATE
            elif "kube-scheduler" in manifest_path:
                template = golden_configs.SCHEDULER_HARDENED_TEMPLATE
            
            if template:
                # Extract dynamic values from current manifest (Smart Inheritance)
                adv_addr = None
                k8s_ver = "v1.28.2" # Default fallback
                etcd_servers = None
                svc_cluster_ip = None
                svc_acc_issuer = None
                
                try:
                    # Safely extract containers ensuring a stable List[Dict[str, Any]] for static analysis
                    containers: List[Dict[str, Any]] = []
                    # Normalize and explicitly type intermediary variables so static analyzers can infer types
                    spec_obj_raw: Optional[Dict[str, Any]] = None
                    # Extract raw candidate and narrow its type using runtime checks so type checkers see Dict[str, Any]
                    spec_candidate_raw: Optional[Any] = data.get('spec')
                    spec_candidate: Optional[Dict[str, Any]] = None
                    # Only cast and assign after an explicit runtime type check
                    if isinstance(spec_candidate_raw, dict):
                        spec_candidate = cast(Dict[str, Any], spec_candidate_raw)
                        spec_obj_raw = spec_candidate

                    # Normalize into a concrete dict for downstream usage to satisfy type checkers
                    spec: Dict[str, Any] = {}
                    if isinstance(spec_obj_raw, dict):
                        spec = spec_obj_raw

                        # Try top-level containers first
                        raw_containers = spec.get('containers')
                        if isinstance(raw_containers, list):
                            # Build validated list of container dicts with explicit casting
                            containers = []
                            raw_containers_list = cast(List[Any], raw_containers)
                            for item in raw_containers_list:
                                if isinstance(item, dict):
                                    containers.append(cast(Dict[str, Any], item))
                        else:
                            # Fallback to template.spec.containers
                            template_obj = spec.get('template')
                            if isinstance(template_obj, dict):
                                # Do not overwrite the outer 'template' (which may be a string template).
                                # Use a local variable for the dict representation of the template to avoid shadowing.
                                template_obj_dict = cast(Dict[str, Any], template_obj)
                                # Safely extract 'spec' with an explicit runtime check so static analyzers
                                # can narrow the type to Dict[str, Any] before using it.
                                t_spec_candidate: Any = template_obj_dict.get('spec')
                                t_spec_obj: Optional[Dict[str, Any]] = None
                                if isinstance(t_spec_candidate, dict):
                                    t_spec_obj = cast(Dict[str, Any], t_spec_candidate)
                                if isinstance(t_spec_obj, dict):
                                    t_spec = t_spec_obj
                                    raw_containers = t_spec.get('containers')
                                    if isinstance(raw_containers, list):
                                        containers = []
                                        raw_containers_list = cast(List[Any], raw_containers)
                                        for item in raw_containers_list:
                                            if isinstance(item, dict):
                                                containers.append(cast(Dict[str, Any], item))

                    # Proceed if we have validated container dicts
                    if containers:
                        # Get version from image
                        image = str(containers[0].get('image', '')) if isinstance(containers[0].get('image', ''), str) else ''
                        if ':' in image:
                            k8s_ver = image.split(':')[-1]
                        
                        # Get vital flags from command (ensure command is a list)
                        command_candidate = containers[0].get('command')
                        # Only treat it as a list if it's actually a list to avoid assigning None/Any to List[str]
                        if isinstance(command_candidate, list):
                            # Help static analyzers by explicitly casting the list items to List[str]
                            command_list_strs = cast(List[str], command_candidate)
                            for arg in command_list_strs:
                                # Coerce to string defensively to avoid unnecessary isinstance checks
                                try:
                                    arg_str = str(arg)
                                except Exception:
                                    continue
                                if arg_str.startswith('--advertise-address='):
                                    adv_addr = arg_str.split('=', 1)[1]
                                elif arg_str.startswith('--etcd-servers='):
                                    etcd_servers = arg_str.split('=', 1)[1]
                                elif arg_str.startswith('--service-cluster-ip-range='):
                                    svc_cluster_ip = arg_str.split('=', 1)[1]
                                elif arg_str.startswith('--service-account-issuer='):
                                    svc_acc_issuer = arg_str.split('=', 1)[1]
                except Exception:
                    pass
                
                # Safety Net: Abort if vital flags are missing for kube-apiserver
                if "kube-apiserver" in manifest_path:
                    missing: List[str] = []
                    if not adv_addr: missing.append("--advertise-address")
                    if not etcd_servers: missing.append("--etcd-servers")
                    if not svc_cluster_ip: missing.append("--service-cluster-ip-range")
                    
                    if missing:
                        print(f"{Colors.RED}[CRITICAL] Could not determine vital flags ({', '.join(missing)}). Aborting to prevent crash.{Colors.ENDC}")
                        for script in scripts:
                            res = self._create_result(script, "FAIL", f"Smart Inheritance failed: missing {missing}", 0)
                            self.results.append(res)
                            self.update_stats(res)
                        return

                # Apply placeholders - ensure template is converted to a concrete string
                tmpl_str = str(template or "")
                final_yaml: str = tmpl_str.replace("{{ADVERTISE_ADDRESS}}", adv_addr or "127.0.0.1") \
                                 .replace("{{K8S_VERSION}}", k8s_ver) \
                                 .replace("{{ETCD_SERVERS}}", etcd_servers or "https://127.0.0.1:2379") \
                                 .replace("{{SERVICE_CLUSTER_IP_RANGE}}", svc_cluster_ip or "10.96.0.0/12") \
                                 .replace("{{SERVICE_ACCOUNT_ISSUER}}", svc_acc_issuer or "https://kubernetes.default.svc.cluster.local")
                
                try:
                    # Create backup first
                    self.atomic_manager.create_backup(manifest_path)
                    static_expected_flags = []
                    if "kube-apiserver" in manifest_path:
                        static_expected_flags = [
                            "--request-timeout=300s",
                            "--service-account-extend-token-expiration=false",
                            "--authorization-mode=Node,RBAC"
                        ]
                    elif "kube-controller-manager" in manifest_path:
                        static_expected_flags = ["--feature-gates=RotateKubeletServerCertificate=true"]

                    save_yaml_robust(manifest_path, final_yaml, expected_flags=static_expected_flags)
                    print(f"{Colors.GREEN}[BATCH] Golden Config applied to {manifest_path}{Colors.ENDC}")
                except Exception as e:
                    print(f"{Colors.RED}[BATCH ERROR] Failed to write Golden Config: {e}{Colors.ENDC}")
                    return
            else:
                print(f"{Colors.RED}[BATCH ERROR] No Golden Config template found for {manifest_path}{Colors.ENDC}")
                return
        else:
            # 3. Dynamic Patching Logic (for etcd, etc.)
            CSV_CHECKS = ["1.2.8", "1.2.11", "1.2.14", "1.2.29"]
            expected_flags: List[str] = []
            
            for script in scripts:
                cfg = self.get_remediation_config_for_check(script['id'])
                mod_type = "csv" if script['id'] in CSV_CHECKS else "string"
                
                # Collect all flags to modify
                modifications: List[str] = []
                flag_name = cfg.get('flag_name')
                expected_value = cfg.get('expected_value')
                if flag_name:
                    if not flag_name.startswith('--'): flag_name = '--' + flag_name
                    entry = f"{flag_name}={expected_value}" if expected_value is not None else flag_name
                    modifications.append(entry)
                    expected_flags.append(entry)
                
                flags_dict = cfg.get('flags', {})
                for f_name, f_val in flags_dict.items():
                    if not f_name.startswith('--'): f_name = '--' + f_name
                    entry = f"{f_name}={f_val}" if f_val is not None else f_name
                    modifications.append(entry)
                    expected_flags.append(entry)
                
                # Apply to the 'data' dictionary
                try:
                    # Ensure 'data' is a dict and extract containers with explicit type checks
                    containers: List[Dict[str, Any]] = []
                    # Normalize and safely extract 'spec' to satisfy static analyzers
                    spec_obj_raw = data.get('spec')
                    spec: Dict[str, Any] = {}
                    if isinstance(spec_obj_raw, dict):
                        spec = spec_obj_raw

                        # Try top-level containers first
                        raw = spec.get('containers')
                        if isinstance(raw, list):
                            containers = []
                            raw_containers_list = cast(List[Any], raw)
                            for item in raw_containers_list:
                                if isinstance(item, dict):
                                    containers.append(cast(Dict[str, Any], item))
                        else:
                            # Fallback to template.spec.containers
                            template_obj = spec.get('template')
                            # Use concrete Any for the candidate so static analyzers don't treat it as Unknown|None
                            t_spec_candidate_raw: Any = None
                            t_spec_obj: Optional[Dict[str, Any]] = None
                            if isinstance(template_obj, dict):
                                # Cast template_obj to a Dict[str, Any] before calling .get to provide a stable type
                                template_obj_dict = cast(Dict[str, Any], template_obj)
                                t_spec_candidate_raw = template_obj_dict.get('spec')
                            # Narrow the candidate to a dict before casting
                            if isinstance(t_spec_candidate_raw, dict):
                                t_spec_obj = cast(Dict[str, Any], t_spec_candidate_raw)
                            if isinstance(t_spec_obj, dict):
                                t_spec = t_spec_obj
                                raw = t_spec.get('containers')
                                if isinstance(raw, list):
                                    containers = []
                                    raw_containers_list = cast(List[Any], raw)
                                    for item in raw_containers_list:
                                        if isinstance(item, dict):
                                            containers.append(cast(Dict[str, Any], item))
                    # Iterate only over validated container dicts
                    for container in containers:
                        command = container.get('command')
                        if not isinstance(command, list):
                            continue
                        # Narrow the type for static analyzers: treat command as List[str]
                        command_list: List[str] = cast(List[str], command)
                        # Process each modification entry individually to ensure f_name is always bound
                        for flag_str in modifications:
                            if '=' in flag_str:
                                f_name_raw, f_val = flag_str.split('=', 1)
                            else:
                                f_name_raw, f_val = flag_str, None
                            
                            # Coerce flag name to a safe string and skip if not possible
                            try:
                                f_name = str(f_name_raw)
                            except Exception:
                                continue
                            
                            prefix = f_name + "="
                            # Update or append flag
                            flag_index = -1
                            for i, item in enumerate(command_list):
                                # Coerce item to string (robust and avoids redundant isinstance checks)
                                item_str = str(item)
                                if item_str.startswith(prefix) or item_str == f_name:
                                    flag_index = i
                                    break
                            
                            if flag_index >= 0:
                                if mod_type == "csv":
                                    existing = command_list[flag_index]
                                    current_val = existing.split("=", 1)[1] if "=" in existing else ""
                                    # Only add if f_val is provided and not already present
                                    if f_val is not None and f_val not in current_val.split(','):
                                        new_val = f"{current_val},{f_val}" if current_val else str(f_val)
                                        command_list[flag_index] = f"{f_name}={new_val}"
                                else:
                                    command_list[flag_index] = f"{f_name}={f_val}" if f_val is not None else f_name
                            else:
                                command_list.append(f"{f_name}={f_val}" if f_val is not None else f_name)
                        # Ensure container reflects the (possibly) modified command list
                        container['command'] = command_list
                except Exception as e:
                    print(f"{Colors.RED}[BATCH ERROR] Failed to modify data for {script['id']}: {e}{Colors.ENDC}")

            # Write the dictionary back to disk with atomic persistence
            try:
                if not yaml:
                    raise RuntimeError("yaml module is required to serialize manifests during batch remediation")
                content = yaml.dump(data, default_flow_style=False, width=1000, indent=2, sort_keys=False)
                self.atomic_manager.create_backup(manifest_path)
                save_yaml_robust(manifest_path, content, expected_flags=expected_flags)
                print(f"{Colors.GREEN}[BATCH] Write verified successfully.{Colors.ENDC}")
            except Exception as e:
                print(f"{Colors.RED}[BATCH ERROR] Failed to write {manifest_path}: {e}{Colors.ENDC}")
                return

        # 4. Robust Restart
        print(f"{Colors.YELLOW}[BATCH] Allowing filesystem to settle (2s)...{Colors.ENDC}")
        time.sleep(2)
        print(f"{Colors.YELLOW}[BATCH] Syncing filesystem (5s sleep)...{Colors.ENDC}")
        time.sleep(5)
        
        print(f"{Colors.YELLOW}[BATCH] Triggering hard kill for {component}...{Colors.ENDC}")
        self._hard_kill_containers(component)
        
        # 5. Wait for health check ONCE
        print(f"{Colors.YELLOW}[BATCH] Waiting for cluster health...{Colors.ENDC}")
        is_healthy, health_msg = self.atomic_manager.wait_for_cluster_healthy()
        
        if not is_healthy:
            print(f"{Colors.RED}[BATCH ERROR] Cluster unhealthy after batch update: {health_msg}{Colors.ENDC}")
            for script in scripts:
                res = self._create_result(script, "FAIL", f"Cluster unhealthy after batch: {health_msg}", 0)
                self.results.append(res)
                self.update_stats(res)
            return

        # 6. Verify each check via aggressive verification loop
        print("======================================================================")
        print("PHASE 1: PRE-REMEDIATION DIAGNOSTIC SCAN")
        print("(Failures listed below are EXPECTED. They will be fixed in Phase 2)")
        print("======================================================================")
        print(f"{Colors.CYAN}[BATCH] Verifying {len(scripts)} checks...{Colors.ENDC}")
        for script in scripts:
            status, reason = self._run_audit_verification_loop(script, component)

            # Requirement 4: Ensure audit_results is updated with NEW status
            if status in ["PASS", "FIXED"]:
                self.audit_results[script['id']] = {
                    'status': status,
                    'role': script['role'],
                    'level': script['level'],
                    'reason': reason,
                    'timestamp': datetime.now().isoformat()
                }

            res = cast(CISResult, {
                "id": script['id'],
                "role": script["role"],
                "level": script["level"],
                "status": status,
                "duration": 0.0,
                "reason": reason,
                "fix_hint": "",
                "cmds": [],
                "output": reason,
                "path": script["path"],
                "component": self.get_component_for_rule(script['id'])
            })
            self.results.append(res)
            self.update_stats(res)
            print(f"   • {script['id']}: {Colors.GREEN if status in ['PASS', 'FIXED'] else Colors.RED}{status}{Colors.ENDC}")

    def print_compliance_report(self):
        """
        Requirement 2: Fix Reporting Logic
        Ensure the final report reflects the LATEST status by iterating through audit_results.
        """
        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"FINAL COMPLIANCE REPORT (Post-Remediation)")
        print(f"{'='*70}{Colors.ENDC}")
        
        new_stats = {
            "master": {"pass": 0, "fail": 0, "manual": 0, "skipped": 0, "error": 0, "total": 0},
            "worker": {"pass": 0, "fail": 0, "manual": 0, "skipped": 0, "error": 0, "total": 0}
        }
        
        for data in self.audit_results.values():
            role = data.get('role', 'master')
            status = data.get('status', 'UNKNOWN')
            
            if status in ["PASS", "FIXED"]:
                new_stats[role]["pass"] += 1
            elif status in ["FAIL", "ERROR", "REMEDIATION_FAILED"]:
                new_stats[role]["fail"] += 1
            elif status == "MANUAL":
                new_stats[role]["manual"] += 1
            elif status in ["SKIPPED", "IGNORED"]:
                new_stats[role]["skipped"] += 1
            
            new_stats[role]["total"] += 1
            
        self.stats = new_stats
        self.print_stats_summary()

    def _run_remediation_with_split_strategy(self, scripts: List[Dict[str, Any]]) -> None:
        """
        Execute remediation with split execution strategy for stability.
        
        SPLIT STRATEGY:
        - GROUP A (Critical/Config): IDs starting with 1., 2., 3., 4.
          These involve service restarts (API/Kubelet/Etcd) and are executed SEQUENTIALLY.
          Health check called after EACH script to catch failures early.
          
        - GROUP B (Resources): IDs starting with 5.
          These are API calls (NetworkPolicies, Namespaces) and are executed in PARALLEL.
          More stable, no service restarts involved.
        
        Rationale: Critical config changes can cause race conditions when parallel.
        Resources are safe to execute in parallel as they don't restart services.
        
        ดำเนินการแก้ไขด้วยกลยุทธ์การแยกการทำงานเพื่อความเสถียร
        กลยุทธ์การแยก:
        - กลุ่ม A (วิกฤต/การตั้งค่า): ID ที่ขึ้นต้นด้วย 1., 2., 3., 4.
          สิ่งเหล่านี้เกี่ยวข้องกับการรีสตาร์ทบริการ (API/Kubelet/Etcd) และรันแบบลำดับ (SEQUENTIAL)
          มีการเรียกตรวจสอบความสมบูรณ์หลังแต่ละสคริปต์เพื่อตรวจจับความล้มเหลวแต่เนิ่นๆ
          
        - กลุ่ม B (ทรัพยากร): ID ที่ขึ้นต้นด้วย 5.
          สิ่งเหล่านี้คือการเรียก API (NetworkPolicies, Namespaces) และรันแบบขนาน (PARALLEL)
          มีความเสถียรกว่า ไม่มีการรีสตาร์ทบริการที่เกี่ยวข้อง
        """
        # Reset manual pending items for this remediation run
        # รีเซ็ตรายการที่รอดำเนินการด้วยตนเองสำหรับการรันการแก้ไขนี้
        self.manual_pending_items = []
        
        # Split scripts into Group A (critical) and Group B (resources)
        # แยกสคริปต์ออกเป็นกลุ่ม A (วิกฤต) และกลุ่ม B (ทรัพยากร)
        group_a = [s for s in scripts if any(s['id'].startswith(prefix) for prefix in ['1.', '2.', '3.', '4.'])]
        group_b = [s for s in scripts if any(s['id'].startswith(prefix) for prefix in ['5.'])]
        
        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"REMEDIATION EXECUTION PLAN")
        print(f"{'='*70}{Colors.ENDC}")
        print(f"  Group A (Critical/Config - SEQUENTIAL): {len(group_a)} checks")
        if group_a:
            for s in group_a:
                print(f"    • {s['id']}")
        print(f"\n  Group B (Resources - PARALLEL): {len(group_b)} checks")
        if group_b:
            for s in group_b:
                print(f"    • {s['id']}")
        print(f"{Colors.CYAN}{'='*70}{Colors.ENDC}\n")
        
        # --- EXECUTE GROUP A: BATCHED & SEQUENTIAL ---
        if group_a:
            print(f"\n{Colors.YELLOW}[*] Executing GROUP A (Critical/Config) - Batching Static Pods...{Colors.ENDC}")
            
            # 1. Filter out Manual checks first
            active_group_a: List[Dict[str, Any]] = []
            for script in group_a:
                is_manual = False
                reason = ""
                remediation_cfg = self.get_remediation_config_for_check(script['id'])
                if remediation_cfg.get("remediation") == "manual":
                    is_manual = True
                    reason = "Remediation marked as manual in configuration"
                elif script['id'] in self.audit_results and self.audit_results[script['id']].get('status') == 'MANUAL':
                    is_manual = True
                    reason = "Audit phase identified this as manual intervention required"
                elif os.path.exists(script['path']) and self._is_manual_check(script['path']):
                    is_manual = True
                    reason = "Script explicitly marked as MANUAL"
                
                if is_manual:
                    print(f"{Colors.YELLOW}[INFO] Check {script['id']} is MANUAL. Skipping automation.{Colors.ENDC}")
                    self.manual_pending_items.append({
                        'id': script['id'],
                        'role': script.get('role', 'unknown'),
                        'level': script.get('level', 'unknown'),
                        'path': script.get('path', ''),
                        'reason': reason,
                        'component': self.get_component_for_rule(script['id'])
                    })
                    self.log_activity("MANUAL_CHECK_SKIPPED", f"{script['id']}: {reason}")
                else:
                    active_group_a.append(script)

            # 2. Group active checks by manifest for Static Pods
            batch_groups: Dict[str, List[Dict[str, Any]]] = {}
            sequential_a: List[Dict[str, Any]] = []
            for script in active_group_a:
                is_yaml, manifest_path = self._classify_remediation_type(script['id'])
                if is_yaml and manifest_path and os.path.exists(manifest_path):
                    # Only batch static pods as requested
                    if any(comp in manifest_path for comp in ["kube-apiserver", "kube-controller-manager", "kube-scheduler"]):
                        if manifest_path not in batch_groups:
                            batch_groups[manifest_path] = []
                        batch_groups[manifest_path].append(script)
                        continue
                sequential_a.append(script)

            # 3. Execute Batches
            for manifest_path, scripts_to_batch in batch_groups.items():
                if self.stop_requested: break
                self.apply_batch_remediation(manifest_path, scripts_to_batch)

            # 4. Execute remaining Sequential
            for idx, script in enumerate(sequential_a, 1):
                if self.stop_requested: break
                
                print(f"\n{Colors.CYAN}[Group A Sequential {idx}/{len(sequential_a)}] Running: {script['id']}...{Colors.ENDC}")
                
                requires_health_check, classification = self._classify_remediation_type(script['id'])
                print(f"{Colors.BLUE}    [Smart Wait] {classification}{Colors.ENDC}")
                
                if script['id'].startswith('2.'):
                    remediation_cfg = self.get_remediation_config_for_check(script['id'])
                    result = self._run_etcd_remediation(script, remediation_cfg)
                else:
                    result = self.run_script(script, "remediate")
                
                typed_result = self._coerce_cis_result(result)
                if typed_result is not None:
                    self.results.append(typed_result)
                    self.update_stats(typed_result)
                    progress_pct = (idx / len(sequential_a)) * 100
                    self._print_progress(typed_result, idx, len(sequential_a), progress_pct)
                    
                    if typed_result['status'] in ['PASS', 'FIXED'] and requires_health_check:
                        print(f"{Colors.YELLOW}    [Health Check] Verifying cluster stability...{Colors.ENDC}")
                        if not self.wait_for_healthy_cluster(skip_health_check=False):
                            print(f"{Colors.RED}⛔ CRITICAL: CLUSTER UNHEALTHY - ABORTING{Colors.ENDC}")
                            self.stop_requested = True
                            break
            
            print(f"\n{Colors.GREEN}[+] GROUP A (Critical/Config) Complete.{Colors.ENDC}")
        
        # --- EXECUTE GROUP B: PARALLEL (Safe, no service restarts) ---
        # --- ดำเนินการกลุ่ม B: แบบขนาน (ปลอดภัย ไม่มีการรีสตาร์ทบริการ) ---
        if group_b:
            print(f"\n{Colors.YELLOW}[*] Executing GROUP B (Resources) - PARALLEL mode...{Colors.ENDC}")
            
            # Filter out MANUAL checks from GROUP B
            # กรองการตรวจสอบด้วยตนเองออกจากกลุ่ม B
            group_b_automated: List[Dict[str, Any]] = []
            group_b_manual: List[Tuple[Dict[str, Any], str]] = []
            
            for script in group_b:
                is_manual = False
                reason = ""
                
                # Check 1: Configuration says remediation is manual
                remediation_cfg = self.get_remediation_config_for_check(script['id'])
                if remediation_cfg.get("remediation") == "manual":
                    is_manual = True
                    reason = "Remediation marked as manual in configuration"
                
                # Check 2: Audit status is MANUAL
                if not is_manual and script['id'] in self.audit_results:
                    if self.audit_results[script['id']].get('status') == 'MANUAL':
                        is_manual = True
                        reason = "Audit phase identified this as manual intervention required"
                
                # Check 3: Script itself is marked as MANUAL
                if not is_manual and os.path.exists(script['path']):
                    is_manual = self._is_manual_check(script['path'])
                    if is_manual:
                        reason = "Script explicitly marked as MANUAL"
                
                if is_manual:
                    group_b_manual.append((script, reason))
                else:
                    group_b_automated.append(script)
            
            # Handle MANUAL checks / จัดการการตรวจสอบด้วยตนเอง
            if group_b_manual:
                print(f"\n{Colors.YELLOW}[*] GROUP B: Skipping {len(group_b_manual)} MANUAL checks...{Colors.ENDC}")
                for script, reason in group_b_manual:
                    print(f"{Colors.YELLOW}    [SKIP] {script['id']}: {reason}{Colors.ENDC}")
                    
                    self.manual_pending_items.append({
                        'id': script['id'],
                        'role': script.get('role', 'unknown'),
                        'level': script.get('level', 'unknown'),
                        'path': script.get('path', ''),
                        'reason': reason,
                        'component': self.get_component_for_rule(script['id'])
                    })
                    
                    self.log_activity("MANUAL_CHECK_SKIPPED", f"{script['id']}: {reason}")
            
            # Execute automated checks in parallel
            # ดำเนินการตรวจสอบอัตโนมัติแบบขนาน
            if group_b_automated:
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    futures = {executor.submit(self.run_script, s, "remediate"): s for s in group_b_automated}
                    
                    try:
                        completed = 0
                        for future in as_completed(futures):
                            if self.stop_requested:
                                break
                            
                            result_raw = future.result()
                            typed_result = self._coerce_cis_result(result_raw)
                            if typed_result is not None:
                                self.results.append(typed_result)
                                self.update_stats(typed_result)
                                completed += 1
                                
                                # Show progress for parallel execution
                                progress_pct = (completed / len(group_b_automated)) * 100
                                self._print_progress(typed_result, completed, len(group_b_automated), progress_pct)
                    
                    except KeyboardInterrupt:
                        self.stop_requested = True
                        print("\n[!] Aborted.")
            
            print(f"\n{Colors.GREEN}[+] GROUP B (Resources) Complete.{Colors.ENDC}")

    def _ensure_aggressive_remediation_confirmation(self) -> bool:
        if self._aggressive_remediation_confirmed:
            return True

        if os.environ.get("CIS_SKIP_AGGRESSIVE_CONFIRM", "").lower() == "true":
            self._aggressive_remediation_confirmed = True
            return True

        prompt = "[!] WARNING: This mode performs AGGRESSIVE container restarts (Hard Kill). The API Server will be briefly unavailable. Proceed with aggressive remediation for ALL components? (y/N) "
        try:
            choice = input(prompt)
        except (EOFError, KeyboardInterrupt):
            print(f"{Colors.YELLOW}[WARN] No response received; aborting aggressive remediation.{Colors.ENDC}")
            return False

        confirmed = choice.strip().lower() == "y"
        self._aggressive_remediation_confirmed = confirmed
        return confirmed

    def _init_stats(self):
        """Initialize statistics dictionary / เริ่มต้นพจนานุกรมสถิติ"""
        self.stats = {
            "master": {"pass": 0, "fail": 0, "manual": 0, "skipped": 0, "error": 0, "total": 0},
            "worker": {"pass": 0, "fail": 0, "manual": 0, "skipped": 0, "error": 0, "total": 0}
        }

    def _prepare_report_dir(self, mode: str) -> None:
        """
        Create report directory with Oracle-style structure
        สร้างไดเรกทอรีรายงานพร้อมโครงสร้างแบบ Oracle
        """
        # Coerce mode to string to handle callers passing non-str values without unnecessary isinstance checks
        mode = str(mode)
        run_folder = f"run_{self.timestamp}"
        self.current_report_dir = os.path.join(self.date_dir, mode, run_folder)
        os.makedirs(self.current_report_dir, exist_ok=True)

    def save_snapshot(self, mode: str, role: str, level: str) -> None:
        """
        Save results for trend comparison
        บันทึกผลลัพธ์เพื่อเปรียบเทียบแนวโน้ม
        """
        history_file = os.path.join(
            self.history_dir,
            f"snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{mode}_{role}_{level}.json"
        )
        
        snapshot = {
            "timestamp": datetime.now().isoformat(),
            "mode": mode,
            "role": role,
            "level": level,
            "stats": self.stats,
            "results": self.results
        }
        
        try:
            with open(history_file, 'w') as f:
                json.dump(snapshot, f, indent=2)
        except Exception as e:
            if self.verbose >= 2:
                print(f"{Colors.RED}[DEBUG] Snapshot error: {e}{Colors.ENDC}")

    def get_previous_snapshot(self, mode: str, role: str, level: str) -> Optional[Dict[str, Any]]:
        """
        Retrieve most recent previous snapshot.
        ดึงข้อมูลสแนปชอตก่อนหน้าล่าสุด

        Returns:
            Optional[Dict[str, Any]]: Loaded snapshot dict, or None if not found or on error.
        """
        pattern = f"snapshot_*_{mode}_{role}_{level}.json"
        snapshots = sorted(
            glob.glob(os.path.join(self.history_dir, pattern)),
            reverse=True
        )
        
        if snapshots:
            try:
                with open(snapshots[0], 'r') as f:
                    return json.load(f)
            except Exception as e:
                if self.verbose >= 1:
                    print(f"{Colors.YELLOW}[!] Snapshot load error: {e}{Colors.ENDC}")
        
        return None

    def calculate_score(self, stats: Dict[str, Dict[str, int]]) -> float:
        """
        Calculate dual-metric compliance scores (LEGACY - deprecated)
        
        This function is maintained for backwards compatibility.
        Use calculate_compliance_scores() instead for new code.
        
        Returns:
            float: Automation Health percentage (Pass / (Pass + Fail))
            
        คำนวณคะแนนการปฏิบัติตามข้อกำหนดแบบเมตริกคู่ (ดั้งเดิม - เลิกใช้งานแล้ว)
        """
        auto_health = self.calculate_compliance_scores(stats)
        # Ensure a stable float return value for static analysis and callers.
        try:
            # Use duck-typing: prefer a 'get' method if available rather than isinstance checks
            getter = getattr(auto_health, 'get', None)
            if callable(getter):
                val = getter('automation_health', 0.0)
            else:
                val = 0.0

            # Coerce to a safe numeric type for float() to avoid type-checker errors.
            if isinstance(val, (int, float)):
                return float(val)
            try:
                # Convert unknown objects to string then to float as a safe fallback.
                return float(str(val))
            except Exception:
                return 0.0
        except Exception:
            return 0.0

    def calculate_compliance_scores(self, stats: Dict[str, Dict[str, int]]) -> Dict[str, Any]:
        """
        Calculate dual compliance metrics for accurate CIS hardening assessment
        
        Two separate metrics address different concerns:
        
        1. AUTOMATION HEALTH (Technical Implementation)
           Score: Pass / (Pass + Fail)
           Purpose: Measures remediation script effectiveness
           Ignores: Manual checks (not script-automated)
           Use: Identify broken automation / script fixes needed
           
        2. AUDIT READINESS (Overall CIS Compliance)
           Score: Pass / Total Checks
           Purpose: Shows true CIS compliance readiness
           Includes: All check types (manual counts as non-passing)
           Use: Formal audit preparation and compliance reporting
        
        Args:
            stats: Dict with structure {role: {pass, fail, manual, skipped, error, total}}
        
        Returns:
            Dict with keys:
                - 'automation_health': float (0-100) - Pass/(Pass+Fail)
                - 'audit_readiness': float (0-100) - Pass/Total
                - 'pass': int - Total passing checks
                - 'fail': int - Total failing checks
                - 'manual': int - Total manual checks
                - 'skipped': int - Total skipped checks
                - 'error': int - Total error checks
                - 'total': int - Total checks processed
                
        คำนวณเมตริกการปฏิบัติตามข้อกำหนดคู่เพื่อการประเมินการทำให้ Kubernetes แข็งแกร่งตาม CIS ที่แม่นยำ
        """
        # Aggregate all roles (master, worker, etc.)
        # รวมทุกบทบาท (master, worker ฯลฯ)
        total_pass = stats["master"]["pass"] + stats["worker"]["pass"]
        total_fail = stats["master"]["fail"] + stats["worker"]["fail"]
        total_manual = stats["master"]["manual"] + stats["worker"]["manual"]
        total_skipped = stats["master"]["skipped"] + stats["worker"]["skipped"]
        total_error = stats["master"]["error"] + stats["worker"]["error"]
        total_all = stats["master"]["total"] + stats["worker"]["total"]
        
        # METRIC 1: Automation Health = Pass / (Pass + Fail)
        # Measures: How well are remediation scripts working?
        # Ignores: Manual checks (they're not automated)
        # เมตริก 1: สุขภาพของระบบอัตโนมัติ = ผ่าน / (ผ่าน + ล้มเหลว)
        automated_checks = total_pass + total_fail
        if automated_checks > 0:
            automation_health = round((total_pass / automated_checks) * 100, 2)
        else:
            automation_health = 0.0
        
        # METRIC 2: Audit Readiness = Pass / Total Checks
        # Measures: What's our true CIS compliance status?
        # Includes: All checks (manual = non-passing)
        # เมตริก 2: ความพร้อมในการตรวจสอบ = ผ่าน / การตรวจสอบทั้งหมด
        if total_all > 0:
            audit_readiness = round((total_pass / total_all) * 100, 2)
        else:
            audit_readiness = 0.0
        
        return {
            'automation_health': automation_health,
            'audit_readiness': audit_readiness,
            'pass': total_pass,
            'fail': total_fail,
            'manual': total_manual,
            'skipped': total_skipped,
            'error': total_error,
            'total': total_all
        }

    def show_trend_analysis(self, current_score: float, previous_snapshot: Dict[str, Any]) -> None:
        """
        Display score trend comparison
        แสดงการเปรียบเทียบแนวโน้มคะแนน
        """
        # Skip if previous snapshot is missing or empty (avoids redundant isinstance on a typed dict)
        if not previous_snapshot:
            return

        previous_stats = previous_snapshot.get("stats", {})
        previous_score = self.calculate_score(previous_stats)
        # Ensure numeric arithmetic for trend calculation
        try:
            trend = float(current_score) - float(previous_score)
        except Exception:
            trend = 0.0

        trend_symbol = "📈" if trend > 0 else ("📉" if trend < 0 else "➡️")
        trend_color = Colors.GREEN if trend > 0 else (Colors.RED if trend < 0 else Colors.YELLOW)

        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"TREND ANALYSIS (Score Comparison)")
        print(f"{'='*70}{Colors.ENDC}")
        print(f"  Current Score:   {Colors.BOLD}{current_score}%{Colors.ENDC}")
        print(f"  Previous Score:  {Colors.BOLD}{previous_score}%{Colors.ENDC}")
        print(f"  Change:          {trend_color}{trend_symbol} {'+' if trend > 0 else ''}{trend:.2f}%{Colors.ENDC}")
        print(f"{Colors.CYAN}{'='*70}{Colors.ENDC}")

    def save_reports(self, mode: str) -> None:
        """
        Save all report formats (CSV, JSON, Text, HTML)
        บันทึกรูปแบบรายงานทั้งหมด (CSV, JSON, Text, HTML)
        """
        if not self.current_report_dir:
            raise ValueError("Report directory not set")
        
        # CSV report / รายงาน CSV
        self._save_csv_report()
        
        # Text reports / รายงานข้อความ
        self._save_text_reports(mode)
        
        # JSON report / รายงาน JSON
        self._save_json_report()
        
        # HTML report / รายงาน HTML
        self._save_html_report(mode)
        
        print(f"\n   [*] Reports saved to: {self.current_report_dir}")

    def _save_csv_report(self):
        """
        Save CSV report with structured data
        บันทึกรายงาน CSV ด้วยข้อมูลที่มีโครงสร้าง
        """
        if not self.current_report_dir:
            raise ValueError("Report directory not set")
        csv_file = os.path.join(self.current_report_dir, "report.csv")
        fieldnames = ["id", "role", "level", "status", "duration", "reason", "fix_hint", "component"]
        
        try:
            with open(csv_file, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(self.results)
        except Exception as e:
            if self.verbose >= 2:
                print(f"{Colors.RED}[DEBUG] CSV save error: {e}{Colors.ENDC}")

    def export_results_to_excel(self, filename: Optional[str] = None) -> None:
        """
        Export results to Excel format with formatting and color coding.
        ส่งออกผลลัพธ์เป็นรูปแบบ Excel พร้อมการจัดรูปแบบและรหัสสี
        """
        if not openpyxl_available:
            print(f"{Colors.RED}[!] openpyxl is not installed. Falling back to CSV.{Colors.ENDC}")
            self._save_csv_report()
            return

        if openpyxl is None:
            print(f"{Colors.RED}[!] openpyxl failed to import. Falling back to CSV.{Colors.ENDC}")
            self._save_csv_report()
            return

        if not self.results:
            print(f"{Colors.YELLOW}[!] No results to export. Run a scan first.{Colors.ENDC}")
            return

        if not filename:
            if self.current_report_dir:
                filename = os.path.join(self.current_report_dir, f"cis_report_{self.timestamp}.xlsx")
            else:
                filename = f"cis_report_{self.timestamp}.xlsx"

        try:
            from openpyxl.styles import Font as ExcelFont, PatternFill as ExcelPatternFill, Alignment as ExcelAlignment
        except ImportError as exc:
            print(f"{Colors.RED}[!] Unable to import openpyxl styling helpers ({exc}). Falling back to CSV.{Colors.ENDC}")
            self._save_csv_report()
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("CIS Benchmark Report")
            ws.title = "CIS Benchmark Report"

            # Define columns
            headers = ["Check ID", "Node Type", "Level", "Status", "Duration (s)", "Remediation Result / Reason", "Component"]
            ws.append(headers)

            # Formatting: Bold Header
            for cell in ws[1]:
                cell.font = ExcelFont(bold=True)
                cell.alignment = ExcelAlignment(horizontal="center")
                cell.fill = ExcelPatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add data
            for res in self.results:
                row = [
                    res.get("id", ""),
                    res.get("role", "").capitalize(),
                    res.get("level", ""),
                    res.get("status", ""),
                    res.get("duration", 0),
                    res.get("reason", ""),
                    res.get("component", "")
                ]
                ws.append(row)
                
                # Color code the Status cell
                status = res.get("status", "")
                status_cell = ws.cell(row=ws.max_row, column=4)
                
                if status in ["PASS", "FIXED"]:
                    status_cell.fill = ExcelPatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
                elif status == "FAIL":
                    status_cell.fill = ExcelPatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
                elif status == "MANUAL":
                    status_cell.fill = ExcelPatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Orange

            # Auto-adjust column width
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = _excel_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            print(f"{Colors.GREEN}[+] Excel report exported successfully: {filename}{Colors.ENDC}")
            self.log_activity("EXPORT_EXCEL", f"File: {filename}")
            
        except Exception as e:
            print(f"{Colors.RED}[!] Excel export error: {e}{Colors.ENDC}")
            self.log_activity("EXPORT_EXCEL_ERROR", str(e))

    def _save_text_reports(self, mode: str):
        """Save summary and detailed text reports / บันทึกรายงานข้อความสรุปและรายละเอียด"""
        if not self.current_report_dir:
            raise ValueError("Report directory not set")
        summary_file = os.path.join(self.current_report_dir, "summary.txt")
        failed_file = os.path.join(self.current_report_dir, "failed_items.txt")
        summary_file = os.path.join(self.current_report_dir, "summary.txt")
        failed_file = os.path.join(self.current_report_dir, "failed_items.txt")
        
        # Categorize results / จำแนกผลลัพธ์
        passed: List[CISResult] = [r for r in self.results if r['status'] in ['PASS', 'FIXED']]
        failed: List[CISResult] = [r for r in self.results if r['status'] in ['FAIL', 'ERROR']]
        manual: List[CISResult] = [r for r in self.results if r['status'] == 'MANUAL']
        skipped: List[CISResult] = [r for r in self.results if r['status'] == 'SKIPPED']
        
        # Summary / สรุป
        with open(summary_file, 'w') as f:
            considered = len(passed) + len(failed) + len(manual)
            score = (len(passed) / considered * 100) if considered > 0 else 0.0
            
            f.write(f"{'='*60}\n")
            f.write(f"CIS BENCHMARK SUMMARY - {mode.upper()}\n")
            f.write(f"Date: {datetime.now()}\n")
            f.write(f"{'='*60}\n\n")
            f.write(f"Total:    {len(self.results)}\n")
            f.write(f"Pass:     {len(passed)}\n")
            f.write(f"Fail:     {len(failed)}\n")
            f.write(f"Manual:   {len(manual)}\n")
            f.write(f"Skipped:  {len(skipped)}\n")
            f.write(f"Score:    {score:.2f}%\n")
        
        # Failed/Manual/Skipped items / รายการที่ล้มเหลว/ด้วยตนเอง/ข้ามไป
        with open(failed_file, 'w') as f:
            f.write(f"DETAILED RESULTS ({mode.upper()})\n")
            f.write(f"{'='*60}\n\n")
            
            for item in failed + manual + skipped:
                title = self.extract_metadata_from_script(item.get('path'))
                f.write(f"CIS ID: {item['id']} [{item['status']}]\n")
                f.write(f"Title:  {title}\n")
                f.write(f"Role:   {item['role'].upper()}\n")
                if item.get('reason'):
                    f.write(f"Reason: {item['reason']}\n")
    def _save_json_report(self):
        """Save JSON report with full details / บันทึกรายงาน JSON ด้วยรายละเอียดเต็มรูปแบบ"""
        if not self.current_report_dir:
            raise ValueError("Report directory not set")
        json_file = os.path.join(self.current_report_dir, "report.json")
        
        try:
            with open(json_file, 'w') as f:
                json.dump({"stats": self.stats, "results": self.results}, f, indent=2)
        except Exception as e:
            if self.verbose >= 2:
                print(f"{Colors.RED}[DEBUG] JSON save error: {e}{Colors.ENDC}")

    def _save_html_report(self, mode: str) -> None:
        """Save HTML report with visualization / บันทึกรายงาน HTML พร้อมการแสดงภาพ"""
        if not self.current_report_dir:
            raise ValueError("Report directory not set")
        html_file = os.path.join(self.current_report_dir, "report.html")
        
        passed = len([r for r in self.results if r['status'] in ['PASS', 'FIXED']])
        failed = len([r for r in self.results if r['status'] in ['FAIL', 'ERROR']])
        manual = len([r for r in self.results if r['status'] == 'MANUAL'])
        score = (passed / (passed + failed + manual) * 100) if (passed + failed + manual) > 0 else 0
        
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>CIS Benchmark Report</title>
    <style>
        body {{ font-family: Arial; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; }}
        h1 {{ color: #333; text-align: center; }}
        .stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin: 20px 0; }}
        .stat-box {{ padding: 15px; text-align: center; border-radius: 4px; }}
        .pass {{ background: #d4edda; color: #155724; }}
        .fail {{ background: #f8d7da; color: #721c24; }}
        .manual {{ background: #fff3cd; color: #856404; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #007bff; color: white; padding: 10px; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>CIS Kubernetes Benchmark Report</h1>
        <p><strong>Mode:</strong> {mode.upper()} | <strong>Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="stats">
            <div class="stat-box pass"><div style="font-size: 24px; font-weight: bold;">{passed}</div>Passed</div>
            <div class="stat-box fail"><div style="font-size: 24px; font-weight: bold;">{failed}</div>Failed</div>
            <div class="stat-box manual"><div style="font-size: 24px; font-weight: bold;">{manual}</div>Manual</div>
            <div class="stat-box" style="background: #d1ecf1; color: #0c5460;"><div style="font-size: 24px; font-weight: bold;">{score:.1f}%</div>Score</div>
        </div>
        
        <table>
            <tr><th>CIS ID</th><th>Role</th><th>Status</th><th>Duration</th></tr>
            {''.join(f'<tr><td>{r["id"]}</td><td>{r["role"].upper()}</td><td>{r["status"]}</td><td>{r["duration"]}s</td></tr>' for r in self.results)}
        </table>
    </div>
</body>
</html>"""
        
        with open(html_file, 'w') as f:
            f.write(html_content)

    def print_stats_summary(self):
        """
        Display comprehensive compliance status with focused scoring.
        
        Metrics:
        1. AUTOMATION HEALTH: Measures remediation script reliability (excludes manual checks).
        2. SELECTED LEVEL COMPLIANCE: Pass rate for the targeted CIS level scope.
        3. OVERALL HARDENING PROGRESS: How much of the combined Level 1 + Level 2 suite has passed.
        
        The per-role breakdown that follows highlights PASS/FAIL/MANUAL counts for transparency, and
        MANUAL CHECKS remain separate so they do not skew automation scores.
        """
        # Calculate compliance scores
        scores = self.calculate_compliance_scores(self.stats)
        
        # Determine role being assessed
        master_total = self.stats["master"]["total"]
        worker_total = self.stats["worker"]["total"]
        if master_total > 0 and worker_total == 0:
            role_name = "MASTER NODE"
        elif worker_total > 0 and master_total == 0:
            role_name = "WORKER NODE"
        else:
            role_name = "CLUSTER"
        
        # Color helpers
        def get_score_color(score: float) -> str:
            # Coerce to float for defensive typing and static analysis
            try:
                s = float(score)
            except Exception:
                s = 0.0
            if s >= 80:
                return Colors.GREEN
            elif s >= 50:
                return Colors.YELLOW
            return Colors.RED
        
        def get_score_status(score: float) -> str:
            # Coerce to float for defensive typing and static analysis
            try:
                s = float(score)
            except Exception:
                s = 0.0
            if s >= 90:
                return "Excellent"
            if s >= 80:
                return "Good"
            if s >= 70:
                return "Acceptable"
            if s >= 50:
                return "Needs Improvement"
            return "Critical"
        
        # Compliance header
        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"COMPLIANCE STATUS: {role_name}")
        print(f"{'='*70}{Colors.ENDC}")
        
        # 1. AUTOMATION HEALTH
        auto_health = scores['automation_health']
        auto_color = get_score_color(auto_health)
        auto_status = get_score_status(auto_health)
        print(f"\n{Colors.BOLD}1. AUTOMATION HEALTH (Technical Implementation){Colors.ENDC}")
        print(f"   [Pass / (Pass + Fail)] - EXCLUDES Manual checks")
        print(f"   - Score: {auto_color}{auto_health:.2f}%{Colors.ENDC}")
        print(f"   - Status: {auto_color}{auto_status}{Colors.ENDC}")
        print(f"   - Meaning: How well remediation scripts are working (excluding manual checks)")
        
        # 2. SELECTED LEVEL COMPLIANCE
        level_label = "1+2" if self.current_level == "all" else self.current_level
        audit_ready = scores['audit_readiness']
        audit_color = get_score_color(audit_ready)
        audit_status = get_score_status(audit_ready)
        print(f"\n{Colors.BOLD}2. SELECTED LEVEL COMPLIANCE (Level {level_label}){Colors.ENDC}")
        print(f"   [Pass / Total Checks for selected scope]")
        print(f"   - Score: {audit_color}{audit_ready:.2f}%{Colors.ENDC}")
        print(f"   - Status: {audit_color}{audit_status}{Colors.ENDC}")
        print(f"   - Meaning: Compliance score restricted to the level you chose")
        if self.current_level == "1":
            print(f"\n{Colors.YELLOW}[!] REALITY CHECK:{Colors.ENDC}")
            print("    This score reflects BASELINE security (Level 1) only.")
            print("    Advanced threats (Level 2) are not yet mitigated.")
            print("    Recommended Next Step: Run Audit/Remediation for Level 2.")
        
        # 3. OVERALL HARDENING PROGRESS
        overall_total = self.total_level_one_two_checks
        if overall_total > 0:
            progress_score = round((scores['pass'] / overall_total) * 100, 2)
        else:
            progress_score = 0.0
        progress_color = get_score_color(progress_score)
        progress_status = get_score_status(progress_score)
        print(f"\n{Colors.BOLD}3. OVERALL HARDENING PROGRESS (Level 1+2){Colors.ENDC}")
        print("   [Passed checks this run / Total Level 1+2 configured checks]")
        print(f"   - Score: {progress_color}{progress_score:.2f}%{Colors.ENDC}")
        print(f"   - Status: {progress_color}{progress_status}{Colors.ENDC}")
        if overall_total > 0:
            print(f"   - Meaning: {scores['pass']} of {overall_total} Level 1+2 checks passed in this run.")
        else:
            print("   - Meaning: Total Level 1+2 check count unavailable (config load issue).")
        
        # 4. AUTOMATED FAILURES (Critical Issues Only)
        automated_failures = scores['fail']
        print(f"\n{Colors.BOLD}4. AUTOMATED FAILURES (❌ Need Script Fixes){Colors.ENDC}")
        if automated_failures > 0:
            print(f"   {Colors.RED}⚠ {automated_failures} automated checks FAILED{Colors.ENDC}")
            print(f"   Action: Debug and fix remediation scripts")
            failed_checks = [r for r in self.results if r['status'] == 'FAIL']
            if failed_checks and self.verbose >= 1:
                print(f"   Failed checks:")
                for check in failed_checks[:10]:
                    print(f"     • {check['id']}")
                if len(failed_checks) > 10:
                    print(f"     ... and {len(failed_checks) - 10} more")
        else:
            print(f"   {Colors.GREEN}✓ All automated checks working{Colors.ENDC}")
        print(f"\n{Colors.CYAN}{'='*70}{Colors.ENDC}")
        
        # Per-role breakdown (detailed for transparency)
        # รายละเอียดแยกตามบทบาท (เพื่อความโปร่งใส)
        print(f"\n{Colors.CYAN}DETAILED BREAKDOWN BY ROLE{Colors.ENDC}")
        print(f"{Colors.CYAN}{'='*70}{Colors.ENDC}")
        
        for role in ["master", "worker"]:
            s = self.stats[role]
            if s['total'] == 0:
                continue
            
            # Calculate per-role automation health (excluding manual)
            # คำนวณสุขภาพของระบบอัตโนมัติตามบทบาท (ไม่รวมแมนนวล)
            role_automated = s['pass'] + s['fail']
            role_auto_health = round((s['pass'] / role_automated) * 100, 2) if role_automated > 0 else 0.0
            
            # Calculate per-role audit readiness
            # คำนวณความพร้อมในการตรวจสอบตามบทบาท
            role_audit = round((s['pass'] / s['total']) * 100, 2) if s['total'] > 0 else 0.0
            
            print(f"\n  {Colors.BOLD}{role.upper()}:{Colors.ENDC}")
            print(f"    {Colors.GREEN}Pass{Colors.ENDC}:      {Colors.GREEN}{s['pass']:3d}{Colors.ENDC}")
            print(f"    {Colors.RED}Fail{Colors.ENDC}:      {Colors.RED}{s['fail']:3d}{Colors.ENDC}")
            print(f"    {Colors.YELLOW}Manual{Colors.ENDC}:    {Colors.YELLOW}{s['manual']:3d}{Colors.ENDC} (Requires human review)")
            print(f"    {Colors.CYAN}Skipped{Colors.ENDC}:   {Colors.CYAN}{s['skipped']:3d}{Colors.ENDC}")
            
            if s.get('error', 0) > 0:
                print(f"    {Colors.RED}Error{Colors.ENDC}:     {Colors.RED}{s['error']:3d}{Colors.ENDC}")
            
            print(f"    {Colors.BOLD}Total{Colors.ENDC}:     {Colors.BOLD}{s['total']:3d}{Colors.ENDC}")
            
            auto_color = get_score_color(role_auto_health)
            audit_color = get_score_color(role_audit)
            print(f"    {Colors.BOLD}Auto Health{Colors.ENDC}:  {auto_color}{role_auto_health:.2f}%{Colors.ENDC} (of automated checks)")
            print(f"    {Colors.BOLD}Audit Ready{Colors.ENDC}:  {audit_color}{role_audit:.2f}%{Colors.ENDC} (overall)")
        
        # ========== NEW SECTION: MANUAL PENDING ITEMS ==========
        # Show items that were skipped from automation
        # ========== ส่วนใหม่: รายการที่รอดำเนินการด้วยตนเอง ==========
        print(f"\n{Colors.CYAN}{'='*70}{Colors.ENDC}")
        print(f"\n{Colors.BOLD}📋 MANUAL INTERVENTION REQUIRED{Colors.ENDC}")
        print(f"{Colors.YELLOW}Items skipped from automation for human review:{Colors.ENDC}\n")
        
        if self.manual_pending_items:
            total_manual = len(self.manual_pending_items)
            print(f"  Total: {total_manual} checks require manual review\n")
            
            # Group by role for clarity
            # กลุ่มตามบทบาทเพื่อความชัดเจน
            manual_by_role: Dict[str, List[Dict[str, Any]]] = {}
            for item in self.manual_pending_items:
                role = str(item.get('role', 'unknown'))
                if role not in manual_by_role:
                    manual_by_role[role] = []
                manual_by_role[role].append(item)
            
            for role in ['master', 'worker']:
                if role in manual_by_role:
                    items: List[Dict[str, Any]] = manual_by_role[role]
                    print(f"  {Colors.BOLD}{role.upper()} NODE ({len(items)} items):{Colors.ENDC}")
                    for item in items:
                        check_id = item['id']
                        reason = item.get('reason', 'No details available')
                        component = item.get('component', '')
                        
                        # Format output nicely
                        component_str = f" [{component}]" if component else ""
                        print(f"    • {Colors.YELLOW}{check_id}{Colors.ENDC}{component_str}")
                        print(f"      └─ {Colors.CYAN}{reason}{Colors.ENDC}")
                    print()
            
            # Guidelines / แนวทางปฏิบัติ
            print(f"  {Colors.CYAN}Notes:{Colors.ENDC}")
            print(f"    • These checks are NOT failures or errors")
            print(f"    • They require human decisions that cannot be automated")
            print(f"    • They do NOT count against Automation Health score")
            print(f"    • They do NOT block remediation success\n")
            
            print(f"  {Colors.CYAN}Recommended Actions:{Colors.ENDC}")
            print(f"    1. Review each manual item and understand what it requires")
            print(f"    2. Determine if the check applies to your cluster architecture")
            print(f"    3. If applicable, implement the fix manually following CIS guidelines")
            print(f"    4. Re-run audit to verify the fix")
            print(f"    5. Document any decisions for compliance audit trail")
        else:
            print(f"  {Colors.GREEN}✓ No manual intervention items{Colors.ENDC}\n")
            print(f"  All checks were either automated or not required for your configuration.")
        
        print(f"\n{Colors.CYAN}{'='*70}{Colors.ENDC}")



    def show_verbose_result(self, res: CISResult) -> None:
        """
        Display detailed result output
        แสดงผลลัพธ์รายละเอียด
        """
        if self.verbose == 0:
            return
        
        title = self.extract_metadata_from_script(res.get('path'))
        if res['status'] in ['PASS', 'FIXED']:
            color = Colors.GREEN
        elif res['status'] == 'MANUAL':
            color = Colors.YELLOW
        else:
            color = Colors.RED
        
        print(f"\n{Colors.CYAN}{'='*70}")
        print(f"{color}[{res['status']}]{Colors.ENDC} {res['id']} - {title}")
        print(f"{'='*70}{Colors.ENDC}")
        
        if self.verbose >= 1:
            if res.get('reason'):
                print(f"  Reason: {res['reason']}")
            if res.get('fix_hint') and res['status'] == 'FAIL':
                print(f"  Fix:    {res['fix_hint']}")
        
        if self.verbose >= 2:
            print(f"\n  {Colors.YELLOW}[Output]{Colors.ENDC}")
            output = res.get('output', '').strip()
            print(output if output else "  (no output)")
        
        print()

    def show_menu(self):
        """
        Display main menu
        แสดงเมนูหลัก
        """
        print(f"\n{Colors.CYAN}{'='*70}")
        print("SELECT MODE")
        print(f"{'='*70}{Colors.ENDC}\n")
        print("  1) Audit only (non-destructive)")
        print("  2) Remediation only (DESTRUCTIVE - ALL checks)")
        print("  3) Remediation only (Fix FAILED items only)")
        print("  4) Both (Audit then Remediation)")
        print("  5) Health Check")
        print("  6) Help")
        print("  0) Exit\n")
        
        while True:
            choice = input(f"{Colors.BOLD}Choose [0-7]: {Colors.ENDC}").strip()
            if choice in ['0', '1', '2', '3', '4', '5', '6', '7']:
                return choice
            print(f"{Colors.RED}Invalid choice.{Colors.ENDC}")

    def get_audit_options(self):
        """
        Get user options for audit
        ได้รับตัวเลือกของผู้ใช้สำหรับการตรวจสอบ
        """
        print(f"\n{Colors.CYAN}AUDIT CONFIGURATION{Colors.ENDC}\n")
        
        # PRIORITY 1: Try to auto-detect node role
        # ลำดับความสำคัญ 1: พยายามตรวจจับบทบาทของโหนดโดยอัตโนมัติ
        detected_role = self.detect_node_role()
        if detected_role:
            print(f"{Colors.GREEN}[+] Auto-detected Node Role: {detected_role.upper()}{Colors.ENDC}")
            role = detected_role
        else:
            # Detection failed - show simplified menu (no 'Both' option)
            # การตรวจจับล้มเหลว - แสดงเมนูแบบย่อ
            print("  Select Target Role:")
            print("    1) Master")
            print("    2) Worker")
            # Validate role input - only accept 1 or 2
            while True:
                role_input = input("\n  Input [1-2]: ").strip()
                if role_input in ["1", "2"]:
                    role = {"1": "master", "2": "worker"}[role_input]
                    break
                print(f"  {Colors.RED}✗ Invalid choice. Please select 1 or 2.{Colors.ENDC}")
        
        # Level selection / การเลือกระดับ
        print(f"\n  CIS Level:")
        print("    1) Level 1")
        print("    2) Level 2")
        print("    3) All")
        # Validate level input - accept 1, 2, 3 or default to 3
        while True:
            level_input = input("\n  Select level [1-3] (default: 3): ").strip() or "3"
            if level_input in ["1", "2", "3"]:
                level = {"1": "1", "2": "2", "3": "all"}[level_input]
                break
            print(f"  {Colors.RED}✗ Invalid choice. Please select 1, 2, or 3.{Colors.ENDC}")
        
        return level, role, self.verbose, False, SCRIPT_TIMEOUT

    def get_remediation_options(self):
        """
        Get user options for remediation
        ได้รับตัวเลือกของผู้ใช้สำหรับการแก้ไข
        """
        print(f"\n{Colors.RED}[!] WARNING: REMEDIATION WILL MODIFY YOUR CLUSTER!{Colors.ENDC}\n")
        
        # PRIORITY 1: Try to auto-detect node role
        # ลำดับความสำคัญ 1: พยายามตรวจจับบทบาทของโหนดโดยอัตโนมัติ
        detected_role = self.detect_node_role()
        if detected_role:
            print(f"{Colors.GREEN}[+] Auto-detected Node Role: {detected_role.upper()}{Colors.ENDC}")
            role = detected_role
        else:
            # Detection failed - show simplified menu (no 'Both' option)
            # การตรวจจับล้มเหลว - แสดงเมนูแบบย่อ
            print("  Select Target Role:")
            print("    1) Master")
            print("    2) Worker")
            # Validate role input - only accept 1 or 2
            while True:
                role_input = input("\n  Input [1-2]: ").strip()
                if role_input in ["1", "2"]:
                    role = {"1": "master", "2": "worker"}[role_input]
                    break
                print(f"  {Colors.RED}✗ Invalid choice. Please select 1 or 2.{Colors.ENDC}")
        
        # Level selection / การเลือกระดับ
        print(f"\n  CIS Level:")
        print("    1) Level 1")
        print("    2) Level 2")
        print("    3) All")
        # Validate level input - accept 1, 2, 3 or default to 3
        while True:
            level_input = input("\n  Select level [1-3] (default: 3): ").strip() or "3"
            if level_input in ["1", "2", "3"]:
                level = {"1": "1", "2": "2", "3": "all"}[level_input]
                break
            print(f"  {Colors.RED}✗ Invalid choice. Please select 1, 2, or 3.{Colors.ENDC}")
        
        return level, role, SCRIPT_TIMEOUT

    def confirm_action(self, message: str) -> bool:
        """
        Ask for user confirmation
        ขอการยืนยันจากผู้ใช้
        """
        while True:
            response = input(f"\n{message} [y/n]: ").strip().lower()
            if response in ['y', 'yes']:
                return True
            elif response in ['n', 'no']:
                return False
            print(f"{Colors.RED}Please enter 'y' or 'n'.{Colors.ENDC}")

    def show_results_menu(self, mode: str) -> None:
        """
        Show menu after scan/fix completion
        แสดงเมนูหลังจากการสแกน/แก้ไขเสร็จสมบูรณ์
        """
        if not self.current_report_dir:
            print(f"{Colors.RED}[!] No report directory available.{Colors.ENDC}")
            return
        
        while True:
            print(f"\n{Colors.CYAN}{'='*70}")
            print("RESULTS MENU")
            print(f"{'='*70}{Colors.ENDC}\n")
            print("  1) View summary")
            print("  2) View failed items")
            print("  3) View HTML report")
            print("  4) Export to Excel")
            print("  5) Return to main menu\n")
            
            choice = input(f"{Colors.BOLD}Choose [1-5]: {Colors.ENDC}").strip()
            
            if choice == '1':
                summary_file = os.path.join(self.current_report_dir, "summary.txt")
                if os.path.exists(summary_file):
                    with open(summary_file, 'r') as f:
                        print(f"\n{f.read()}")
            elif choice == '2':
                failed_file = os.path.join(self.current_report_dir, "failed_items.txt")
                if os.path.exists(failed_file):
                    with open(failed_file, 'r') as f:
                        print(f"\n{f.read()}")
            elif choice == '3':
                html_file = os.path.join(self.current_report_dir, "report.html")
                print(f"   HTML report: {html_file}")
            elif choice == '4':
                self.export_results_to_excel()
            elif choice == '5':
                break

    def show_help(self):
        """Display help information / แสดงข้อมูลความช่วยเหลือ"""
        print(f"""
{Colors.CYAN}{'='*70}
CIS Kubernetes Benchmark - HELP
{'='*70}{Colors.ENDC}

{Colors.BOLD}[1] AUDIT{Colors.ENDC}
    Scan compliance checks (non-destructive)
    ตรวจสอบการปฏิบัติตามข้อกำหนด (ไม่ทำลาย)
    
    Output: Stores audit results for targeted remediation

{Colors.BOLD}[2] REMEDIATION (ALL){Colors.ENDC}
    Apply fixes to ALL items regardless of audit status (MODIFIES CLUSTER)
    ใช้การแก้ไขเพื่อแก้ไขรายการ ALL (แก้ไขคลัสเตอร์)
    
    Use Case: Fresh cluster remediation, drift detection, force full compliance

{Colors.BOLD}[3] REMEDIATION (FAILED ONLY){Colors.ENDC}
    Fix ONLY items that FAILED or returned MANUAL in the previous audit
    ใช้การแก้ไขเพื่อแก้ไขเฉพาะรายการที่ล้มเหลวหรือต้องการแทรกแซง
    
    Use Case: Efficient remediation after audit, fix only what failed
    Requires: Must run Audit first to capture failed items
    Performance: Significantly faster on large clusters with few failures

{Colors.BOLD}[4] BOTH{Colors.ENDC}
    Run audit first, then remediate ALL items
    รันการตรวจสอบก่อน จากนั้นทำการแก้ไขทั้งหมด

{Colors.BOLD}[5] HEALTH CHECK{Colors.ENDC}
    Check Kubernetes cluster status
    ตรวจสอบสถานะของคลัสเตอร์ Kubernetes

{Colors.BOLD}SMART WAIT FEATURE{Colors.ENDC}
    Intelligently skips health checks for safe operations:
    - SKIP: CIS 1.1.x (file permissions/ownership - no service impact)
    - CHECK: All others (config/service changes - requires verification)
    
    Result: 50% faster remediation on large checks

{Colors.CYAN}{'='*70}{Colors.ENDC}
""")

    def main_loop(self):
        """
        Main application loop
        ลูปแอปพลิเคชันหลัก
        """
        self.show_banner()
        self.check_health()
        
        while True:
            choice = self.show_menu()
            
            if choice == '1':  # Audit / การตรวจสอบ
                level, role, verbose, skip_manual, timeout = self.get_audit_options()
                self.verbose = verbose
                self.skip_manual = skip_manual
                self.script_timeout = timeout
                self.log_activity("AUDIT", f"Level:{level}, Role:{role}")
                self.scan(level, role)
                
            elif choice == '2':  # Remediation ALL (Force Run) / การแก้ไขทั้งหมด
                level, role, timeout = self.get_remediation_options()
                self.script_timeout = timeout
                
                if self.confirm_action("Confirm remediation (ALL checks)?"):
                    self.log_activity("FIX_ALL", f"Level:{level}, Role:{role}")
                    self.fix(level, role, fix_failed_only=False)
                    
            elif choice == '3':  # Remediation FAILED ONLY / การแก้ไขเฉพาะรายการที่ล้มเหลว
                level, role, timeout = self.get_remediation_options()
                self.script_timeout = timeout
                
                # AUTO-AUDIT: If no audit results, run silent audit first
                # การตรวจสอบอัตโนมัติ: หากไม่มีผลการตรวจสอบ ให้รันการตรวจสอบแบบเงียบก่อน
                if not self.audit_results:
                    print(f"{Colors.CYAN}[INFO] No previous audit found. Running auto-audit to identify failures...{Colors.ENDC}")
                    # Run audit silently with same level/role settings
                    self.scan(level, role, skip_menu=True)
                    print(f"\n{Colors.CYAN}[+] Auto-audit complete. Proceeding to remediation...{Colors.ENDC}")
                
                # Show summary of audit findings
                # แสดงสรุปผลการตรวจสอบ
                failed_count = sum(1 for r in self.audit_results.values() if r.get('status') in ['FAIL', 'ERROR', 'MANUAL'])
                passed_count = sum(1 for r in self.audit_results.values() if r.get('status') in ['PASS', 'FIXED'])
                
                print(f"\n{Colors.CYAN}{'='*70}")
                print("AUDIT SUMMARY")
                print(f"{'='*70}{Colors.ENDC}")
                print(f"  Total Audited:    {len(self.audit_results)}")
                print(f"  PASSED:           {Colors.GREEN}{passed_count}{Colors.ENDC}")
                print(f"  FAILED/MANUAL:    {Colors.RED}{failed_count}{Colors.ENDC}")
                print(f"{Colors.CYAN}{'='*70}{Colors.ENDC}\n")
                
                if failed_count == 0:
                    print(f"{Colors.GREEN}[+] All checks passed! No remediation needed.{Colors.ENDC}")
                    continue
                
                if self.confirm_action(f"Remediate {failed_count} failed/manual items?"):
                    self.log_activity("FIX_FAILED_ONLY", f"Level:{level}, Role:{role}, Failed:{failed_count}")
                    self.fix(level, role, fix_failed_only=True)
                    
            elif choice == '4':  # Both / ทั้งสอง
                level, role, verbose, skip_manual, timeout = self.get_audit_options()
                self.verbose = verbose
                self.script_timeout = timeout
                self.log_activity("AUDIT_THEN_FIX", f"Level:{level}, Role:{role}")
                self.scan(level, role, skip_menu=True)
                
                if self.confirm_action("Proceed to remediation (ALL checks)?"):
                    self.fix(level, role, fix_failed_only=False)
                    
            elif choice == '5':  # Health check / ตรวจสอบสุขภาพ
                self.log_activity("HEALTH_CHECK", "Initiated")
                self.check_health()
                
            elif choice == '6':  # Help / ความช่วยเหลือ
                self.show_help()
                
            elif choice == '7':  # Export to Excel / ส่งออกเป็น Excel
                self.export_results_to_excel()
                
            elif choice == '0':  # Exit / ออก
                self.log_activity("EXIT", "Application terminated")
                print(f"\n{Colors.CYAN}Goodbye!{Colors.ENDC}\n")
                sys.exit(0)


if __name__ == "__main__":
    # Entry point for the CIS Kubernetes Benchmark Runner
    # จุดเริ่มต้นสำหรับ CIS Kubernetes Benchmark Runner
    parser = argparse.ArgumentParser(
        description="CIS Kubernetes Benchmark Runner"
    )
    parser.add_argument(
        "-v", "--verbose", action="count", default=0,
        help="Verbosity level (-v: detailed, -vv: debug)"
    )
    parser.add_argument(
        "--config", default="cis_config.json",
        help="Path to configuration file"
    )
    parser.add_argument(
        "--mode", choices=["audit", "remediate", "both", "health"],
        help="Run in specific mode without menu"
    )
    parser.add_argument(
        "--level", choices=["1", "2", "all"], default="all",
        help="CIS Level to run"
    )
    parser.add_argument(
        "--role", choices=["master", "worker", "all"],
        help="Node role to target"
    )
    parser.add_argument(
        "--failed-only", action="store_true",
        help="Only remediate items that failed audit"
    )
    parser.add_argument(
        "--excel", action="store_true",
        help="Export results to Excel automatically"
    )
    
    args = parser.parse_args()
    
    runner = CISUnifiedRunner(verbose=args.verbose, config_path=args.config)
    
    try:
        if args.mode:
            # CLI Mode (Non-interactive)
            # โหมด CLI (ไม่โต้ตอบ)
            runner.run_preflight_checks()
            
            # Determine role if not provided
            # กำหนดบทบาทหากไม่ได้ระบุ
            target_role = args.role or runner.detect_node_role() or "all"
            
            if args.mode == "audit":
                runner.scan(args.level, target_role, skip_menu=True)
            elif args.mode == "remediate":
                runner.fix(args.level, target_role, fix_failed_only=args.failed_only)
            elif args.mode == "both":
                runner.scan(args.level, target_role, skip_menu=True)
                runner.fix(args.level, target_role, fix_failed_only=args.failed_only)
            elif args.mode == "health":
                runner.check_health()
            
            # Auto-export to Excel if requested
            # ส่งออกเป็น Excel โดยอัตโนมัติหากมีการร้องขอ
            if args.excel:
                runner.export_results_to_excel()
        else:
            # Interactive Menu Mode
            # โหมดเมนูแบบโต้ตอบ
            runner.main_loop()
            
    except KeyboardInterrupt:
        print("\n[!] Interrupted by user. Exiting...")
        sys.exit(1)
    except Exception as e:
        print(f"\n{Colors.RED}[!] CRITICAL ERROR: {e}{Colors.ENDC}")
        if args.verbose >= 2:
            import traceback
            traceback.print_exc()
        sys.exit(1)
